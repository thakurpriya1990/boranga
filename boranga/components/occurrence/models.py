import hashlib
import json
import logging
import mimetypes
import os
import random
import string
import traceback
import zipfile
import zoneinfo
from abc import abstractmethod
from datetime import datetime
from datetime import timezone as dt_timezone
from decimal import Decimal
from io import BytesIO

import openpyxl
import pyproj
import reversion
from colorfield.fields import ColorField
from django.apps import apps
from django.conf import settings
from django.contrib.contenttypes import fields
from django.contrib.contenttypes import models as ct_models
from django.contrib.gis.db import models as gis_models
from django.contrib.gis.db.models.functions import Area
from django.contrib.gis.geos import GEOSGeometry, Polygon
from django.core.cache import cache
from django.core.exceptions import FieldDoesNotExist, FieldError, ValidationError
from django.core.files.storage import FileSystemStorage
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.validators import MaxValueValidator, MinValueValidator
from django.db import IntegrityError, models, transaction
from django.db.models import (
    CharField,
    Count,
    Func,
    ManyToManyField,
    Max,
    OuterRef,
    Q,
    Subquery,
)
from django.db.models.functions import Cast, Length
from django.utils import timezone
from django.utils.functional import cached_property
from ledger_api_client.ledger_models import EmailUserRO as EmailUser
from ledger_api_client.managed_models import SystemGroup
from multiselectfield import MultiSelectField
from openpyxl.styles import NamedStyle
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from ordered_model.models import OrderedModel, OrderedModelManager
from rest_framework import serializers
from taggit.managers import TaggableManager

from boranga import exceptions
from boranga.components.conservation_status.models import ProposalAmendmentReason
from boranga.components.main.models import (
    AbstractOrderedList,
    ArchivableModel,
    CommunicationsLogEntry,
    Document,
    OrderedArchivableManager,
    RevisionedMixin,
    UserAction,
)
from boranga.components.main.related_item import RelatedItem
from boranga.components.occurrence.email import (
    send_approve_email_notification,
    send_approver_approve_email_notification,
    send_approver_back_to_assessor_email_notification,
    send_approver_decline_email_notification,
    send_decline_email_notification,
    send_occurrence_report_amendment_email_notification,
    send_occurrence_report_referral_complete_email_notification,
    send_occurrence_report_referral_email_notification,
    send_occurrence_report_referral_recall_email_notification,
)
from boranga.components.species_and_communities.models import (
    Community,
    CurrentImpact,
    District,
    DocumentCategory,
    DocumentSubCategory,
    GroupType,
    PotentialImpact,
    PotentialThreatOnset,
    Region,
    Species,
    Taxonomy,
    ThreatAgent,
    ThreatCategory,
)
from boranga.components.users.models import (
    SubmitterInformation,
    SubmitterInformationModelMixin,
)
from boranga.helpers import (
    belongs_to_by_user_id,
    clone_model,
    get_choices_for_field,
    get_display_field_for_model,
    get_mock_request,
    get_openpyxl_data_validation_type_for_django_field,
    is_django_admin,
    is_occurrence_approver,
    is_occurrence_assessor,
    member_ids,
    no_commas_validator,
)
from boranga.ledger_api_utils import retrieve_email_user
from boranga.settings import (
    GROUP_NAME_OCCURRENCE_APPROVER,
    GROUP_NAME_OCCURRENCE_ASSESSOR,
)

logger = logging.getLogger(__name__)

private_storage = FileSystemStorage(
    location=settings.BASE_DIR + "/private-media/", base_url="/private-media/"
)


def update_occurrence_report_comms_log_filename(instance, filename):
    return (
        f"{settings.MEDIA_APP_DIR}/occurrence_report/"
        f"{instance.log_entry.occurrence_report.id}/communications/{filename}"
    )


def update_occurrence_report_doc_filename(instance, filename):
    return f"{settings.MEDIA_APP_DIR}/occurrence_report/{instance.occurrence_report.id}/documents/{filename}"


def update_occurrence_doc_filename(instance, filename):
    return f"{settings.MEDIA_APP_DIR}/occurrence/{instance.occurrence.id}/documents/{filename}"


class OccurrenceReportManager(models.Manager):
    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("group_type", "species", "community")
            .annotate(
                observer_count=Count(
                    "observer_detail", filter=Q(observer_detail__visible=True)
                )
            )
        )


class OccurrenceReport(SubmitterInformationModelMixin, RevisionedMixin):
    """
    Occurrence Report for any particular species or community

    Used by:
    - Occurrence
    """

    objects = OccurrenceReportManager()

    BULK_IMPORT_ABBREVIATION = "ocr"
    BULK_IMPORT_EXCLUDE_FIELDS = ["occurrence_report_number", "import_hash"]

    CUSTOMER_STATUS_DRAFT = "draft"
    CUSTOMER_STATUS_WITH_ASSESSOR = "with_assessor"
    CUSTOMER_STATUS_WITH_APPROVER = "with_approver"
    CUSTOMER_STATUS_APPROVED = "approved"
    CUSTOMER_STATUS_DECLINED = "declined"
    CUSTOMER_STATUS_DISCARDED = "discarded"
    CUSTOMER_STATUS_CLOSED = "closed"
    CUSTOMER_STATUS_CHOICES = (
        (CUSTOMER_STATUS_DRAFT, "Draft"),
        (CUSTOMER_STATUS_WITH_ASSESSOR, "Under Review"),
        (CUSTOMER_STATUS_WITH_APPROVER, "Under Review"),
        (CUSTOMER_STATUS_APPROVED, "Approved"),
        (CUSTOMER_STATUS_DECLINED, "Declined"),
        (CUSTOMER_STATUS_DISCARDED, "Discarded"),
        (CUSTOMER_STATUS_CLOSED, "DeListed"),
    )

    # List of statuses from above that allow a customer to edit an occurrence report.
    CUSTOMER_EDITABLE_STATE = [
        "draft",
        "discarded",
        "amendment_required",
    ]

    # List of statuses from above that allow a customer to view an occurrence report (read-only)
    CUSTOMER_VIEWABLE_STATE = [
        "with_assessor",
        "with_approver",
        "under_review",
        "approved",
        "declined",
        "closed",
    ]

    PROCESSING_STATUS_DRAFT = "draft"
    PROCESSING_STATUS_WITH_ASSESSOR = "with_assessor"
    PROCESSING_STATUS_WITH_REFERRAL = "with_referral"
    PROCESSING_STATUS_WITH_APPROVER = "with_approver"
    PROCESSING_STATUS_APPROVED = "approved"
    PROCESSING_STATUS_DECLINED = "declined"
    PROCESSING_STATUS_UNLOCKED = "unlocked"
    PROCESSING_STATUS_DISCARDED = "discarded"
    PROCESSING_STATUS_CLOSED = "closed"
    PROCESSING_STATUS_CHOICES = (
        (PROCESSING_STATUS_DRAFT, "Draft"),
        (PROCESSING_STATUS_WITH_ASSESSOR, "With Assessor"),
        (PROCESSING_STATUS_WITH_REFERRAL, "With Referral"),
        (PROCESSING_STATUS_WITH_APPROVER, "With Approver"),
        (PROCESSING_STATUS_APPROVED, "Approved"),
        (PROCESSING_STATUS_DECLINED, "Declined"),
        (PROCESSING_STATUS_UNLOCKED, "Unlocked"),
        (PROCESSING_STATUS_DISCARDED, "Discarded"),
        (PROCESSING_STATUS_CLOSED, "DeListed"),
    )

    VALID_BULK_IMPORT_PROCESSING_STATUSES = [
        (PROCESSING_STATUS_WITH_ASSESSOR, "With Assessor"),
        (PROCESSING_STATUS_WITH_APPROVER, "With Approver"),
        (PROCESSING_STATUS_APPROVED, "Approved"),
    ]

    FINALISED_STATUSES = [
        PROCESSING_STATUS_APPROVED,
        PROCESSING_STATUS_DECLINED,
        PROCESSING_STATUS_DISCARDED,
        PROCESSING_STATUS_CLOSED,
    ]

    customer_status = models.CharField(
        "Customer Status",
        max_length=40,
        choices=CUSTOMER_STATUS_CHOICES,
        default=CUSTOMER_STATUS_CHOICES[0][0],
    )

    APPLICATION_TYPE_CHOICES = (
        ("new_proposal", "New Application"),
        ("amendment", "Amendment"),
        ("renewal", "Renewal"),
        ("external", "External"),
    )

    RELATED_ITEM_CHOICES = [
        ("species", "Species"),
        ("community", "Community"),
        ("agendaitem", "Meeting Agenda Item"),
    ]

    # group_type of report
    group_type = models.ForeignKey(
        GroupType, on_delete=models.SET_NULL, blank=True, null=True
    )
    #
    proposal_type = models.CharField(
        "Application Status Type",
        max_length=40,
        choices=APPLICATION_TYPE_CHOICES,
        default=APPLICATION_TYPE_CHOICES[0][0],
    )

    # species related occurrence
    species = models.ForeignKey(
        Species,
        on_delete=models.CASCADE,
        related_name="occurrence_report",
        null=True,
        blank=True,
    )

    # communties related occurrence
    community = models.ForeignKey(
        Community,
        on_delete=models.CASCADE,
        related_name="occurrence_report",
        null=True,
        blank=True,
    )

    occurrence = models.ForeignKey(
        "Occurrence",
        on_delete=models.PROTECT,
        related_name="occurrence_reports",
        null=True,
        blank=True,
    )

    occurrence_report_number = models.CharField(max_length=9, blank=True, default="")

    # Field to use when importing data from the legacy system
    migrated_from_id = models.CharField(
        max_length=50, blank=True, null=True, unique=True
    )

    observation_date = models.DateTimeField(null=True, blank=True)
    reported_date = models.DateTimeField(auto_now_add=True, null=False, blank=False)
    submitter_information = models.OneToOneField(
        SubmitterInformation,
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
        related_name="occurrence_report",
    )
    submitter = models.IntegerField(null=True)  # EmailUserRO
    lodgement_date = models.DateTimeField(blank=True, null=True)

    assigned_officer = models.IntegerField(null=True)  # EmailUserRO
    assigned_approver = models.IntegerField(null=True)  # EmailUserRO
    approved_by = models.IntegerField(null=True)  # EmailUserRO
    # internal user who edits the approved conservation status(only specific fields)
    # modified_by = models.IntegerField(null=True) #EmailUserRO
    processing_status = models.CharField(
        "Processing Status",
        max_length=30,
        choices=PROCESSING_STATUS_CHOICES,
        default=PROCESSING_STATUS_CHOICES[0][0],
    )

    proposed_decline_status = models.BooleanField(default=False)
    deficiency_data = models.TextField(null=True, blank=True)  # deficiency comment
    assessor_data = models.TextField(null=True, blank=True)  # assessor comment
    approver_comment = models.TextField(blank=True)
    internal_application = models.BooleanField(default=False)
    site = models.TextField(null=True, blank=True)
    comments = models.TextField(null=True, blank=True)
    # Allows the OCR submitter to hint the assessor to which occurrence to assign to
    # without forcefully linking the occurrence to the OCR
    ocr_for_occ_number = models.CharField(max_length=9, blank=True, default="")
    ocr_for_occ_name = models.CharField(max_length=250, blank=True, default="")

    # If this OCR was created as part of a bulk import task, this field will be populated
    bulk_import_task = models.ForeignKey(
        "OccurrenceReportBulkImportTask",
        on_delete=models.PROTECT,
        null=True,
        blank=True,
        related_name="occurrence_reports",
    )
    # A hash of the import row data to allow for duplicate detection
    import_hash = models.CharField(max_length=64, null=True, blank=True)

    class Meta:
        app_label = "boranga"
        ordering = ["-id"]

    def __str__(self):
        return str(self.occurrence_report_number)

    def save(self, *args, **kwargs):
        # Clear the cache
        cache.delete(settings.CACHE_KEY_MAP_OCCURRENCE_REPORTS)
        if self.occurrence_report_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(no_revision=True, force_insert=force_insert)
            new_occurrence_report_id = f"OCR{str(self.pk)}"
            self.occurrence_report_number = new_occurrence_report_id
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)

    @property
    def reference(self):
        return f"{self.occurrence_report_number}"

    @property
    def applicant(self):
        if self.submitter:
            email_user = retrieve_email_user(self.submitter)
            return f"{email_user.first_name} {email_user.last_name}"

    @property
    def applicant_email(self):
        if self.submitter:
            email_user = retrieve_email_user(self.submitter)
            return email_user.email

    @property
    def applicant_details(self):
        if self.submitter:
            email_user = retrieve_email_user(self.submitter)
            return f"{email_user.first_name} {email_user.last_name}"

    @property
    def applicant_address(self):
        if self.submitter:
            email_user = retrieve_email_user(self.submitter)
            return email_user.residential_address

    @property
    def applicant_id(self):
        if self.submitter:
            email_user = retrieve_email_user(self.submitter)
            return email_user.id

    @property
    def applicant_type(self):
        if self.submitter:
            # return self.APPLICANT_TYPE_SUBMITTER
            return "SUB"

    @property
    def applicant_field(self):
        # if self.org_applicant:
        #     return 'org_applicant'
        # elif self.proxy_applicant:
        #     return 'proxy_applicant'
        # else:
        #     return 'submitter'
        return "submitter"

    def log_user_action(self, action, request):
        return OccurrenceReportUserAction.log_action(self, action, request.user.id)

    def can_user_edit(self, request):
        """
        :return: True if the occurrence report is in one of the editable status.
        """
        if self.customer_status not in self.CUSTOMER_EDITABLE_STATE:
            return False

        if not request.user.id == self.submitter:
            return False

        return True

    @property
    def can_user_view(self):
        """
        :return: True if the occurrence report is in one of the approved status.
        """
        return self.customer_status in self.CUSTOMER_VIEWABLE_STATE

    @property
    def is_flora_application(self):
        if self.group_type.name == GroupType.GROUP_TYPE_FLORA:
            return True
        return False

    @property
    def is_fauna_application(self):
        if self.group_type.name == GroupType.GROUP_TYPE_FAUNA:
            return True
        return False

    @property
    def is_community_application(self):
        if self.group_type.name == GroupType.GROUP_TYPE_COMMUNITY:
            return True
        return False

    @property
    def finalised(self):
        return self.processing_status in self.FINALISED_STATUSES

    @property
    def allowed_assessors(self):
        group_ids = None
        if self.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
        ]:
            group_ids = member_ids(GROUP_NAME_OCCURRENCE_APPROVER)
            users = (
                list(
                    map(
                        lambda id: retrieve_email_user(id),
                        group_ids,
                    )
                )
                if group_ids
                else []
            )
            return users
        elif self.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
            OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
            OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
        ]:
            group_ids = member_ids(GROUP_NAME_OCCURRENCE_ASSESSOR)
            users = (
                list(
                    map(
                        lambda id: retrieve_email_user(id),
                        group_ids,
                    )
                )
                if group_ids
                else []
            )

            return list(set(users))
        else:
            return []

    @property
    def number_of_observers(self):
        return self.observer_count

    @property
    def has_main_observer(self):
        return self.observer_detail.filter(visible=True, main_observer=True).exists()

    def has_assessor_mode(self, request):
        status_with_assessor = [
            OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
            OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
        ]
        if self.processing_status not in status_with_assessor:
            return False

        if request.user.is_superuser:
            return True

        if not self.assigned_officer:
            return False

        if not self.assigned_officer == request.user.id:
            return False

        return is_occurrence_assessor(request)

    def has_approver_mode(self, request):
        status_with_approver = [
            "with_approver",
        ]
        if self.processing_status not in status_with_approver:
            return False

        if request.user.is_superuser:
            return True

        if not self.assigned_approver:
            return False

        if not self.assigned_approver == request.user.id:
            return False

        return is_occurrence_approver(request) or request.user.is_superuser

    def has_unlocked_mode(self, request):
        status_with_assessor = [
            "unlocked",
        ]
        if self.processing_status not in status_with_assessor:
            return False

        if not self.assigned_officer:
            return False

        if not self.assigned_officer == request.user.id:
            return False

        return (
            is_occurrence_assessor(request)
            or is_occurrence_approver(request)
            or request.user.is_superuser
        )

    def get_approver_group(self):
        return SystemGroup.objects.get(name=GROUP_NAME_OCCURRENCE_APPROVER)

    @property
    def related_item_identifier(self):
        return self.occurrence_report_number

    @property
    def related_item_descriptor(self):
        if self.species:
            if self.species.taxonomy and self.species.taxonomy.scientific_name:
                return self.species.taxonomy.scientific_name
        return "Descriptor not available"

    @property
    def related_item_status(self):
        return self.get_processing_status_display()

    @property
    def as_related_item(self):
        related_item = RelatedItem(
            identifier=self.related_item_identifier,
            model_name=self._meta.verbose_name.title(),
            descriptor=self.related_item_descriptor,
            status=self.related_item_status,
            action_url=(
                f'<a href="/internal/occurrence-report/{self.id}'
                f'?action=view" target="_blank">View '
                '<i class="bi bi-box-arrow-up-right"></i></a>'
            ),
        )
        return related_item

    def can_assess(self, request):
        if self.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
            OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
            OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
        ]:
            return (
                is_occurrence_assessor(request)
                or is_occurrence_approver(request)
                or request.user.is_superuser
            )

        elif self.processing_status == OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER:
            return is_occurrence_approver(request) or request.user.is_superuser

        return False

    def can_change_lock(self, request):
        if self.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
            OccurrenceReport.PROCESSING_STATUS_APPROVED,
        ]:
            return (
                is_occurrence_assessor(request)
                or is_occurrence_approver(request)
                or request.user.is_superuser
            )

    @transaction.atomic
    def discard(self, request):
        if not self.processing_status == OccurrenceReport.PROCESSING_STATUS_DRAFT:
            raise exceptions.OccurrenceReportNotAuthorized()

        self.processing_status = OccurrenceReport.PROCESSING_STATUS_DISCARDED
        self.customer_status = OccurrenceReport.CUSTOMER_STATUS_DISCARDED
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_DISCARD_PROPOSAL.format(
                self.occurrence_report_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_DISCARD_PROPOSAL.format(
                self.occurrence_report_number
            ),
            request,
        )

    @transaction.atomic
    def reinstate(self, request):
        if not self.processing_status == OccurrenceReport.PROCESSING_STATUS_DISCARDED:
            raise exceptions.OccurrenceReportNotAuthorized()

        self.processing_status = OccurrenceReport.PROCESSING_STATUS_DRAFT
        self.customer_status = OccurrenceReport.CUSTOMER_STATUS_DRAFT
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_REINSTATE_PROPOSAL.format(
                self.occurrence_report_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_REINSTATE_PROPOSAL.format(
                self.occurrence_report_number
            ),
            request,
        )

    @transaction.atomic
    def reassign_draft_to_user(self, request, user_id):
        if not self.processing_status == OccurrenceReport.PROCESSING_STATUS_DRAFT:
            raise exceptions.OccurrenceReportNotAuthorized()

        try:
            new_submitter = EmailUser.objects.get(id=user_id)
        except EmailUser.DoesNotExist:
            raise serializers.ValidationError(
                f"EmailUserRO with id {user_id} does not exist"
            )

        previous_submitter = EmailUser.objects.get(id=self.submitter)
        self.submitter = new_submitter.id
        self.save(version_user=request.user)

        self.log_user_action(
            OccurrenceReportUserAction.ACTION_REASSIGN_DRAFT_TO_USER.format(
                self.occurrence_report_number,
                previous_submitter.get_full_name(),
                new_submitter.get_full_name(),
            ),
            request,
        )
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_REASSIGN_DRAFT_TO_USER.format(
                self.occurrence_report_number,
                previous_submitter.get_full_name(),
                new_submitter.get_full_name(),
            ),
            request,
        )

    @transaction.atomic
    def assign_officer(self, request, officer):
        if not self.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        if self.processing_status == OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER:
            if officer.id != self.assigned_approver:
                self.assigned_approver = officer.id
                self.save(version_user=request.user)

                # Create a log entry for the proposal
                self.log_user_action(
                    OccurrenceReportUserAction.ACTION_ASSIGN_TO_APPROVER.format(
                        self.occurrence_report_number,
                        f"{officer.get_full_name()}({officer.email})",
                    ),
                    request,
                )

                # Create a log entry for the user
                request.user.log_user_action(
                    OccurrenceReportUserAction.ACTION_ASSIGN_TO_APPROVER.format(
                        self.occurrence_report_number,
                        f"{officer.get_full_name()}({officer.email})",
                    ),
                    request,
                )
        else:
            if officer.id != self.assigned_officer:
                self.assigned_officer = officer.id
                self.save(version_user=request.user)

                # Create a log entry for the proposal
                self.log_user_action(
                    OccurrenceReportUserAction.ACTION_ASSIGN_TO_ASSESSOR.format(
                        self.occurrence_report_number,
                        f"{officer.get_full_name()}({officer.email})",
                    ),
                    request,
                )
                # Create a log entry for the user
                request.user.log_user_action(
                    OccurrenceReportUserAction.ACTION_ASSIGN_TO_ASSESSOR.format(
                        self.occurrence_report_number,
                        f"{officer.get_full_name()}({officer.email})",
                    ),
                    request,
                )

    def unassign(self, request):
        if not self.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        if self.processing_status == OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER:
            if self.assigned_approver:
                self.assigned_approver = None
                self.save(version_user=request.user)

                # Create a log entry for the proposal
                self.log_user_action(
                    OccurrenceReportUserAction.ACTION_UNASSIGN_APPROVER.format(
                        self.occurrence_report_number
                    ),
                    request,
                )

                # Create a log entry for the user
                request.user.log_user_action(
                    OccurrenceReportUserAction.ACTION_UNASSIGN_APPROVER.format(
                        self.occurrence_report_number
                    ),
                    request,
                )
        else:
            if self.assigned_officer:
                self.assigned_officer = None
                self.save(version_user=request.user)

                # Create a log entry for the proposal
                self.log_user_action(
                    OccurrenceReportUserAction.ACTION_UNASSIGN_ASSESSOR.format(
                        self.occurrence_report_number
                    ),
                    request,
                )

                # Create a log entry for the user
                request.user.log_user_action(
                    OccurrenceReportUserAction.ACTION_UNASSIGN_ASSESSOR.format(
                        self.occurrence_report_number
                    ),
                    request,
                )

    @property
    def amendment_requests(self):
        return OccurrenceReportAmendmentRequest.objects.filter(occurrence_report=self)

    @transaction.atomic
    def propose_decline(self, request, details):
        if not self.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        if self.processing_status != OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR:
            raise ValidationError(
                f"You cannot propose to decline Occurrence Report {self} as the processing status is not "
                f"{OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR}"
            )

        reason = details.get("reason")
        OccurrenceReportDeclinedDetails.objects.update_or_create(
            occurrence_report=self,
            defaults={
                "officer": request.user.id,
                "reason": reason,
                "cc_email": details.get("cc_email", None),
            },
        )

        self.proposed_decline_status = True
        self.approver_comment = ""
        OccurrenceReportApprovalDetails.objects.filter(occurrence_report=self).delete()
        self.processing_status = OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_PROPOSED_DECLINE.format(
                self.occurrence_report_number
            ),
            request,
        )

        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_PROPOSED_DECLINE.format(
                self.occurrence_report_number
            ),
            request,
        )

        send_approver_decline_email_notification(reason, request, self)

    @transaction.atomic
    def decline(self, request, details):
        if not self.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        if self.processing_status != OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER:
            raise ValidationError(
                f"You cannot decline Occurrence Report {self} as the processing status is not "
                f"{OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER}"
            )

        reason = details.get("reason")

        self.processing_status = OccurrenceReport.PROCESSING_STATUS_DECLINED
        self.customer_status = OccurrenceReport.CUSTOMER_STATUS_DECLINED
        self.occurrence = None
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_DECLINE.format(
                self.occurrence_report_number,
                reason,
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_DECLINE.format(
                self.occurrence_report_number,
                reason,
            ),
            request,
        )

        send_decline_email_notification(reason, self)

    def validate_submit(self):
        missing_values = []

        if not self.observation_date:
            missing_values.append("Observation Date")

        if self.observer_detail.count() < 1:
            missing_values.append("Observer Details")

        if not self.location or not self.location.location_description:
            missing_values.append("Location Description")

        if self.ocr_geometry.count() < 1:
            missing_values.append("Location")

        if missing_values:
            raise ValidationError(
                "Cannot submit this report due to missing values: "
                + ", ".join(missing_values)
            )

    def validate_propose_approve(self):
        self.validate_submit()

        missing_values = []

        if not self.identification or not self.identification.identification_certainty:
            missing_values.append("Identification Certainty")

        if not self.location or not self.location.location_accuracy:
            missing_values.append("Location Accuracy")

        if missing_values:
            raise ValidationError(
                "Cannot submit this report due to missing values: "
                + ", ".join(missing_values)
            )

    @transaction.atomic
    def propose_approve(self, request, validated_data):

        self.validate_propose_approve()

        if not self.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        if self.processing_status != OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR:
            raise ValidationError(
                f"You cannot propose to decline Occurrence Report {self} as the processing status is not "
                f"{OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR}"
            )

        occurrence = None
        occurrence_id = validated_data.get("occurrence_id", None)
        if occurrence_id:
            try:
                occurrence = Occurrence.objects.get(id=occurrence_id)
            except Occurrence.DoesNotExist:
                raise ValidationError(
                    f"Occurrence with id {occurrence_id} does not exist"
                )

        details = validated_data.get("details", None)
        new_occurrence_name = validated_data.get("new_occurrence_name", None)

        if new_occurrence_name and (
            Occurrence.objects.filter(occurrence_name=new_occurrence_name).exists()
            or OccurrenceReportApprovalDetails.objects.filter(
                new_occurrence_name=new_occurrence_name
            ).exists()
        ):
            raise ValidationError(
                f'Occurrence with name "{new_occurrence_name}" already exists or has been proposed for approval'
            )

        OccurrenceReportApprovalDetails.objects.update_or_create(
            occurrence_report=self,
            defaults={
                "officer": request.user.id,
                "occurrence": occurrence,
                "new_occurrence_name": new_occurrence_name,
                "copy_ocr_comments_to_occ_comments": validated_data.get(
                    "copy_ocr_comments_to_occ_comments", True
                ),
                "details": details,
                "cc_email": validated_data.get("cc_email", None),
            },
        )

        self.approver_comment = ""
        self.proposed_decline_status = False
        OccurrenceReportDeclinedDetails.objects.filter(occurrence_report=self).delete()
        self.processing_status = OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_PROPOSED_APPROVAL.format(
                self.occurrence_report_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_PROPOSED_APPROVAL.format(
                self.occurrence_report_number
            ),
            request,
        )

        send_approver_approve_email_notification(request, self)

    @transaction.atomic
    def approve(self, request):
        if not self.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        if self.processing_status != OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER:
            raise ValidationError(
                f"You cannot approve Occurrence Report {self} as the processing status is not "
                f"{OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER}"
            )

        if not self.approval_details:
            raise ValidationError(
                f"Approval details are required to approve Occurrence Report {self}"
            )

        self.processing_status = OccurrenceReport.PROCESSING_STATUS_APPROVED
        self.customer_status = OccurrenceReport.CUSTOMER_STATUS_APPROVED
        self.approved_by = request.user.id

        if self.approval_details.occurrence:
            occurrence = self.approval_details.occurrence

            if self.approval_details.copy_ocr_comments_to_occ_comments:
                if not occurrence.comment:
                    occurrence.comment = self.comments
                else:
                    occurrence.comment = occurrence.comment + "\n\n" + self.comments
                occurrence.save(version_user=request.user)

        else:
            if not self.approval_details.new_occurrence_name:
                raise ValidationError(
                    "New occurrence name is required to approve Occurrence Report"
                )
            occurrence = Occurrence.clone_from_occurrence_report(self)
            occurrence.occurrence_name = self.approval_details.new_occurrence_name
            occurrence.save(version_user=request.user)

        self.occurrence = occurrence
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_APPROVE.format(
                self.occurrence_report_number,
                request.user.get_full_name(),
            ),
            request,
        )

        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_APPROVE.format(
                self.occurrence_report_number,
                request.user.get_full_name(),
            ),
            request,
        )

        send_approve_email_notification(self)

    @transaction.atomic
    def back_to_assessor(self, request, validated_data):
        if not self.can_assess(request) or self.processing_status not in [
            OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
            OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
        ]:
            raise exceptions.OccurrenceReportNotAuthorized()

        self.processing_status = OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR
        self.save(version_user=request.user)

        reason = validated_data.get("reason", "")

        # Create a log entry for the proposal
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_BACK_TO_ASSESSOR.format(
                self.occurrence_report_number,
                reason,
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_BACK_TO_ASSESSOR.format(
                self.occurrence_report_number,
                reason,
            ),
            request,
        )

        send_approver_back_to_assessor_email_notification(request, self, reason)

    def lock(self, request):
        if (
            self.can_change_lock(request)
            and self.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.processing_status = OccurrenceReport.PROCESSING_STATUS_APPROVED
            self.save(version_user=request.user)

            self.log_user_action(
                OccurrenceReportUserAction.ACTION_LOCK.format(
                    self.occurrence_report_number
                ),
                request,
            )

            request.user.log_user_action(
                OccurrenceReportUserAction.ACTION_LOCK.format(
                    self.occurrence_report_number
                ),
                request,
            )

    def unlock(self, request):
        if (
            self.can_change_lock(request)
            and self.processing_status == OccurrenceReport.PROCESSING_STATUS_APPROVED
        ):
            self.processing_status = OccurrenceReport.PROCESSING_STATUS_UNLOCKED
            if self.assigned_officer != request.user.id:
                self.assigned_officer = request.user.id
            self.save(version_user=request.user)

            self.log_user_action(
                OccurrenceReportUserAction.ACTION_UNLOCK.format(
                    self.occurrence_report_number
                ),
                request,
            )

            request.user.log_user_action(
                OccurrenceReportUserAction.ACTION_UNLOCK.format(
                    self.occurrence_report_number
                ),
                request,
            )

    @property
    def latest_referrals(self):
        return self.referrals.all()[: settings.RECENT_REFERRAL_COUNT]

    def assessor_comments_view(self, request):
        if self.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
            OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
            OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
            OccurrenceReport.PROCESSING_STATUS_APPROVED,
            OccurrenceReport.PROCESSING_STATUS_DECLINED,
            OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
            OccurrenceReport.PROCESSING_STATUS_CLOSED,
        ]:
            if OccurrenceReportReferral.objects.filter(
                occurrence_report=self, referral=request.user.id
            ).exists():
                return True

            return (
                is_occurrence_assessor(request)
                or is_occurrence_approver(request)
                or request.user.is_superuser
            )
        return False

    @transaction.atomic
    def send_referral(self, request, referral_email, referral_text):
        referral_email = referral_email.lower()
        if self.processing_status not in [
            OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
            OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
        ]:
            raise exceptions.OccurrenceReportReferralCannotBeSent()

        if (
            not self.processing_status
            == OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL
        ):
            self.processing_status = OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL
            self.save(version_user=request.user)

        referral = None

        # Check if the user is in ledger
        try:
            referee = EmailUser.objects.get(email__iexact=referral_email.strip())
        except EmailUser.DoesNotExist:
            raise ValidationError(
                "The user you want to send the referral to does not exist in the ledger database"
            )

        # Don't allow the user to refer to themselves
        if referee.id == request.user.id:
            raise ValidationError("You cannot refer to yourself")

        # Don't allow the user to refer to the submitter
        if referee.id == self.submitter:
            raise ValidationError("You cannot refer to the submitter")

        # Check if the referral has already been sent to this user
        if OccurrenceReportReferral.objects.filter(
            referral=referee.id, occurrence_report=self
        ).exists():
            raise ValidationError("A referral has already been sent to this user")

        # Create Referral
        referral = OccurrenceReportReferral.objects.create(
            occurrence_report=self,
            referral=referee.id,
            sent_by=request.user.id,
            text=referral_text,
            assigned_officer=request.user.id,
        )

        # Create a log entry for the proposal
        self.log_user_action(
            OccurrenceReportUserAction.ACTION_SEND_REFERRAL_TO.format(
                referral.id,
                self.occurrence_report_number,
                f"{referee.get_full_name()}({referee.email})",
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_SEND_REFERRAL_TO.format(
                referral.id,
                self.occurrence_report_number,
                f"{referee.get_full_name()}({referee.email})",
            ),
            request,
        )

        # send email
        send_occurrence_report_referral_email_notification(referral, request)

    @property
    def external_referral_invites(self):
        return self.external_referee_invites.filter(
            archived=False, datetime_first_logged_in__isnull=True
        )

    @transaction.atomic
    def copy(self, request_user_id):
        ocr_copy = OccurrenceReport.objects.get(id=self.id)
        ocr_copy.pk = None
        ocr_copy.processing_status = OccurrenceReport.PROCESSING_STATUS_DRAFT
        ocr_copy.customer_status = OccurrenceReport.CUSTOMER_STATUS_DRAFT
        ocr_copy.occurrence_report_number = ""
        ocr_copy.lodgement_date = None
        ocr_copy.observation_date = None
        ocr_copy.comments = None
        ocr_copy.assigned_officer = None
        ocr_copy.assigned_approver = None
        ocr_copy.approved_by = None
        ocr_copy.submitter_information = None
        if request_user_id != self.submitter:
            ocr_copy.submitter = request_user_id
            ocr_copy.internal_application = True
        ocr_copy.save(no_revision=True)

        if request_user_id == self.submitter:
            # Use the same submitter category as the previous proposal when the user copying is the submitter
            ocr_copy.submitter_information.submitter_category_id = (
                self.submitter_information.submitter_category_id
            )
            ocr_copy.submitter_information.save()

        # Clone all the associated models
        if hasattr(self, "location") and self.location:
            location = clone_model(
                OCRLocation,
                OCRLocation,
                self.location,
            )
            if location:
                location.occurrence_report = ocr_copy
                location.save()

        if hasattr(self, "habitat_composition") and self.habitat_composition:
            habitat_composition = clone_model(
                OCRHabitatComposition,
                OCRHabitatComposition,
                self.habitat_composition,
            )
            if habitat_composition:
                habitat_composition.occurrence_report = ocr_copy
                habitat_composition.save()

        if hasattr(self, "habitat_condition") and self.habitat_condition:
            habitat_condition = clone_model(
                OCRHabitatCondition,
                OCRHabitatCondition,
                self.habitat_condition,
            )
            if habitat_condition:
                habitat_condition.occurrence_report = ocr_copy
                habitat_condition.save()

        if hasattr(self, "vegetation_structure") and self.vegetation_structure:
            vegetation_structure = clone_model(
                OCRVegetationStructure,
                OCRVegetationStructure,
                self.vegetation_structure,
            )
            if vegetation_structure:
                vegetation_structure.occurrence_report = ocr_copy
                vegetation_structure.save()

        if hasattr(self, "fire_history") and self.fire_history:
            fire_history = clone_model(
                OCRFireHistory, OCRFireHistory, self.fire_history
            )
            if fire_history:
                fire_history.occurrence_report = ocr_copy
                fire_history.save()

        if hasattr(self, "associated_species") and self.associated_species:
            associated_species = clone_model(
                OCRAssociatedSpecies,
                OCRAssociatedSpecies,
                self.associated_species,
            )
            if associated_species:
                associated_species.occurrence_report = ocr_copy
                associated_species.save()
                # copy over related species separately
                for i in self.associated_species.related_species.all():
                    associated_species.related_species.add(i)

        # Clone the threats
        for threat in self.ocr_threats.all():
            ocr_threat = clone_model(
                OCRConservationThreat, OCRConservationThreat, threat
            )
            if ocr_threat:
                ocr_threat.occurrence_report = ocr_copy
                ocr_threat.occurrence_report_threat = threat
                ocr_threat.save()

        # Clone the documents
        for doc in self.documents.all():
            ocr_doc = clone_model(
                OccurrenceReportDocument, OccurrenceReportDocument, doc
            )
            if ocr_doc:
                ocr_doc.occurrence_report = ocr_copy
                ocr_doc.save()

        # Clone any observers
        observer_qs = self.observer_detail.all()
        if request_user_id == self.submitter:
            # If the user copying is not the submitter, only copy the main observer
            observer_qs = self.observer_detail.filter(main_observer=True)
        for observer in observer_qs:
            ocr_observer = clone_model(OCRObserverDetail, OCRObserverDetail, observer)
            if ocr_observer:
                ocr_observer.occurrence_report = ocr_copy
                ocr_observer.save()

        # Clone any occurrence geometries
        for geom in self.ocr_geometry.all():
            ocr_geom = clone_model(
                OccurrenceReportGeometry, OccurrenceReportGeometry, geom
            )
            if ocr_geom:
                ocr_geom.occurrence_report = ocr_copy
                ocr_geom.save()

        # For flora create an empty plant count
        if self.group_type.name == GroupType.GROUP_TYPE_FLORA:
            plant_count = OCRPlantCount()
            plant_count.occurrence_report = ocr_copy
            plant_count.save()

        # For fauna create an empty animal observation
        if self.group_type.name == GroupType.GROUP_TYPE_FAUNA:
            animal_observation = OCRAnimalObservation()
            animal_observation.occurrence_report = ocr_copy
            animal_observation.save()

        # Create an empty observation detail
        observation_detail = OCRObservationDetail()
        observation_detail.occurrence_report = ocr_copy
        observation_detail.save()

        # Create an empty identification
        identification = OCRIdentification()
        identification.occurrence_report = ocr_copy
        identification.save()

        return ocr_copy


class OccurrenceReportDeclinedDetails(models.Model):
    occurrence_report = models.OneToOneField(
        OccurrenceReport, on_delete=models.CASCADE, related_name="declined_details"
    )
    officer = models.IntegerField()  # EmailUserRO
    reason = models.TextField(blank=True)
    cc_email = models.TextField(null=True)

    class Meta:
        app_label = "boranga"


class OccurrenceReportApprovalDetails(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrapp"

    occurrence_report = models.OneToOneField(
        OccurrenceReport, on_delete=models.CASCADE, related_name="approval_details"
    )
    occurrence = models.ForeignKey(
        "Occurrence", on_delete=models.PROTECT, null=True, blank=True
    )  # If being added to an existing occurrence
    new_occurrence_name = models.CharField(max_length=200, null=True, blank=True)
    officer = models.IntegerField()  # EmailUserRO
    copy_ocr_comments_to_occ_comments = models.BooleanField(default=True)
    details = models.TextField(blank=True)
    cc_email = models.TextField(null=True)

    class Meta:
        app_label = "boranga"

    def save(self, *args, **kwargs):
        if self.occurrence and self.new_occurrence_name:
            raise ValidationError(
                "You can't have both an existing occurrence and a new occurrence name"
            )
        if not self.occurrence and not self.new_occurrence_name:
            raise ValidationError(
                "You must have either an existing occurrence or a new occurrence name"
            )
        super().save(*args, **kwargs)

    @property
    def officer_name(self):
        if not self.officer:
            return None

        return retrieve_email_user(self.officer).get_full_name()


class OccurrenceReportLogEntry(CommunicationsLogEntry):
    occurrence_report = models.ForeignKey(
        OccurrenceReport, related_name="comms_logs", on_delete=models.CASCADE
    )

    def __str__(self):
        return f"{self.reference} - {self.subject}"

    class Meta:
        app_label = "boranga"

    def save(self, **kwargs):
        # save the application reference if the reference not provided
        if not self.reference:
            self.reference = self.occurrence_report.reference
        super().save(**kwargs)


class OccurrenceReportLogDocument(Document):
    log_entry = models.ForeignKey(
        "OccurrenceReportLogEntry", related_name="documents", on_delete=models.CASCADE
    )
    _file = models.FileField(
        upload_to=update_occurrence_report_comms_log_filename,
        max_length=512,
        storage=private_storage,
    )

    class Meta:
        app_label = "boranga"

    def get_parent_instance(self) -> models.Model:
        return self.log_entry


class OccurrenceReportUserAction(UserAction):
    # OccurrenceReport Proposal
    ACTION_EDIT_OCCURRENCE_REPORT = "Edit occurrence report {}"
    ACTION_LODGE_PROPOSAL = "Lodge occurrence report {}"
    ACTION_SAVE_APPLICATION = "Save occurrence report {}"
    ACTION_EDIT_APPLICATION = "Edit occurrence report {}"
    ACTION_ASSIGN_TO_ASSESSOR = "Assign occurrence report {} to {} as the assessor"
    ACTION_UNASSIGN_ASSESSOR = "Unassign assessor from occurrence report {}"
    ACTION_ASSIGN_TO_APPROVER = "Assign occurrence report {} to {} as the approver"
    ACTION_UNASSIGN_APPROVER = "Unassign approver from occurrence report {}"
    ACTION_DECLINE = "Occurrence Report {} has been declined. Reason: {}"
    ACTION_APPROVE = "Occurrence Report {} has been approved by {}"
    ACTION_UNLOCK = "Unlock occurrence report {}"
    ACTION_LOCK = "Lock occurrence report {}"
    ACTION_CLOSE_OccurrenceReport = "De list occurrence report {}"
    ACTION_DISCARD_PROPOSAL = "Discard occurrence report {}"
    ACTION_REINSTATE_PROPOSAL = "Reinstate occurrence report {}"
    ACTION_APPROVAL_LEVEL_DOCUMENT = "Assign Approval level document {}"
    ACTION_UPDATE_OBSERVER_DETAIL = "Update Observer {} on occurrence report {}"
    ACTION_COPY = "Created occurrence report {} from a copy of occurrence report {}"
    ACTION_COPY_TO = "Copy occurrence report to {}"
    ACTION_REASSIGN_DRAFT_TO_USER = (
        "Occurrence Report {} (draft) reassigned from {} to {}"
    )

    # Amendment
    ACTION_ID_REQUEST_AMENDMENTS = "Request amendments"

    # Assessors
    ACTION_SAVE_ASSESSMENT_ = "Save assessment {}"
    ACTION_CONCLUDE_ASSESSMENT_ = "Conclude assessment {}"
    ACTION_PROPOSED_READY_FOR_AGENDA = (
        "Occurrence report {} has been proposed as 'ready for agenda'"
    )
    ACTION_PROPOSED_APPROVAL = (
        "Occurrence report {} has been proposed as 'for approval'"
    )
    ACTION_PROPOSED_DECLINE = "Occurrence report {} has been proposed as 'for decline'"

    # Referrals
    ACTION_SEND_REFERRAL_TO = "Send referral {} for occurrence report {} to {}"
    ACTION_RESEND_REFERRAL_TO = "Resend referral {} for occurrence report {} to {}"
    ACTION_REMIND_REFERRAL = (
        "Send reminder for referral {} for occurrence report {} to {}"
    )
    ACTION_BACK_TO_ASSESSOR = "{} sent back to assessor. Reason: {}"
    RECALL_REFERRAL = "Referral {} for occurrence report {} has been recalled by {}"
    SAVE_REFERRAL = "Referral {} for occurrence report {} has been saved by {}"
    CONCLUDE_REFERRAL = "Referral {} for occurrence report {} has been concluded by {}"

    # Document
    ACTION_ADD_DOCUMENT = "Document {} added for occurrence report {}"
    ACTION_UPDATE_DOCUMENT = "Document {} updated for occurrence report {}"
    ACTION_DISCARD_DOCUMENT = "Document {} discarded for occurrence report {}"
    ACTION_REINSTATE_DOCUMENT = "Document {} reinstated for occurrence report {}"

    # Threat
    ACTION_ADD_THREAT = "Threat {} added for occurrence report {}"
    ACTION_UPDATE_THREAT = "Threat {} updated for occurrence report {}"
    ACTION_DISCARD_THREAT = "Threat {} discarded for occurrence report {}"
    ACTION_REINSTATE_THREAT = "Threat {} reinstated for occurrence report {}"

    class Meta:
        app_label = "boranga"
        ordering = ("-when",)

    @classmethod
    def log_action(cls, occurrence_report, action, user):
        return cls.objects.create(
            occurrence_report=occurrence_report, who=user, what=str(action)
        )

    occurrence_report = models.ForeignKey(
        OccurrenceReport, related_name="action_logs", on_delete=models.CASCADE
    )


def update_occurrence_report_referral_doc_filename(instance, filename):
    return "{}/occurrence_report/{}/referral/{}".format(
        settings.MEDIA_APP_DIR, instance.referral.occurrence_report.id, filename
    )


class OccurrenceReportProposalRequest(models.Model):
    occurrence_report = models.ForeignKey(OccurrenceReport, on_delete=models.CASCADE)
    text = models.TextField(blank=True)

    class Meta:
        app_label = "boranga"
        ordering = ["id"]


class OccurrenceReportAmendmentRequest(OccurrenceReportProposalRequest):
    STATUS_CHOICE_REQUESTED = "requested"
    STATUS_CHOICE_AMENDED = "amended"

    STATUS_CHOICES = (
        (STATUS_CHOICE_REQUESTED, "Requested"),
        (STATUS_CHOICE_AMENDED, "Amended"),
    )

    status = models.CharField(
        "Status", max_length=30, choices=STATUS_CHOICES, default=STATUS_CHOICES[0][0]
    )
    reason = models.ForeignKey(
        ProposalAmendmentReason, blank=True, null=True, on_delete=models.SET_NULL
    )

    class Meta:
        app_label = "boranga"
        ordering = ["id"]

    @transaction.atomic
    def generate_amendment(self, request):
        if not self.occurrence_report.can_assess(request):
            raise exceptions.ProposalNotAuthorized()

        if self.status == "requested":
            occurrence_report = self.occurrence_report
            if occurrence_report.processing_status != "draft":
                occurrence_report.processing_status = "draft"
                occurrence_report.customer_status = "draft"
                occurrence_report.save(version_user=request.user)

            # Create a log entry for the occurrence report
            occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_ID_REQUEST_AMENDMENTS, request
            )

            # Create a log entry for the user
            request.user.log_user_action(
                OccurrenceReportUserAction.ACTION_ID_REQUEST_AMENDMENTS, request
            )

            # send email
            send_occurrence_report_amendment_email_notification(
                self, request, occurrence_report
            )

        self.save()

    @transaction.atomic
    def add_documents(self, request):
        # save the files
        data = json.loads(request.data.get("data"))

        if not data.get("update"):
            documents_qs = self.amendment_request_documents.filter(
                input_name="amendment_request_doc", visible=True
            )
            documents_qs.delete()

        for idx in range(data["num_files"]):
            _file = request.data.get("file-" + str(idx))
            document = self.amendment_request_documents.create(
                _file=_file, name=_file.name
            )
            document.check_file(request.data.get("file-" + str(idx)))
            document.input_name = data["input_name"]
            document.save()

        # end save documents
        self.save()


def update_occurrence_report_amendment_request_doc_filename(instance, filename):
    return "occurrence_report/{}/amendment_request_documents/{}".format(
        instance.occurrence_report_amendment_request.occurrence_report.id, filename
    )


class OccurrenceReportAmendmentRequestDocument(Document):
    occurrence_report_amendment_request = models.ForeignKey(
        "OccurrenceReportAmendmentRequest",
        related_name="amendment_request_documents",
        on_delete=models.CASCADE,
    )
    _file = models.FileField(
        upload_to=update_occurrence_report_amendment_request_doc_filename,
        max_length=500,
        storage=private_storage,
    )
    input_name = models.CharField(max_length=255, null=True, blank=True)

    def get_parent_instance(self) -> models.Model:
        return self.occurrence_report_amendment_request


class OccurrenceReportReferral(models.Model):
    SENT_CHOICE_FROM_ASSESSOR = 1
    SENT_CHOICE_FROM_REFERRAL = 2

    SENT_CHOICES = (
        (SENT_CHOICE_FROM_ASSESSOR, "Sent From Assessor"),
        (SENT_CHOICE_FROM_REFERRAL, "Sent From Referral"),
    )
    PROCESSING_STATUS_WITH_REFERRAL = "with_referral"
    PROCESSING_STATUS_RECALLED = "recalled"
    PROCESSING_STATUS_COMPLETED = "completed"
    PROCESSING_STATUS_CHOICES = (
        (PROCESSING_STATUS_WITH_REFERRAL, "Awaiting"),
        (PROCESSING_STATUS_RECALLED, "Recalled"),
        (PROCESSING_STATUS_COMPLETED, "Completed"),
    )
    lodged_on = models.DateTimeField(auto_now_add=True)
    occurrence_report = models.ForeignKey(
        OccurrenceReport, related_name="referrals", on_delete=models.CASCADE
    )
    sent_by = models.IntegerField()  # EmailUserRO
    referral = models.IntegerField()  # EmailUserRO
    linked = models.BooleanField(default=False)
    processing_status = models.CharField(
        "Processing Status",
        max_length=30,
        choices=PROCESSING_STATUS_CHOICES,
        default=PROCESSING_STATUS_CHOICES[0][0],
    )
    text = models.TextField(blank=True)  # Assessor text when send_referral
    referral_comment = models.TextField(blank=True, null=True)  # Referral Comment
    assigned_officer = models.IntegerField(null=True)  # EmailUserRO
    is_external = models.BooleanField(default=False)

    class Meta:
        app_label = "boranga"
        ordering = ("-lodged_on",)

    def __str__(self):
        return "Occurrence Report {} - Referral {}".format(
            self.occurrence_report.id, self.id
        )

    @property
    def can_be_completed(self):
        # Referral cannot be completed until second level referral sent by referral has been completed/recalled
        return not OccurrenceReportReferral.objects.filter(
            sent_by=self.referral,
            occurrence_report=self.occurrence_report,
            processing_status=OccurrenceReportReferral.PROCESSING_STATUS_WITH_REFERRAL,
        ).exists()

    @property
    def referral_as_email_user(self):
        return retrieve_email_user(self.referral)

    @transaction.atomic
    def remind(self, request):
        if not self.occurrence_report.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        # Create a log entry for the proposal
        self.occurrence_report.log_user_action(
            OccurrenceReportUserAction.ACTION_REMIND_REFERRAL.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                f"{self.referral_as_email_user.get_full_name()}",
            ),
            request,
        )

        # Create a log entry for the submitter
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_REMIND_REFERRAL.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                f"{self.referral_as_email_user.get_full_name()}",
            ),
            request,
        )

        # send email
        send_occurrence_report_referral_email_notification(
            self,
            request,
            reminder=True,
        )

    @transaction.atomic
    def recall(self, request):
        if not self.occurrence_report.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        self.processing_status = self.PROCESSING_STATUS_RECALLED
        self.save()

        send_occurrence_report_referral_recall_email_notification(self, request)

        outstanding = self.occurrence_report.referrals.filter(
            processing_status=self.PROCESSING_STATUS_WITH_REFERRAL
        )
        if len(outstanding) == 0:
            self.occurrence_report.processing_status = (
                OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR
            )
            self.occurrence_report.save(version_user=request.user)

        # Create a log entry for the occurrence report
        self.occurrence_report.log_user_action(
            OccurrenceReportUserAction.RECALL_REFERRAL.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                request.user.get_full_name(),
            ),
            request,
        )

        # Create a log entry for the submitter
        request.user.log_user_action(
            OccurrenceReportUserAction.RECALL_REFERRAL.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                request.user.get_full_name(),
            ),
            request,
        )

    @transaction.atomic
    def resend(self, request):
        if not self.occurrence_report.can_assess(request):
            raise exceptions.OccurrenceReportNotAuthorized()

        self.processing_status = self.PROCESSING_STATUS_WITH_REFERRAL
        self.occurrence_report.processing_status = self.PROCESSING_STATUS_WITH_REFERRAL
        self.occurrence_report.save(version_user=request.user)

        self.save()

        # Create a log entry for the occurrence report
        self.occurrence_report.log_user_action(
            OccurrenceReportUserAction.ACTION_RESEND_REFERRAL_TO.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                "{}({})".format(
                    self.referral_as_email_user.get_full_name(),
                    self.referral_as_email_user.email,
                ),
            ),
            request,
        )

        # Create a log entry for the submitter
        request.user.log_user_action(
            OccurrenceReportUserAction.ACTION_RESEND_REFERRAL_TO.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                "{}({})".format(
                    self.referral_as_email_user.get_full_name(),
                    self.referral_as_email_user.email,
                ),
            ),
            request,
        )

        # send email
        send_occurrence_report_referral_email_notification(self, request)

    @transaction.atomic
    def complete(self, request):
        if request.user.id != self.referral:
            raise exceptions.ReferralNotAuthorized()

        self.processing_status = self.PROCESSING_STATUS_COMPLETED
        self.save()

        outstanding = self.occurrence_report.referrals.filter(
            processing_status=self.PROCESSING_STATUS_WITH_REFERRAL
        )
        if len(outstanding) == 0:
            self.occurrence_report.processing_status = (
                OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR
            )
            self.occurrence_report.save(version_user=request.user)

        # Create a log entry for the occurrence report
        self.occurrence_report.log_user_action(
            OccurrenceReportUserAction.CONCLUDE_REFERRAL.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                "{}({})".format(
                    self.referral_as_email_user.get_full_name(),
                    self.referral_as_email_user.email,
                ),
            ),
            request,
        )

        # Create a log entry for the submitter
        request.user.log_user_action(
            OccurrenceReportUserAction.CONCLUDE_REFERRAL.format(
                self.id,
                self.occurrence_report.occurrence_report_number,
                "{}({})".format(
                    self.referral_as_email_user.get_full_name(),
                    self.referral_as_email_user.email,
                ),
            ),
            request,
        )

        send_occurrence_report_referral_complete_email_notification(self, request)

    def can_assess_referral(self):
        return self.processing_status == self.PROCESSING_STATUS_WITH_REFERRAL

    @property
    def can_be_processed(self):
        return self.processing_status == self.PROCESSING_STATUS_WITH_REFERRAL


class Datum(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()

    """
    # Admin List

    Used by:
    - OCRLocation
    - OCCLocation

    """

    srid = models.IntegerField(blank=False, null=False, unique=True)

    @property
    def name(self):
        return f"EPSG:{str(self.srid)} - {pyproj.CRS.from_string(str(self.srid)).name}"

    class Meta(OrderedModel.Meta):
        app_label = "boranga"

    def __str__(self):
        return str(self.srid)

    def save(self, *args, **kwargs):
        if not self.srid:
            raise ValidationError("SRID is required")

        try:
            pyproj.CRS.from_string(str(self.srid))
        except pyproj.exceptions.CRSError:
            raise ValidationError(f"Invalid SRID: {self.srid}")
        else:
            cache_key = settings.CACHE_KEY_EPSG_CODES.format(
                **{"auth_name": "EPSG", "pj_type": "CRS"}
            )
            cache.delete(cache_key)

        super().save(*args, **kwargs)


class CoordinateSource(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - OCRLocation
    - OCCLocation

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Coordinate Source"
        verbose_name_plural = "Coordinate Sources"

    def __str__(self):
        return str(self.name)


class LocationAccuracy(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - OCRLocation
    - OCCLocation

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Location Accuracy"
        verbose_name_plural = "Location Accuracy"

    def __str__(self):
        return str(self.name)


# NOTE: this and OCCLocation have a number of unused fields that should be removed
class OCRLocation(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrloc"

    """
    Location data  for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport, on_delete=models.CASCADE, null=True, related_name="location"
    )
    location_description = models.TextField(null=True, blank=True)
    boundary_description = models.TextField(null=True, blank=True)
    mapped_boundary = models.BooleanField(null=True, blank=True)
    buffer_radius = models.IntegerField(null=True, blank=True, default=0)
    datum = models.ForeignKey(Datum, on_delete=models.SET_NULL, null=True, blank=True)
    epsg_code = models.IntegerField(null=False, blank=False, default=4326)
    coordinate_source = models.ForeignKey(
        CoordinateSource, on_delete=models.SET_NULL, null=True, blank=True
    )
    location_accuracy = models.ForeignKey(
        LocationAccuracy, on_delete=models.SET_NULL, null=True, blank=True
    )

    region = models.ForeignKey(
        Region, default=None, on_delete=models.CASCADE, null=True, blank=True
    )
    district = models.ForeignKey(
        District, default=None, on_delete=models.CASCADE, null=True, blank=True
    )
    locality = models.TextField(default=None, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return (
            f"OCR Location: {self.id} for Occurrence Report: {self.occurrence_report}"
        )


class GeometryManager(models.Manager):
    def get_queryset(self):
        qs = super().get_queryset()
        polygon_ids = qs.extra(
            where=["geometrytype(geometry) LIKE 'POLYGON'"]
        ).values_list("id", flat=True)
        return qs.annotate(
            area=models.Case(
                models.When(
                    models.Q(geometry__isnull=False) & models.Q(id__in=polygon_ids),
                    then=Area(
                        Cast("geometry", gis_models.PolygonField(geography=True))
                    ),
                ),
                default=None,
            )
        )


class GeometryBase(models.Model):
    """
    Base class for geometry models
    """

    objects = GeometryManager()

    EXTENT = (112.5, -35.5, 129.0, -13.5)

    geometry = gis_models.GeometryField(extent=EXTENT, blank=True, null=True)
    original_geometry_ewkb = models.BinaryField(
        blank=True, null=True, editable=True
    )  # original geometry as uploaded by the user in EWKB format (keeps the srid)

    content_type = models.ForeignKey(
        ct_models.ContentType,
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
        related_name="content_type_%(class)s",
    )
    object_id = models.PositiveIntegerField(blank=True, null=True)
    content_object = fields.GenericForeignKey("content_type", "object_id")

    copied_from = fields.GenericRelation("self", related_query_name="copied_to")

    created_date = models.DateTimeField(auto_now_add=True, null=False, blank=False)
    updated_date = models.DateTimeField(auto_now=True, null=False, blank=False)

    class Meta:
        abstract = True
        indexes = [
            models.Index(fields=["content_type", "object_id"]),
        ]

    def save(self, *args, **kwargs):
        if not self.geometry:
            raise ValidationError("Geometry is required")

        if not self.geometry.valid:
            raise ValidationError("Invalid geometry")

        if self.geometry.empty:
            raise ValidationError("Geometry is empty")

        if self.geometry.srid != 4326:
            raise ValidationError(
                f"Cannot save a geometry with SRID {self.geometry.srid} into a WGS-84 (SRID 4326) geometry field."
            )

        if not self.geometry.within(
            GEOSGeometry(Polygon.from_bbox(self.EXTENT), srid=4326)
        ):
            raise ValidationError(
                "Geometry is not within the extent of Western Australia"
            )

        super().save(*args, **kwargs)

    @abstractmethod
    def related_model_field(self):
        """Returns the model field (foreign key) that this geometry model is the geometry of.
        E.g. OccurrenceGeometry is the geometry model of Occurrence"""

        raise NotImplementedError(
            f"Class {self.__class__.__name__} inheriting from {self.__class__.__base__.__name__} "
            f"needs to implement a related_model_field function."
        )

    def __str__(self):
        wkt_ellipsis = ""
        if self.geometry:
            wkt_ellipsis = (
                (self.geometry.wkt[:85] + "..")
                if len(self.geometry.wkt) > 75
                else self.geometry.wkt
            )
        return f"{self.__class__.__name__} of <{self.related_model_field()}>: {wkt_ellipsis}"

    @property
    def area_sqm(self):
        if not hasattr(self, "area") or not self.area:
            return None
        return self.area.sq_m

    @property
    def area_sqhm(self):
        if not hasattr(self, "area") or not self.area:
            return None
        return self.area.sq_m / 10000

    @property
    def original_geometry(self):
        if self.original_geometry_ewkb:
            return GEOSGeometry(self.original_geometry_ewkb)
        return None

    @property
    def original_geometry_srid(self):
        if self.original_geometry_ewkb:
            return GEOSGeometry(self.original_geometry_ewkb).srid
        return None

    def created_from_instance(self):
        if not self.content_type or not self.object_id:
            return None

        InstanceModel = self.content_type.model_class()
        try:
            model_instance = InstanceModel.objects.get(id=self.object_id)
        except InstanceModel.DoesNotExist:
            return None
        else:
            return model_instance

    @property
    def created_from(self):
        """Returns the __str__-representation of the object that this geometry was created from."""

        instance = self.created_from_instance()
        if instance:
            return instance.__str__()
        return None

    def source_of_objects(self):
        content_type = ct_models.ContentType.objects.get_for_model(self.__class__)

        parent_subclasses = self.__class__.__base__.__subclasses__()
        # Get a list of content types for the parent classes of this geometry model
        subclasses_content_types = [
            ct_models.ContentType.objects.get_for_model(psc)
            for psc in parent_subclasses
        ]
        # Get a list of filtered objects (the objects that have been created from self) for each subclass content type
        source_of_objects = [
            sc_ct.get_all_objects_for_this_type().filter(
                content_type=content_type, object_id=self.id
            )
            for sc_ct in subclasses_content_types
        ]
        return [soo for soo in source_of_objects if soo.exists()]

    @property
    def source_of(self):
        """Returns a list of the __str__-representations of the objects that have been created from this geometry.
        I.e. the geometry objects for which this geometry is the source.
        """

        return [source.__str__() for qs in self.source_of_objects() for source in qs]


class DrawnByGeometry(models.Model):
    drawn_by = models.IntegerField(blank=True, null=True)  # EmailUserRO
    last_updated_by = models.IntegerField(blank=True, null=True)  # EmailUserRO

    class Meta:
        abstract = True


class OccurrenceReportGeometry(GeometryBase, DrawnByGeometry):
    BULK_IMPORT_ABBREVIATION = "ocrgeo"

    occurrence_report = models.ForeignKey(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="ocr_geometry",
    )
    locked = models.BooleanField(default=False)
    show_on_map = models.BooleanField(default=True)

    color = ColorField(blank=True, null=True)
    stroke = ColorField(blank=True, null=True)

    class Meta:
        app_label = "boranga"

    def related_model_field(self):
        return self.occurrence_report

    def save(self, *args, **kwargs):
        if (
            self.occurrence_report.group_type.name == GroupType.GROUP_TYPE_FAUNA
            and type(self.geometry).__name__ in ["Polygon", "MultiPolygon"]
        ):
            raise ValidationError("Fauna occurrence reports cannot have polygons")

        super().save(*args, **kwargs)


class ObserverRole(AbstractOrderedList):
    class Meta:
        ordering = ["order"]
        app_label = "boranga"


class ObserverCategory(AbstractOrderedList):
    class Meta:
        ordering = ["order"]
        app_label = "boranga"
        verbose_name_plural = "Observer Categories"


class OCRObserverDetail(RevisionedMixin):
    BULK_IMPORT_ABBREVIATION = "ocrcon"

    """
    Observer data  for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.ForeignKey(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="observer_detail",
    )
    observer_name = models.CharField(max_length=250, blank=True, null=True)
    role = models.ForeignKey(
        ObserverRole, on_delete=models.PROTECT, null=True, blank=True
    )
    category = models.ForeignKey(
        ObserverCategory, on_delete=models.PROTECT, null=True, blank=True
    )
    contact = models.TextField(max_length=250, blank=True, null=True)
    organisation = models.CharField(max_length=250, blank=True, null=True)
    main_observer = models.BooleanField(null=True, blank=True)
    visible = models.BooleanField(default=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCRObserver Detail: {self.id} for Occurrence Report: {self.occurrence_report}"


# Is used in HabitatComposition for multiple selection
class LandForm(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - HabitatComposition

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Land Form"
        verbose_name_plural = "Land Forms"

    def __str__(self):
        return str(self.name)


class RockType(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - HabitatComposition

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Rock Type"
        verbose_name_plural = "Rock Types"

    def __str__(self):
        return str(self.name)


class SoilType(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - HabitatComposition

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Soil Type"
        verbose_name_plural = "Soil Types"

    def __str__(self):
        return str(self.name)


class SoilColour(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - HabitatComposition

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Soil Colour"
        verbose_name_plural = "Soil Colours"

    def __str__(self):
        return str(self.name)


class Drainage(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - HabitatComposition

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Drainage"
        verbose_name_plural = "Drainages"

    def __str__(self):
        return str(self.name)


class SoilCondition(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - HabitatComposition

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Soil Condition"
        verbose_name_plural = "Soil Conditions"

    def __str__(self):
        return str(self.name)


class OCRHabitatComposition(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrhab"
    """
    Habitat data  for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="habitat_composition",
    )
    # TODO: Consider fixing these to use a function that returns the choices
    # as setting them in the __init__ method creates issues in other parts of the application
    land_form = MultiSelectField(max_length=250, blank=True, choices=[], null=True)
    rock_type = models.ForeignKey(
        RockType, on_delete=models.SET_NULL, null=True, blank=True
    )
    loose_rock_percent = models.IntegerField(
        null=True, blank=True, validators=[MinValueValidator(1), MaxValueValidator(100)]
    )
    soil_type = MultiSelectField(max_length=250, blank=True, choices=[], null=True)
    soil_colour = models.ForeignKey(
        SoilColour, on_delete=models.SET_NULL, null=True, blank=True
    )
    soil_condition = models.ForeignKey(
        SoilCondition, on_delete=models.SET_NULL, null=True, blank=True
    )
    drainage = models.ForeignKey(
        Drainage, on_delete=models.SET_NULL, null=True, blank=True
    )
    water_quality = models.CharField(max_length=500, null=True, blank=True)
    habitat_notes = models.CharField(max_length=1000, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCRHabitat Composition: {self.id} for Occurrence Report: {self.occurrence_report}"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._meta.get_field("land_form").choices = tuple(
            LandForm.objects.values_list("id", "name")
        )
        self._meta.get_field("soil_type").choices = tuple(
            SoilType.objects.values_list("id", "name")
        )


class OCRHabitatCondition(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrhq"
    """
    Habitat Condition data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="habitat_condition",
    )
    pristine = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    excellent = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    very_good = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    good = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    degraded = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    completely_degraded = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    count_date = models.DateTimeField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCRHabitat Condition: {self.id} for Occurrence Report: {self.occurrence_report}"


class OCRVegetationStructure(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrveg"

    """
    Vegetation Structure data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="vegetation_structure",
    )

    vegetation_structure_layer_one = models.TextField(null=True, blank=True)
    vegetation_structure_layer_two = models.TextField(null=True, blank=True)
    vegetation_structure_layer_three = models.TextField(null=True, blank=True)
    vegetation_structure_layer_four = models.TextField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCR Vegetation Structure: {self.id} for Occurrence Report: {self.occurrence_report}"


class Intensity(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - FireHistory

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Intensity"
        verbose_name_plural = "Intensities"

    def __str__(self):
        return str(self.name)


class OCRFireHistory(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrfh"

    """
    Fire History data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="fire_history",
    )
    last_fire_estimate = models.DateField(null=True, blank=True)
    intensity = models.ForeignKey(
        Intensity, on_delete=models.SET_NULL, null=True, blank=True
    )
    comment = models.CharField(max_length=1000, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCR Fire History: {self.id} for Occurrence Report: {self.occurrence_report}"


class OCRAssociatedSpecies(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrspe"

    """
    Associated Species data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="associated_species",
    )
    comment = models.TextField(blank=True)

    related_species = models.ManyToManyField(Taxonomy, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCR Associated Species {self.id} for Occurrence Report {self.occurrence_report}"


class ObservationMethod(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - ObservationDetail

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Observation Method"
        verbose_name_plural = "Observation Methods"

    def __str__(self):
        return str(self.name)


class OCRObservationDetail(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrobs"

    """
    Observation Details data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="observation_detail",
    )
    observation_method = models.ForeignKey(
        ObservationMethod, on_delete=models.SET_NULL, null=True, blank=True
    )
    area_surveyed = models.IntegerField(null=True, blank=True, default=0)
    survey_duration = models.IntegerField(null=True, blank=True, default=0)
    comments = models.TextField(blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCR Observation Detail: {self.id} for Occurrence Report: {self.occurrence_report}"


class PlantCountMethod(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - PlantCount

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Plant Count Method"
        verbose_name_plural = "Plant Count Methods"

    def __str__(self):
        return str(self.name)


class PlantCountAccuracy(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - PlantCount

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Plant Count Accuracy"
        verbose_name_plural = "Plant Count Accuracies"

    def __str__(self):
        return str(self.name)


class CountedSubject(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()

    """
    # Admin List

    Used by:
    - PlantCount

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Counted Subject"
        verbose_name_plural = "Counted Subjects"

    def __str__(self):
        return str(self.name)


class PlantCondition(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - PlantCount

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Plant Condition"
        verbose_name_plural = "Plant Conditions"

    def __str__(self):
        return str(self.name)


class OCRPlantCount(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrnum"

    """
    Plant Count data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="plant_count",
    )
    plant_count_method = models.ForeignKey(
        PlantCountMethod, on_delete=models.SET_NULL, null=True, blank=True
    )
    plant_count_accuracy = models.ForeignKey(
        PlantCountAccuracy, on_delete=models.SET_NULL, null=True, blank=True
    )
    counted_subject = models.ForeignKey(
        CountedSubject, on_delete=models.SET_NULL, null=True, blank=True
    )
    plant_condition = models.ForeignKey(
        PlantCondition, on_delete=models.SET_NULL, null=True, blank=True
    )
    estimated_population_area = models.IntegerField(null=True, blank=True, default=0)

    count_status = models.CharField(
        choices=settings.COUNT_STATUS_CHOICES,
        max_length=20,
        default=settings.COUNT_STATUS_NOT_COUNTED,
    )

    detailed_alive_mature = models.IntegerField(null=True, blank=True)
    detailed_dead_mature = models.IntegerField(null=True, blank=True)
    detailed_alive_juvenile = models.IntegerField(null=True, blank=True)
    detailed_dead_juvenile = models.IntegerField(null=True, blank=True)
    detailed_alive_seedling = models.IntegerField(null=True, blank=True)
    detailed_dead_seedling = models.IntegerField(null=True, blank=True)
    detailed_alive_unknown = models.IntegerField(null=True, blank=True)
    detailed_dead_unknown = models.IntegerField(null=True, blank=True)

    simple_alive = models.IntegerField(null=True, blank=True)
    simple_dead = models.IntegerField(null=True, blank=True)

    quadrats_present = models.BooleanField(null=True, blank=True)
    quadrats_data_attached = models.BooleanField(null=True, blank=True)
    quadrats_surveyed = models.IntegerField(null=True, blank=True, default=0)
    individual_quadrat_area = models.IntegerField(null=True, blank=True, default=0)
    total_quadrat_area = models.IntegerField(null=True, blank=True, default=0)
    flowering_plants_per = models.IntegerField(
        null=True,
        blank=True,
        default=0,
        validators=[MinValueValidator(0), MaxValueValidator(100)],
    )

    clonal_reproduction_present = models.BooleanField(null=True, blank=True)
    vegetative_state_present = models.BooleanField(null=True, blank=True)
    flower_bud_present = models.BooleanField(null=True, blank=True)
    flower_present = models.BooleanField(null=True, blank=True)
    immature_fruit_present = models.BooleanField(null=True, blank=True)
    ripe_fruit_present = models.BooleanField(null=True, blank=True)
    dehisced_fruit_present = models.BooleanField(null=True, blank=True)
    pollinator_observation = models.CharField(max_length=1000, null=True, blank=True)
    comment = models.CharField(max_length=1000, null=True, blank=True)
    count_date = models.DateTimeField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def save(self, *args, **kwargs):
        # Set fields to None based on count status field
        if self.count_status == settings.COUNT_STATUS_NOT_COUNTED:
            self.detailed_alive_mature = None
            self.detailed_dead_mature = None
            self.detailed_alive_juvenile = None
            self.detailed_dead_juvenile = None
            self.detailed_alive_seedling = None
            self.detailed_dead_seedling = None
            self.detailed_alive_unknown = None
            self.detailed_dead_unknown = None

            self.simple_alive = None
            self.simple_dead = None
        elif self.count_status == settings.COUNT_STATUS_SIMPLE_COUNT:
            self.detailed_alive_mature = None
            self.detailed_dead_mature = None
            self.detailed_alive_juvenile = None
            self.detailed_dead_juvenile = None
            self.detailed_alive_seedling = None
            self.detailed_dead_seedling = None
            self.detailed_alive_unknown = None
            self.detailed_dead_unknown = None
        elif self.count_status == settings.COUNT_STATUS_COUNTED:
            self.simple_alive = None
            self.simple_dead = None

        super().save(*args, **kwargs)

    def __str__(self):
        return f"OCR Plant Count: {self.id} for Occurrence Report: {self.occurrence_report}"


# used for Animal Observation(MultipleSelect)
class PrimaryDetectionMethod(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - AnimalObservation (MultipleSelect)

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"

    def __str__(self):
        return str(self.name)


# used for Animal Observation(MultipleSelect)
class ReproductiveState(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - AnimalObservation (MultipleSelect)

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Reproductive State"
        verbose_name_plural = "Reproductive States"

    def __str__(self):
        return str(self.name)


class AnimalHealth(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - AnimalObservation

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Animal Health"
        verbose_name_plural = "Animal Health"

    def __str__(self):
        return str(self.name)


class DeathReason(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - AnimalObservation

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"

    def __str__(self):
        return str(self.name)


# sed for Animal Observation(MultipleSelect)
class SecondarySign(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - AnimalObservation (MultipleSelect)

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"

    def __str__(self):
        return str(self.name)


class OCRAnimalObservation(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrnum"

    """
    Animal Observation data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="animal_observation",
    )
    primary_detection_method = MultiSelectField(
        max_length=250, blank=True, choices=[], null=True
    )
    reproductive_state = models.ForeignKey(
        ReproductiveState, on_delete=models.SET_NULL, null=True, blank=True
    )
    animal_health = models.ForeignKey(
        AnimalHealth, on_delete=models.SET_NULL, null=True, blank=True
    )
    death_reason = models.ForeignKey(
        DeathReason, on_delete=models.SET_NULL, null=True, blank=True
    )
    secondary_sign = models.ForeignKey(
        SecondarySign, on_delete=models.SET_NULL, null=True, blank=True
    )

    distinctive_feature = models.CharField(max_length=1000, null=True, blank=True)
    action_taken = models.CharField(max_length=1000, null=True, blank=True)
    action_required = models.CharField(max_length=1000, null=True, blank=True)
    observation_detail_comment = models.CharField(
        max_length=1000, null=True, blank=True
    )

    count_status = models.CharField(
        choices=settings.COUNT_STATUS_CHOICES,
        max_length=20,
        default=settings.COUNT_STATUS_NOT_COUNTED,
    )

    alive_adult_male = models.IntegerField(null=True, blank=True, default=0)
    dead_adult_male = models.IntegerField(null=True, blank=True, default=0)
    alive_adult_female = models.IntegerField(null=True, blank=True, default=0)
    dead_adult_female = models.IntegerField(null=True, blank=True, default=0)
    alive_adult_unknown = models.IntegerField(null=True, blank=True, default=0)
    dead_adult_unknown = models.IntegerField(null=True, blank=True, default=0)

    alive_juvenile_male = models.IntegerField(null=True, blank=True, default=0)
    dead_juvenile_male = models.IntegerField(null=True, blank=True, default=0)
    alive_juvenile_female = models.IntegerField(null=True, blank=True, default=0)
    dead_juvenile_female = models.IntegerField(null=True, blank=True, default=0)
    alive_juvenile_unknown = models.IntegerField(null=True, blank=True, default=0)
    dead_juvenile_unknown = models.IntegerField(null=True, blank=True, default=0)

    alive_unsure_male = models.IntegerField(null=True, blank=True, default=0)
    dead_unsure_male = models.IntegerField(null=True, blank=True, default=0)
    alive_unsure_female = models.IntegerField(null=True, blank=True, default=0)
    dead_unsure_female = models.IntegerField(null=True, blank=True, default=0)
    alive_unsure_unknown = models.IntegerField(null=True, blank=True, default=0)
    dead_unsure_unknown = models.IntegerField(null=True, blank=True, default=0)

    simple_alive = models.IntegerField(null=True, blank=True, default=0)
    simple_dead = models.IntegerField(null=True, blank=True, default=0)

    count_date = models.DateTimeField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCR Animal Observation: {self.id} for Occurrence Report: {self.occurrence_report}"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._meta.get_field("primary_detection_method").choices = tuple(
            PrimaryDetectionMethod.objects.values_list("id", "name")
        )

    @property
    def total_count(self):
        state = ["alive", "dead"]
        sex = ["male", "female", "unknown"]
        age = ["adult", "juvenile", "unsure"]
        total = 0
        for st in state:
            for a in age:
                for s in sex:
                    value = getattr(self, f"{st}_{a}_{s}")
                    if value:
                        total += value
        return total


class IdentificationCertainty(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List
    May be a mandatory field that assessor needs to complete

    Used by:
    - Identification

    """

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Identification Certainty"
        verbose_name_plural = "Identification Certainties"

    def __str__(self):
        return str(self.name)


class SampleType(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - Identification

    """

    name = models.CharField(
        max_length=250, blank=False, null=False, validators=[no_commas_validator]
    )
    group_type = models.ForeignKey(
        GroupType, on_delete=models.SET_NULL, null=True, blank=True
    )
    order_with_respect_to = "group_type"

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        ordering = ["group_type", "order"]

    def __str__(self):
        return str(self.name)


class SampleDestination(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - Identification

    """

    name = models.CharField(
        max_length=250, blank=False, null=False, validators=[no_commas_validator]
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"

    def __str__(self):
        return str(self.name)


class PermitType(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - Identification

    """

    name = models.CharField(
        max_length=250, blank=False, null=False, validators=[no_commas_validator]
    )
    group_type = models.ForeignKey(
        GroupType, on_delete=models.SET_NULL, null=True, blank=True
    )
    order_with_respect_to = "group_type"

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        ordering = ["group_type", "order"]

    def __str__(self):
        return str(self.name)


class OCRIdentification(models.Model):
    BULK_IMPORT_ABBREVIATION = "ocrid"

    """
    Identification data for occurrence report

    Used for:
    - Occurrence Report
    Is:
    - Table
    """

    occurrence_report = models.OneToOneField(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        related_name="identification",
    )
    id_confirmed_by = models.CharField(max_length=1000, null=True, blank=True)
    identification_certainty = models.ForeignKey(
        IdentificationCertainty, on_delete=models.SET_NULL, null=True, blank=True
    )
    sample_type = models.ForeignKey(
        SampleType, on_delete=models.SET_NULL, null=True, blank=True
    )
    sample_destination = models.ForeignKey(
        SampleDestination, on_delete=models.SET_NULL, null=True, blank=True
    )
    permit_type = models.ForeignKey(
        PermitType, on_delete=models.SET_NULL, null=True, blank=True
    )
    permit_id = models.CharField(max_length=500, null=True, blank=True)
    collector_number = models.CharField(max_length=500, null=True, blank=True)
    barcode_number = models.CharField(max_length=500, null=True, blank=True)
    identification_comment = models.TextField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCR Identification: {self.id} for Occurrence Report: {self.occurrence_report}"


class OccurrenceReportDocument(Document):
    BULK_IMPORT_ABBREVIATION = "ocrdoc"

    document_number = models.CharField(max_length=9, blank=True, default="")
    occurrence_report = models.ForeignKey(
        "OccurrenceReport", related_name="documents", on_delete=models.CASCADE
    )
    _file = models.FileField(
        upload_to=update_occurrence_report_doc_filename,
        max_length=512,
        storage=private_storage,
    )
    input_name = models.CharField(max_length=255, null=True, blank=True)
    can_submitter_access = models.BooleanField(default=False)
    document_category = models.ForeignKey(
        DocumentCategory, null=True, blank=True, on_delete=models.SET_NULL
    )
    document_sub_category = models.ForeignKey(
        DocumentSubCategory, null=True, blank=True, on_delete=models.SET_NULL
    )
    uploaded_by = models.IntegerField(null=True)  # EmailUserRO

    class Meta:
        app_label = "boranga"
        verbose_name = "Occurrence Report Document"

    def save(self, *args, **kwargs):
        # Prefix "D" char to document_number.
        if self.document_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(no_revision=True, force_insert=force_insert)
            new_document_id = f"D{str(self.pk)}"
            self.document_number = new_document_id
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)

    def get_parent_instance(self) -> models.Model:
        return self.occurrence_report

    @transaction.atomic
    def add_documents(self, request, *args, **kwargs):
        # save the files
        data = json.loads(request.data.get("data"))
        # if not data.get('update'):
        #     documents_qs = self.filter(input_name='species_doc', visible=True)
        #     documents_qs.delete()
        for idx in range(data["num_files"]):
            self.check_file(request.data.get("file-" + str(idx)))
            _file = request.data.get("file-" + str(idx))
            self._file = _file
            self.name = _file.name
            self.input_name = data["input_name"]
            self.save(no_revision=True)
        # end save documents
        self.save(*args, **kwargs)


class ShapefileDocumentQueryset(models.QuerySet):
    """Using a custom manager to make sure shapfiles are removed when a bulk .delete is called
    as having multiple files with the shapefile extensions in the same folder causes issues.
    """

    def delete(self):
        for obj in self:
            obj._file.delete()
        super().delete()


class OccurrenceReportShapefileDocument(Document):
    objects = ShapefileDocumentQueryset.as_manager()
    occurrence_report = models.ForeignKey(
        "OccurrenceReport", related_name="shapefile_documents", on_delete=models.CASCADE
    )
    _file = models.FileField(
        upload_to=update_occurrence_report_doc_filename,
        max_length=512,
        storage=private_storage,
    )
    input_name = models.CharField(max_length=255, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def get_parent_instance(self) -> models.Model:
        return self.occurrence_report

    def delete(self, *args, **kwargs):
        # By pass the custom delete method in super and just do a regular delete
        models.Model.delete(self, *args, **kwargs)


class OCRConservationThreat(RevisionedMixin):
    BULK_IMPORT_ABBREVIATION = "ocrthr"
    """
    Threat for a occurrence_report in a particular location.

    NB: Maybe make many to many

    Has a:
    - occurrence_report
    Used for:
    - OccurrenceReport
    Is:
    - Table
    """

    occurrence_report = models.ForeignKey(
        OccurrenceReport,
        on_delete=models.CASCADE,
        null=True,
        blank=True,
        related_name="ocr_threats",
    )
    threat_number = models.CharField(max_length=9, blank=True, default="")
    threat_category = models.ForeignKey(
        ThreatCategory, on_delete=models.CASCADE, default=None, null=True, blank=True
    )
    threat_agent = models.ForeignKey(
        ThreatAgent, on_delete=models.SET_NULL, default=None, null=True, blank=True
    )
    current_impact = models.ForeignKey(
        CurrentImpact, on_delete=models.SET_NULL, default=None, null=True, blank=True
    )
    potential_impact = models.ForeignKey(
        PotentialImpact, on_delete=models.SET_NULL, default=None, null=True, blank=True
    )
    potential_threat_onset = models.ForeignKey(
        PotentialThreatOnset,
        on_delete=models.SET_NULL,
        default=None,
        null=True,
        blank=True,
    )
    comment = models.CharField(max_length=512, blank=True, null=True)
    date_observed = models.DateField(blank=True, null=True)
    visible = models.BooleanField(
        default=True
    )  # to prevent deletion, hidden and still be available in history

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCRConservationThreat: {self.threat_number} for Occurrence Report: {self.occurrence_report}"

    def save(self, *args, **kwargs):
        if self.threat_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(no_revision=True, force_insert=force_insert)
            new_threat_id = f"T{str(self.pk)}"
            self.threat_number = new_threat_id
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)

    @property
    def source(self):
        return self.occurrence_report.occurrence_report_number


class WildStatus(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()

    name = models.CharField(
        max_length=250,
        blank=False,
        null=False,
        unique=True,
        validators=[no_commas_validator],
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Wild Status"
        verbose_name_plural = "Wild Statuses"

    def __str__(self):
        return str(self.name)


class OccurrenceManager(models.Manager):
    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("group_type", "species", "community")
            .annotate(occurrence_report_count=Count("occurrence_reports"))
        )


class Occurrence(RevisionedMixin):
    BULK_IMPORT_ABBREVIATION = "occ"

    REVIEW_STATUS_CHOICES = (
        ("not_reviewed", "Not Reviewed"),
        ("awaiting_amendments", "Awaiting Amendments"),
        ("amended", "Amended"),
        ("accepted", "Accepted"),
    )

    RELATED_ITEM_CHOICES = [
        ("species", "Species"),
        ("community", "Community"),
        ("occurrence_report", "Occurrence Report"),
    ]

    OCCURRENCE_CHOICE_OCR = "ocr"
    OCCURRENCE_CHOICE_NON_OCR = "non-ocr"
    OCCURRENCE_SOURCE_CHOICES = (
        (OCCURRENCE_CHOICE_OCR, "OCR"),
        (OCCURRENCE_CHOICE_NON_OCR, "Non-OCR (describe in comments)"),
    )

    objects = OccurrenceManager()
    occurrence_number = models.CharField(max_length=9, blank=True, default="")

    # Field to use when importing data from the legacy system
    migrated_from_id = models.CharField(
        max_length=50, blank=True, null=True, unique=True
    )

    occurrence_name = models.CharField(max_length=250, blank=True, null=True)
    group_type = models.ForeignKey(
        GroupType, on_delete=models.PROTECT, null=True, blank=True
    )

    species = models.ForeignKey(
        Species,
        on_delete=models.PROTECT,
        null=True,
        blank=True,
        related_name="occurrences",
    )
    community = models.ForeignKey(
        Community,
        on_delete=models.PROTECT,
        null=True,
        blank=True,
        related_name="occurrences",
    )

    submitter = models.IntegerField(null=True)  # EmailUserRO
    wild_status = models.ForeignKey(
        WildStatus, on_delete=models.PROTECT, null=True, blank=True
    )
    occurrence_source = MultiSelectField(
        max_length=250, blank=True, choices=OCCURRENCE_SOURCE_CHOICES, null=True
    )

    comment = models.TextField(null=True, blank=True)

    review_due_date = models.DateField(null=True, blank=True)
    review_status = models.CharField(
        "Review Status",
        max_length=30,
        choices=REVIEW_STATUS_CHOICES,
        default=REVIEW_STATUS_CHOICES[0][0],
    )

    created_date = models.DateTimeField(auto_now_add=True, null=False, blank=False)
    updated_date = models.DateTimeField(auto_now=True, null=False, blank=False)

    combined_occurrence = models.ForeignKey(
        "self",
        on_delete=models.PROTECT,
        null=True,
        blank=True,
        related_name="combined_occurrences",
    )

    PROCESSING_STATUS_DRAFT = "draft"
    PROCESSING_STATUS_ACTIVE = "active"
    PROCESSING_STATUS_LOCKED = "locked"
    PROCESSING_STATUS_SPLIT = "split"
    PROCESSING_STATUS_COMBINE = "combine"
    PROCESSING_STATUS_HISTORICAL = "historical"
    PROCESSING_STATUS_DISCARDED = "discarded"
    PROCESSING_STATUS_CHOICES = (
        (PROCESSING_STATUS_DRAFT, "Draft"),
        (PROCESSING_STATUS_ACTIVE, "Active"),
        (PROCESSING_STATUS_LOCKED, "Locked"),
        (PROCESSING_STATUS_SPLIT, "Split"),
        (PROCESSING_STATUS_COMBINE, "Combine"),
        (PROCESSING_STATUS_HISTORICAL, "Historical"),
        (PROCESSING_STATUS_DISCARDED, "Discarded"),
    )
    processing_status = models.CharField(
        "Processing Status",
        max_length=30,
        choices=PROCESSING_STATUS_CHOICES,
        default=PROCESSING_STATUS_DRAFT,
    )

    class Meta:
        indexes = [
            models.Index(fields=["group_type"]),
            models.Index(fields=["species"]),
            models.Index(fields=["community"]),
        ]
        unique_together = (("occurrence_name", "species", "community"),)

        app_label = "boranga"

    def save(self, *args, **kwargs):
        # Clear the cache
        cache.delete(settings.CACHE_KEY_MAP_OCCURRENCES)
        if self.occurrence_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(no_revision=True, force_insert=force_insert)
            self.occurrence_number = f"OCC{str(self.pk)}"
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)

    def __str__(self):
        if self.species:
            return f"{self.occurrence_number} - {self.species} ({self.group_type})"
        elif self.community:
            return f"{self.occurrence_number} - {self.community} ({self.group_type})"
        else:
            return f"{self.occurrence_number} - {self.group_type}"

    @property
    def number_of_reports(self):
        return self.occurrence_report_count

    @property
    def related_item_identifier(self):
        return self.occurrence_number

    @property
    def related_item_descriptor(self):
        if self.species:
            if self.species.taxonomy and self.species.taxonomy.scientific_name:
                return self.species.taxonomy.scientific_name
        return "Descriptor not available"

    @property
    def related_item_status(self):
        return self.get_processing_status_display()

    @property
    def as_related_item(self):
        related_item = RelatedItem(
            identifier=self.related_item_identifier,
            model_name=self._meta.verbose_name.title(),
            descriptor=self.related_item_descriptor,
            status=self.related_item_status,
            action_url=(
                f'<a href="/internal/occurrence/{self.id}'
                f'?group_type_name={self.group_type.name}" target="_blank">View '
                '<i class="bi bi-box-arrow-up-right"></i></a>'
            ),
        )
        return related_item

    @transaction.atomic
    def combine(self, request):
        # only Active OCCs may be combined to
        if not (self.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE):
            raise ValidationError("Occurrence not Active, cannot be combined to")

        occ_combine_data = json.loads(request.POST.get("data"))

        # OCCs being combined must not be discarded or historical
        combine_occurrences = Occurrence.objects.exclude(id=self.id).filter(
            id__in=occ_combine_data["combine_ids"]
        )

        if not combine_occurrences.exists():
            raise ValidationError("No Occurrences selected to be combined")

        if combine_occurrences.filter(
            Q(processing_status=Occurrence.PROCESSING_STATUS_DISCARDED)
            | Q(processing_status=Occurrence.PROCESSING_STATUS_HISTORICAL)
        ).exists():
            raise ValidationError("Closed or Discarded Occurrences may not be combined")

        # validate species/community
        if combine_occurrences.exclude(group_type=self.group_type).exists():
            raise ValidationError("Selected Occurrence has mismatched group type")

        # dictionary pairing request value keys with corresponding model attrs/foreign relations
        FORM_KEYS = {
            "occurrence_source": "occurrence_source",
            "wild_status": "wild_status",
            "review_due_date": "review_due_date",
            "comment": "comment",
        }
        SECTION_KEYS = {
            "chosen_location_section": "location",
            "chosen_habitat_composition_section": "habitat_composition",
            "chosen_habitat_condition_section": "habitat_condition",
            "chosen_vegetation_structure_section": "vegetation_structure",
            "chosen_fire_history_section": "fire_history",
            "chosen_associated_species_section": "associated_species",
            "chosen_observation_detail_section": "observation_detail",
            "chosen_animal_observation_section": "animal_observation",
            "chosen_plant_count_section": "plant_count",
            "chosen_identification_section": "identification",
        }
        COPY_TABLE_KEYS = {
            "combine_key_contact_ids": OCCContactDetail,
            "combine_document_ids": OccurrenceDocument,
            "combine_site_ids": OccurrenceSite,
        }
        MOVE_TABLE_KEYS = {"combine_threat_ids": OCCConservationThreat}

        # assess and assign form values
        for key in FORM_KEYS:
            if key in occ_combine_data and occ_combine_data[key] != self.id:
                try:  # handle in case somehow the combined occurrence record does not exist
                    setattr(
                        self,
                        key,
                        getattr(combine_occurrences.get(id=occ_combine_data[key]), key),
                    )
                except Exception as e:
                    logger.exception(e)

        # assess and copy section values
        for key in SECTION_KEYS:
            if key in occ_combine_data and occ_combine_data[key] != self.id:
                try:  # handle in case somehow the combined occurrence record does not exist
                    # or does not have the specified section
                    src_section = getattr(
                        combine_occurrences.get(id=occ_combine_data[key]),
                        SECTION_KEYS[key],
                    )
                    section = getattr(self, SECTION_KEYS[key])
                    section_fields = type(section)._meta.get_fields()
                    for i in section_fields:
                        if (
                            i.name != "id"
                            and i.name != "occurrence"
                            and hasattr(section, i.name)
                        ):
                            if isinstance(i, models.ManyToManyField):
                                src_value = getattr(src_section, i.name)
                                value = getattr(section, i.name)
                                value.clear()
                                for i in src_value.all():
                                    value.add(i)
                            else:
                                value = getattr(src_section, i.name)
                                setattr(section, i.name, value)
                    section.save()
                except Exception as e:
                    logger.exception(e)

        # assess and copy table values (contacts, documents, and sites)
        for key in COPY_TABLE_KEYS:
            if key in occ_combine_data:
                for record in (
                    COPY_TABLE_KEYS[key]
                    .objects.filter(id__in=occ_combine_data[key])
                    .exclude(occurrence=self)
                ):
                    copy = clone_model(
                        COPY_TABLE_KEYS[key],
                        COPY_TABLE_KEYS[key],
                        record,
                    )
                    if copy:
                        copy.occurrence = self
                        copy.save()

        # assess and move threat table values
        for key in MOVE_TABLE_KEYS:
            if key in occ_combine_data:
                for record in (
                    MOVE_TABLE_KEYS[key]
                    .objects.filter(id__in=occ_combine_data[key])
                    .exclude(occurrence=self)
                ):
                    record.occurrence = self
                    record.save()

        # special handling is required for tenure records
        # current
        for record in (
            OccurrenceTenure.objects.filter(
                id__in=occ_combine_data["combine_tenure_ids"]
            )
            .filter(status=OccurrenceTenure.STATUS_CURRENT)
            .exclude(occurrence_geometry__occurrence=self)
        ):
            # if current, move by changing the geometry occurrence
            if record.occurrence_geometry:
                occurrence_geometry = record.occurrence_geometry
                occurrence_geometry.occurrence = self
                occurrence_geometry.save()

        # historical
        for record in (
            OccurrenceTenure.objects.filter(
                id__in=occ_combine_data["combine_tenure_ids"]
            )
            .filter(status=OccurrenceTenure.STATUS_HISTORICAL)
            .exclude(historical_occurrence=self.id)
        ):
            # if historical, move by changing historical occurrence
            record.historical_occurrence = self.id
            record.save(override_datetime_updated=True)

        # NOTE: not validating OCR species/community - already validated at OCC level
        # move OCRs
        ocrs = OccurrenceReport.objects.filter(occurrence__in=combine_occurrences)
        ocrs.update(occurrence=self)

        # update combined OCCs to note that they have been combined and close
        for i in combine_occurrences:
            i.processing_status = Occurrence.PROCESSING_STATUS_HISTORICAL
            i.combined_occurrence = self
            i.save(version_user=request.user)

        # save
        self.save(version_user=request.user)

        # action log
        self.log_user_action(
            OccurrenceUserAction.ACTION_COMBINE_OCCURRENCE.format(
                ", ".join(
                    list(
                        combine_occurrences.values_list("occurrence_number", flat=True)
                    )
                ),
                self.occurrence_number,
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_COMBINE_OCCURRENCE.format(
                ", ".join(
                    list(
                        combine_occurrences.values_list("occurrence_number", flat=True)
                    )
                ),
                self.occurrence_number,
            ),
            request,
        )

    def validate_activate(self):
        missing_values = []

        occ_points = self.occ_geometry.annotate(
            geom_type=GeometryType("geometry")
        ).filter(geom_type="POINT")
        occ_boundaries = self.occ_geometry.annotate(
            geom_type=GeometryType("geometry")
        ).filter(geom_type="POLYGON")

        if (
            self.group_type.name
            in [GroupType.GROUP_TYPE_FLORA, GroupType.GROUP_TYPE_COMMUNITY]
            and not self.occurrence_name
        ):
            missing_values.append("Occurrence Name")

        if (
            self.group_type.name
            in [GroupType.GROUP_TYPE_FLORA, GroupType.GROUP_TYPE_COMMUNITY]
            and not occ_boundaries.exists()
        ):
            missing_values.append("Boundary on Map")

        if (
            self.group_type.name == GroupType.GROUP_TYPE_FAUNA
            and not occ_points.exists()
        ):
            missing_values.append("Point on Map")

        if not self.identification or not self.identification.identification_certainty:
            missing_values.append("Identification Certainty")

        if not self.location or not self.location.location_accuracy:
            missing_values.append("Location Accuracy")

        if missing_values:
            raise ValidationError(
                "Cannot activate this occurrence due to missing values: "
                + ", ".join(missing_values)
            )

    @transaction.atomic
    def discard(self, request):
        if not self.processing_status == Occurrence.PROCESSING_STATUS_DRAFT:
            raise exceptions.OccurrenceNotAuthorized()

        self.processing_status = Occurrence.PROCESSING_STATUS_DISCARDED
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_DISCARD_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_DISCARD_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

    @transaction.atomic
    def reinstate(self, request):
        if not self.processing_status == Occurrence.PROCESSING_STATUS_DISCARDED:
            raise exceptions.OccurrenceNotAuthorized()

        self.processing_status = Occurrence.PROCESSING_STATUS_DRAFT
        self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_REINSTATE_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_REINSTATE_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

    def activate(self, request):
        self.validate_activate()
        if (
            is_occurrence_approver(request)
            and self.processing_status == Occurrence.PROCESSING_STATUS_DRAFT
        ):
            self.processing_status = Occurrence.PROCESSING_STATUS_ACTIVE
            self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_ACTIVATE_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_ACTIVATE_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

    def lock(self, request):
        if (
            is_occurrence_approver(request)
            and self.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE
        ):
            self.processing_status = Occurrence.PROCESSING_STATUS_LOCKED
            self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_LOCK_OCCURRENCE.format(self.occurrence_number),
            request,
        )
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_LOCK_OCCURRENCE.format(self.occurrence_number),
            request,
        )

    def unlock(self, request):
        if (
            is_occurrence_approver(request)
            and self.processing_status == Occurrence.PROCESSING_STATUS_LOCKED
        ):
            self.processing_status = Occurrence.PROCESSING_STATUS_ACTIVE
            self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_UNLOCK_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_UNLOCK_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

    def close(self, request):
        if (
            is_occurrence_approver(request)
            and self.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE
        ):
            self.processing_status = Occurrence.PROCESSING_STATUS_HISTORICAL
            self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_CLOSE_OCCURRENCE.format(self.occurrence_number),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_CLOSE_OCCURRENCE.format(self.occurrence_number),
            request,
        )

    def reopen(self, request):
        if (
            is_occurrence_approver(request)
            and self.processing_status == Occurrence.PROCESSING_STATUS_HISTORICAL
        ):
            self.processing_status = Occurrence.PROCESSING_STATUS_ACTIVE
            self.save(version_user=request.user)

        # Log proposal action
        self.log_user_action(
            OccurrenceUserAction.ACTION_REOPEN_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

        # Create a log entry for the user
        request.user.log_user_action(
            OccurrenceUserAction.ACTION_REOPEN_OCCURRENCE.format(
                self.occurrence_number
            ),
            request,
        )

    # if this function is called and the OCC has no associated OCRs, discard it
    def check_ocr_count_for_discard(self, request):
        discardable = [Occurrence.PROCESSING_STATUS_DRAFT]
        if (
            self.processing_status in discardable
            and (is_occurrence_assessor(request) or is_occurrence_approver(request))
            and OccurrenceReport.objects.filter(occurrence=self).count() < 1
        ):
            self.processing_status = Occurrence.PROCESSING_STATUS_DISCARDED
            self.save(version_user=request.user)

    def can_user_edit(self, request):
        user_editable_state = [
            Occurrence.PROCESSING_STATUS_ACTIVE,
            Occurrence.PROCESSING_STATUS_DRAFT,
            Occurrence.PROCESSING_STATUS_DISCARDED,
        ]
        if self.processing_status not in user_editable_state:
            return False

        return is_occurrence_approver(request) or request.user.is_superuser

    def can_user_reopen(self, request):
        user_editable_state = [
            Occurrence.PROCESSING_STATUS_HISTORICAL,
        ]
        if self.processing_status not in user_editable_state:
            return False

        return is_occurrence_approver(request) or request.user.is_superuser

    def log_user_action(self, action, request):
        return OccurrenceUserAction.log_action(self, action, request.user.id)

    def get_related_occurrence_reports(self, **kwargs):

        return OccurrenceReport.objects.filter(occurrence=self)

    def get_related_items(self, filter_type, **kwargs):
        return_list = []
        if filter_type == "all":
            related_field_names = [
                "species",
                "community",
                "occurrence_report",
                "conservation_status",
            ]
        elif filter_type == "all_except_parent_species":
            related_field_names = [
                "conservation_status",
                "occurrences",
                "occurrence_report",
            ]
        else:
            related_field_names = [
                filter_type,
            ]
        all_fields = self._meta.get_fields()
        for a_field in all_fields:
            if a_field.name in related_field_names:
                field_objects = []
                if a_field.is_relation:
                    if a_field.many_to_many:
                        field_objects = a_field.related_model.objects.filter(
                            **{a_field.remote_field.name: self}
                        )
                    elif a_field.many_to_one:  # foreign key
                        field_objects = [
                            getattr(self, a_field.name),
                        ]
                    elif a_field.one_to_many:  # reverse foreign key
                        field_objects = a_field.related_model.objects.filter(
                            **{a_field.remote_field.name: self}
                        )
                    elif a_field.one_to_one:
                        if hasattr(self, a_field.name):
                            field_objects = [
                                getattr(self, a_field.name),
                            ]
                for field_object in field_objects:
                    if field_object:
                        related_item = field_object.as_related_item
                        if related_item not in return_list:
                            return_list.append(related_item)

                # Add parent species related items to the list (limited to one degree of separation)
                if a_field.name == "species" and self.species:
                    return_list.extend(self.species.get_related_items("for_occurrence"))

                # Add renamed from / renamed to community related items to the list
                if a_field.name == "community" and self.community:
                    return_list.extend(
                        self.community.get_related_items("for_occurrence")
                    )

        # Remove the occurrence itself from the list if it ended up there
        for item in return_list:
            if (
                item.model_name == "Occurrence"
                and item.identifier == self.occurrence_number
            ):
                return_list.remove(item)

        return return_list

    @classmethod
    @transaction.atomic
    def clone_from_occurrence_report(self, occurrence_report):
        occurrence = Occurrence()

        occurrence.group_type = occurrence_report.group_type

        occurrence.species = occurrence_report.species
        occurrence.community = occurrence_report.community
        occurrence.comment = occurrence_report.comments

        occurrence.save(no_revision=True)

        # Clone all the associated models

        location = clone_model(
            OCRLocation,
            OCCLocation,
            occurrence_report.location,
        )
        if location:
            location.occurrence = occurrence
            location.copied_ocr_location = occurrence_report.location
            location.save()

        habitat_composition = clone_model(
            OCRHabitatComposition,
            OCCHabitatComposition,
            occurrence_report.habitat_composition,
        )
        if habitat_composition:
            habitat_composition.occurrence = occurrence
            habitat_composition.copied_ocr_habitat_composition = (
                occurrence_report.habitat_composition
            )
            habitat_composition.save()

        habitat_condition = clone_model(
            OCRHabitatCondition,
            OCCHabitatCondition,
            occurrence_report.habitat_condition,
        )
        if habitat_condition:
            habitat_condition.occurrence = occurrence
            habitat_condition.copied_ocr_habitat_condition = (
                occurrence_report.habitat_condition
            )
            habitat_condition.save()

        vegetation_structure = clone_model(
            OCRVegetationStructure,
            OCCVegetationStructure,
            occurrence_report.vegetation_structure,
        )
        if vegetation_structure:
            vegetation_structure.occurrence = occurrence
            vegetation_structure.copied_ocr_vegetation_structure = (
                occurrence_report.vegetation_structure
            )
            vegetation_structure.save()

        fire_history = clone_model(
            OCRFireHistory, OCCFireHistory, occurrence_report.fire_history
        )
        if fire_history:
            fire_history.occurrence = occurrence
            fire_history.copied_ocr_fire_history = occurrence_report.fire_history
            fire_history.save()

        associated_species = clone_model(
            OCRAssociatedSpecies,
            OCCAssociatedSpecies,
            occurrence_report.associated_species,
        )
        if associated_species:
            associated_species.occurrence = occurrence
            associated_species.copied_ocr_associated_species = (
                occurrence_report.associated_species
            )
            associated_species.save()
            # copy over related species separately
            for i in occurrence_report.associated_species.related_species.all():
                associated_species.related_species.add(i)

        observation_detail = clone_model(
            OCRObservationDetail,
            OCCObservationDetail,
            occurrence_report.observation_detail,
        )
        if observation_detail:
            observation_detail.occurrence = occurrence
            observation_detail.copied_ocr_observation_detail = (
                occurrence_report.observation_detail
            )
            observation_detail.save()

        plant_count = clone_model(
            OCRPlantCount, OCCPlantCount, occurrence_report.plant_count
        )
        if plant_count:
            plant_count.occurrence = occurrence
            plant_count.copied_ocr_plant_count = occurrence_report.plant_count
            plant_count.save()

        animal_observation = clone_model(
            OCRAnimalObservation,
            OCCAnimalObservation,
            occurrence_report.animal_observation,
        )
        if animal_observation:
            animal_observation.occurrence = occurrence
            animal_observation.copied_ocr_animal_observation = (
                occurrence_report.animal_observation
            )
            animal_observation.save()

        identification = clone_model(
            OCRIdentification, OCCIdentification, occurrence_report.identification
        )
        if identification:
            identification.occurrence = occurrence
            identification.copied_ocr_identification = occurrence_report.identification
            identification.save()

        # Clone the threats
        for threat in occurrence_report.ocr_threats.all():
            occ_threat = clone_model(
                OCRConservationThreat, OCCConservationThreat, threat
            )
            if occ_threat:
                occ_threat.occurrence = occurrence
                occ_threat.occurrence_report_threat = threat
                occ_threat.save()

        # Clone the documents
        for doc in occurrence_report.documents.all():
            occ_doc = clone_model(OccurrenceReportDocument, OccurrenceDocument, doc)
            if occ_doc:
                occ_doc.occurrence = occurrence
                occ_doc.save()

        return occurrence


class OccurrenceLogEntry(CommunicationsLogEntry):
    occurrence = models.ForeignKey(
        Occurrence, related_name="comms_logs", on_delete=models.CASCADE
    )

    def __str__(self):
        return f"{self.reference} - {self.subject}"

    class Meta:
        app_label = "boranga"

    def save(self, **kwargs):
        # save the occurrence number as the reference if the reference not provided
        if not self.reference:
            self.reference = self.occurrence.occurrence_number
        super().save(**kwargs)


def update_occurrence_comms_log_filename(instance, filename):
    return "{}/occurrence/{}/communications/{}".format(
        settings.MEDIA_APP_DIR, instance.log_entry.occurrence.id, filename
    )


class OccurrenceLogDocument(Document):
    log_entry = models.ForeignKey(
        OccurrenceLogEntry, related_name="documents", on_delete=models.CASCADE
    )
    _file = models.FileField(
        upload_to=update_occurrence_comms_log_filename,
        max_length=512,
        storage=private_storage,
    )

    class Meta:
        app_label = "boranga"

    def get_parent_instance(self) -> models.Model:
        return self.log_entry


class OccurrenceUserAction(UserAction):
    ACTION_VIEW_OCCURRENCE = "View occurrence {}"
    ACTION_SAVE_OCCURRENCE = "Save occurrence {}"
    ACTION_EDIT_OCCURRENCE = "Edit occurrence {}"
    ACTION_DISCARD_OCCURRENCE = "Discard occurrence {}"
    ACTION_REINSTATE_OCCURRENCE = "Reinstate occurrence {}"
    ACTION_COMBINE_OCCURRENCE = "{} combined in to occurrence {}"
    ACTION_ACTIVATE_OCCURRENCE = "Activate occurrence {}"
    ACTION_LOCK_OCCURRENCE = "Lock occurrence {}"
    ACTION_UNLOCK_OCCURRENCE = "Unlock occurrence {}"
    ACTION_CLOSE_OCCURRENCE = "Close occurrence {}"
    ACTION_REOPEN_OCCURRENCE = "Reopen occurrence {}"

    # Document
    ACTION_ADD_DOCUMENT = "Document {} added for occurrence {}"
    ACTION_UPDATE_DOCUMENT = "Document {} updated for occurrence {}"
    ACTION_DISCARD_DOCUMENT = "Document {} discarded for occurrence {}"
    ACTION_REINSTATE_DOCUMENT = "Document {} reinstated for occurrence {}"

    # Threat
    ACTION_ADD_THREAT = "Threat {} added for occurrence {}"
    ACTION_UPDATE_THREAT = "Threat {} updated for occurrence {}"
    ACTION_DISCARD_THREAT = "Threat {} discarded for occurrence {}"
    ACTION_REINSTATE_THREAT = "Threat {} reinstated for occurrence {}"

    class Meta:
        app_label = "boranga"
        ordering = ("-when",)

    @classmethod
    def log_action(cls, occurrence, action, user):
        return cls.objects.create(occurrence=occurrence, who=user, what=str(action))

    occurrence = models.ForeignKey(
        Occurrence, related_name="action_logs", on_delete=models.CASCADE
    )


class OccurrenceDocument(Document):
    document_number = models.CharField(max_length=9, blank=True, default="")
    occurrence = models.ForeignKey(
        "Occurrence", related_name="documents", on_delete=models.CASCADE
    )
    _file = models.FileField(
        upload_to=update_occurrence_doc_filename,
        max_length=512,
        storage=private_storage,
    )
    input_name = models.CharField(max_length=255, null=True, blank=True)
    document_category = models.ForeignKey(
        DocumentCategory, null=True, blank=True, on_delete=models.SET_NULL
    )
    document_sub_category = models.ForeignKey(
        DocumentSubCategory, null=True, blank=True, on_delete=models.SET_NULL
    )
    uploaded_by = models.IntegerField(null=True)  # EmailUserRO

    class Meta:
        app_label = "boranga"
        verbose_name = "Occurrence Document"

    def save(self, *args, **kwargs):
        # Prefix "D" char to document_number.
        if self.document_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(no_revision=True, force_insert=force_insert)
            new_document_id = f"D{str(self.pk)}"
            self.document_number = new_document_id
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)

    def get_parent_instance(self) -> models.Model:
        return self.occurrence

    @transaction.atomic
    def add_documents(self, request, *args, **kwargs):
        # save the files
        data = json.loads(request.data.get("data"))
        # if not data.get('update'):
        #     documents_qs = self.filter(input_name='species_doc', visible=True)
        #     documents_qs.delete()
        for idx in range(data["num_files"]):
            self.check_file(request.data.get("file-" + str(idx)))
            _file = request.data.get("file-" + str(idx))
            self._file = _file
            self.name = _file.name
            self.input_name = data["input_name"]
            self.save(no_revision=True)
        # end save documents
        self.save(*args, **kwargs)


class OCCLocation(models.Model):
    """
    Location data  for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence, on_delete=models.CASCADE, null=True, related_name="location"
    )
    copied_ocr_location = models.ForeignKey(
        OCRLocation, on_delete=models.SET_NULL, null=True, blank=True
    )
    location_description = models.TextField(null=True, blank=True)
    boundary_description = models.TextField(null=True, blank=True)
    mapped_boundary = models.BooleanField(null=True, blank=True)
    buffer_radius = models.IntegerField(null=True, blank=True, default=0)
    datum = models.ForeignKey(Datum, on_delete=models.SET_NULL, null=True, blank=True)
    epsg_code = models.IntegerField(null=False, blank=False, default=4326)
    coordinate_source = models.ForeignKey(
        CoordinateSource, on_delete=models.SET_NULL, null=True, blank=True
    )
    location_accuracy = models.ForeignKey(
        LocationAccuracy, on_delete=models.SET_NULL, null=True, blank=True
    )

    region = models.ForeignKey(
        Region, default=None, on_delete=models.CASCADE, null=True, blank=True
    )
    district = models.ForeignKey(
        District, default=None, on_delete=models.CASCADE, null=True, blank=True
    )
    locality = models.TextField(default=None, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCCLocation: {self.id} for Occurrence: {self.occurrence}"


class GeometryType(Func):
    function = "GeometryType"
    output_field = CharField()


class OccurrenceGeometry(GeometryBase, DrawnByGeometry):
    occurrence = models.ForeignKey(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="occ_geometry",
    )
    locked = models.BooleanField(default=False)
    buffer_radius = models.FloatField(null=True, blank=True, default=0)

    color = ColorField(default="#3333FF")  # Light blue
    stroke = ColorField(default="#0033CC")  # Dark blue
    opacity = models.FloatField(
        default=0.5, validators=[MinValueValidator(0.0), MaxValueValidator(1.0)]
    )  # Used for map layer opacity

    class Meta:
        app_label = "boranga"

    def related_model_field(self):
        return self.occurrence

    def save(self, *args, **kwargs):
        if self.occurrence.group_type.name == GroupType.GROUP_TYPE_FAUNA and type(
            self.geometry
        ).__name__ in ["Polygon", "MultiPolygon"]:
            raise ValidationError("Fauna occurrences cannot have polygons")

        super().save(*args, **kwargs)


class OCCContactDetail(RevisionedMixin):
    """
    Observer data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.ForeignKey(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="contact_detail",
    )
    contact_name = models.CharField(max_length=250, blank=True, null=True)
    role = models.CharField(max_length=250, blank=True, null=True)
    contact = models.CharField(max_length=250, blank=True, null=True)
    organisation = models.CharField(max_length=250, blank=True, null=True)
    notes = models.CharField(max_length=512, blank=True, null=True)
    visible = models.BooleanField(default=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCCContactDetail {self.id} for Occurrence: {self.occurrence}"


class OCCConservationThreat(RevisionedMixin):
    """
    Threat for an occurrence in a particular location.

    NB: Maybe make many to many

    Has a:
    - occurrence
    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.ForeignKey(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        blank=True,
        related_name="occ_threats",
    )

    # original ocr, if any
    occurrence_report_threat = models.ForeignKey(
        OCRConservationThreat,
        on_delete=models.CASCADE,
        null=True,
        blank=True,
        related_name="original_report_threat",
    )

    threat_number = models.CharField(max_length=9, blank=True, default="")
    threat_category = models.ForeignKey(
        ThreatCategory, on_delete=models.CASCADE, default=None, null=True, blank=True
    )
    threat_agent = models.ForeignKey(
        ThreatAgent, on_delete=models.SET_NULL, default=None, null=True, blank=True
    )
    current_impact = models.ForeignKey(
        CurrentImpact, on_delete=models.SET_NULL, default=None, null=True, blank=True
    )
    potential_impact = models.ForeignKey(
        PotentialImpact, on_delete=models.SET_NULL, default=None, null=True, blank=True
    )
    potential_threat_onset = models.ForeignKey(
        PotentialThreatOnset,
        on_delete=models.SET_NULL,
        default=None,
        null=True,
        blank=True,
    )
    comment = models.CharField(max_length=512, blank=True, null=True)
    date_observed = models.DateField(blank=True, null=True)
    visible = models.BooleanField(
        default=True
    )  # to prevent deletion, hidden and still be available in history

    class Meta:
        app_label = "boranga"
        unique_together = (
            "occurrence",
            "occurrence_report_threat",
        )

    def __str__(self):
        return f"OCCConservationThreat {self.id} for Occurrence: {self.occurrence}"

    def save(self, *args, **kwargs):
        if self.threat_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(no_revision=True, force_insert=force_insert)
            new_threat_id = f"T{str(self.pk)}"
            self.threat_number = new_threat_id
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)

    @property
    def source(self):
        if self.occurrence_report_threat:
            return (
                self.occurrence_report_threat.occurrence_report.occurrence_report_number
            )
        return self.occurrence.occurrence_number


class OCCHabitatComposition(models.Model):
    """
    Habitat data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="habitat_composition",
    )
    copied_ocr_habitat_composition = models.ForeignKey(
        OCRHabitatComposition, on_delete=models.SET_NULL, null=True, blank=True
    )
    land_form = MultiSelectField(max_length=250, blank=True, choices=[], null=True)
    rock_type = models.ForeignKey(
        RockType, on_delete=models.SET_NULL, null=True, blank=True
    )
    loose_rock_percent = models.IntegerField(
        null=True, blank=True, validators=[MinValueValidator(1), MaxValueValidator(100)]
    )
    soil_type = MultiSelectField(max_length=250, blank=True, choices=[], null=True)
    soil_colour = models.ForeignKey(
        SoilColour, on_delete=models.SET_NULL, null=True, blank=True
    )
    soil_condition = models.ForeignKey(
        SoilCondition, on_delete=models.SET_NULL, null=True, blank=True
    )
    drainage = models.ForeignKey(
        Drainage, on_delete=models.SET_NULL, null=True, blank=True
    )
    water_quality = models.CharField(max_length=500, null=True, blank=True)
    habitat_notes = models.CharField(max_length=1000, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return f"OCCHabitatComposition {self.id} for Occurrence: {self.occurrence}"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._meta.get_field("land_form").choices = tuple(
            LandForm.objects.values_list("id", "name")
        )
        self._meta.get_field("soil_type").choices = tuple(
            SoilType.objects.values_list("id", "name")
        )


class OCCHabitatCondition(models.Model):
    """
    Habitat Condition data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="habitat_condition",
    )
    copied_ocr_habitat_condition = models.ForeignKey(
        OCRHabitatCondition, on_delete=models.SET_NULL, null=True, blank=True
    )
    pristine = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    excellent = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    very_good = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    good = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    degraded = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    completely_degraded = models.DecimalField(
        null=True,
        blank=True,
        max_digits=5,
        decimal_places=2,
        default=Decimal("0.00"),
        validators=[
            MinValueValidator(Decimal("0.00")),
            MaxValueValidator(Decimal("100.00")),
        ],
    )
    count_date = models.DateTimeField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)


class OCCVegetationStructure(models.Model):
    """
    Vegetation Structure data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="vegetation_structure",
    )
    copied_ocr_vegetation_structure = models.ForeignKey(
        OCRVegetationStructure, on_delete=models.SET_NULL, null=True, blank=True
    )
    vegetation_structure_layer_one = models.TextField(null=True, blank=True)
    vegetation_structure_layer_two = models.TextField(null=True, blank=True)
    vegetation_structure_layer_three = models.TextField(null=True, blank=True)
    vegetation_structure_layer_four = models.TextField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)


class OCCFireHistory(models.Model):
    """
    Fire History data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="fire_history",
    )
    copied_ocr_fire_history = models.ForeignKey(
        OCRFireHistory, on_delete=models.SET_NULL, null=True, blank=True
    )
    last_fire_estimate = models.DateField(null=True, blank=True)
    intensity = models.ForeignKey(
        Intensity, on_delete=models.SET_NULL, null=True, blank=True
    )
    comment = models.CharField(max_length=1000, null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)


class OCCAssociatedSpecies(models.Model):
    """
    Associated Species data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="associated_species",
    )
    copied_ocr_associated_species = models.ForeignKey(
        OCRAssociatedSpecies, on_delete=models.SET_NULL, null=True, blank=True
    )
    comment = models.TextField(blank=True)

    related_species = models.ManyToManyField(Taxonomy, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)


class OCCObservationDetail(models.Model):
    """
    Observation Details data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="observation_detail",
    )
    copied_ocr_observation_detail = models.ForeignKey(
        OCRObservationDetail, on_delete=models.SET_NULL, null=True, blank=True
    )
    observation_method = models.ForeignKey(
        ObservationMethod, on_delete=models.SET_NULL, null=True, blank=True
    )
    area_surveyed = models.IntegerField(null=True, blank=True, default=0)
    survey_duration = models.IntegerField(null=True, blank=True, default=0)
    comments = models.TextField(blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)


class OCCPlantCount(models.Model):
    """
    Plant Count data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="plant_count",
    )
    copied_ocr_plant_count = models.ForeignKey(
        OCRPlantCount, on_delete=models.SET_NULL, null=True, blank=True
    )
    plant_count_method = models.ForeignKey(
        PlantCountMethod, on_delete=models.SET_NULL, null=True, blank=True
    )
    plant_count_accuracy = models.ForeignKey(
        PlantCountAccuracy, on_delete=models.SET_NULL, null=True, blank=True
    )
    counted_subject = models.ForeignKey(
        CountedSubject, on_delete=models.SET_NULL, null=True, blank=True
    )
    plant_condition = models.ForeignKey(
        PlantCondition, on_delete=models.SET_NULL, null=True, blank=True
    )
    estimated_population_area = models.IntegerField(null=True, blank=True, default=0)

    count_status = models.CharField(
        choices=settings.COUNT_STATUS_CHOICES,
        max_length=20,
        default=settings.COUNT_STATUS_NOT_COUNTED,
    )

    detailed_alive_mature = models.IntegerField(null=True, blank=True)
    detailed_dead_mature = models.IntegerField(null=True, blank=True)
    detailed_alive_juvenile = models.IntegerField(null=True, blank=True)
    detailed_dead_juvenile = models.IntegerField(null=True, blank=True)
    detailed_alive_seedling = models.IntegerField(null=True, blank=True)
    detailed_dead_seedling = models.IntegerField(null=True, blank=True)
    detailed_alive_unknown = models.IntegerField(null=True, blank=True)
    detailed_dead_unknown = models.IntegerField(null=True, blank=True)

    simple_alive = models.IntegerField(null=True, blank=True)
    simple_dead = models.IntegerField(null=True, blank=True)

    quadrats_present = models.BooleanField(null=True, blank=True)
    quadrats_data_attached = models.BooleanField(null=True, blank=True)
    quadrats_surveyed = models.IntegerField(null=True, blank=True, default=0)
    individual_quadrat_area = models.IntegerField(null=True, blank=True, default=0)
    total_quadrat_area = models.IntegerField(null=True, blank=True, default=0)
    flowering_plants_per = models.IntegerField(
        null=True,
        blank=True,
        default=0,
        validators=[MinValueValidator(0), MaxValueValidator(100)],
    )

    clonal_reproduction_present = models.BooleanField(null=True, blank=True)
    vegetative_state_present = models.BooleanField(null=True, blank=True)
    flower_bud_present = models.BooleanField(null=True, blank=True)
    flower_present = models.BooleanField(null=True, blank=True)
    immature_fruit_present = models.BooleanField(null=True, blank=True)
    ripe_fruit_present = models.BooleanField(null=True, blank=True)
    dehisced_fruit_present = models.BooleanField(null=True, blank=True)
    pollinator_observation = models.CharField(max_length=1000, null=True, blank=True)
    comment = models.CharField(max_length=1000, null=True, blank=True)
    count_date = models.DateTimeField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)

    def save(self, *args, **kwargs):
        # Set fields to None based on count status field
        if self.count_status == settings.COUNT_STATUS_NOT_COUNTED:
            self.detailed_alive_mature = None
            self.detailed_dead_mature = None
            self.detailed_alive_juvenile = None
            self.detailed_dead_juvenile = None
            self.detailed_alive_seedling = None
            self.detailed_dead_seedling = None
            self.detailed_alive_unknown = None
            self.detailed_dead_unknown = None

            self.simple_alive = None
            self.simple_dead = None
        elif self.count_status == settings.COUNT_STATUS_SIMPLE_COUNT:
            self.detailed_alive_mature = None
            self.detailed_dead_mature = None
            self.detailed_alive_juvenile = None
            self.detailed_dead_juvenile = None
            self.detailed_alive_seedling = None
            self.detailed_dead_seedling = None
            self.detailed_alive_unknown = None
            self.detailed_dead_unknown = None
        elif self.count_status == settings.COUNT_STATUS_COUNTED:
            self.simple_alive = None
            self.simple_dead = None

        super().save(*args, **kwargs)


class OCCAnimalObservation(models.Model):
    """
    Animal Observation data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="animal_observation",
    )
    copied_ocr_animal_observation = models.ForeignKey(
        OCRAnimalObservation, on_delete=models.SET_NULL, null=True, blank=True
    )
    primary_detection_method = MultiSelectField(
        max_length=250, blank=True, choices=[], null=True
    )
    reproductive_state = models.ForeignKey(
        ReproductiveState, on_delete=models.SET_NULL, null=True, blank=True
    )
    animal_health = models.ForeignKey(
        AnimalHealth, on_delete=models.SET_NULL, null=True, blank=True
    )
    death_reason = models.ForeignKey(
        DeathReason, on_delete=models.SET_NULL, null=True, blank=True
    )
    secondary_sign = models.ForeignKey(
        SecondarySign, on_delete=models.SET_NULL, null=True, blank=True
    )

    distinctive_feature = models.CharField(max_length=1000, null=True, blank=True)
    action_taken = models.CharField(max_length=1000, null=True, blank=True)
    action_required = models.CharField(max_length=1000, null=True, blank=True)
    observation_detail_comment = models.CharField(
        max_length=1000, null=True, blank=True
    )

    count_status = models.CharField(
        choices=settings.COUNT_STATUS_CHOICES,
        max_length=20,
        default=settings.COUNT_STATUS_NOT_COUNTED,
    )

    alive_adult_male = models.IntegerField(null=True, blank=True, default=0)
    dead_adult_male = models.IntegerField(null=True, blank=True, default=0)
    alive_adult_female = models.IntegerField(null=True, blank=True, default=0)
    dead_adult_female = models.IntegerField(null=True, blank=True, default=0)
    alive_adult_unknown = models.IntegerField(null=True, blank=True, default=0)
    dead_adult_unknown = models.IntegerField(null=True, blank=True, default=0)

    alive_juvenile_male = models.IntegerField(null=True, blank=True, default=0)
    dead_juvenile_male = models.IntegerField(null=True, blank=True, default=0)
    alive_juvenile_female = models.IntegerField(null=True, blank=True, default=0)
    dead_juvenile_female = models.IntegerField(null=True, blank=True, default=0)
    alive_juvenile_unknown = models.IntegerField(null=True, blank=True, default=0)
    dead_juvenile_unknown = models.IntegerField(null=True, blank=True, default=0)

    alive_unsure_male = models.IntegerField(null=True, blank=True, default=0)
    dead_unsure_male = models.IntegerField(null=True, blank=True, default=0)
    alive_unsure_female = models.IntegerField(null=True, blank=True, default=0)
    dead_unsure_female = models.IntegerField(null=True, blank=True, default=0)
    alive_unsure_unknown = models.IntegerField(null=True, blank=True, default=0)
    dead_unsure_unknown = models.IntegerField(null=True, blank=True, default=0)

    simple_alive = models.IntegerField(null=True, blank=True, default=0)
    simple_dead = models.IntegerField(null=True, blank=True, default=0)

    count_date = models.DateTimeField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._meta.get_field("primary_detection_method").choices = tuple(
            PrimaryDetectionMethod.objects.values_list("id", "name")
        )

    @property
    def total_count(self):
        state = ["alive", "dead"]
        sex = ["male", "female", "unknown"]
        age = ["adult", "juvenile", "unsure"]
        total = 0
        for st in state:
            for a in age:
                for s in sex:
                    value = getattr(self, f"{st}_{a}_{s}")
                    if value:
                        total += value
        return total


class OCCIdentification(models.Model):
    """
    Identification data for occurrence

    Used for:
    - Occurrence
    Is:
    - Table
    """

    occurrence = models.OneToOneField(
        Occurrence,
        on_delete=models.CASCADE,
        null=True,
        related_name="identification",
    )
    copied_ocr_identification = models.ForeignKey(
        OCRIdentification, on_delete=models.SET_NULL, null=True, blank=True
    )
    id_confirmed_by = models.CharField(max_length=1000, null=True, blank=True)
    identification_certainty = models.ForeignKey(
        IdentificationCertainty, on_delete=models.SET_NULL, null=True, blank=True
    )
    sample_type = models.ForeignKey(
        SampleType, on_delete=models.SET_NULL, null=True, blank=True
    )
    sample_destination = models.ForeignKey(
        SampleDestination, on_delete=models.SET_NULL, null=True, blank=True
    )
    permit_type = models.ForeignKey(
        PermitType, on_delete=models.SET_NULL, null=True, blank=True
    )
    permit_id = models.CharField(max_length=500, null=True, blank=True)
    collector_number = models.CharField(max_length=500, null=True, blank=True)
    barcode_number = models.CharField(max_length=500, null=True, blank=True)
    identification_comment = models.TextField(null=True, blank=True)

    class Meta:
        app_label = "boranga"

    def __str__(self):
        return str(self.occurrence)


class OCRExternalRefereeInvite(models.Model):
    email = models.EmailField()
    first_name = models.CharField(max_length=100)
    last_name = models.CharField(max_length=100)
    datetime_sent = models.DateTimeField(null=True, blank=True)
    datetime_first_logged_in = models.DateTimeField(null=True, blank=True)
    occurrence_report = models.ForeignKey(
        OccurrenceReport,
        related_name="external_referee_invites",
        on_delete=models.CASCADE,
    )
    sent_by = models.IntegerField()
    invite_text = models.TextField(blank=True)
    archived = models.BooleanField(default=False)

    class Meta:
        app_label = "boranga"
        verbose_name = "External Occurrence Report Referral"
        verbose_name_plural = "External Occurrence Report Referrals"

    def __str__(self):
        return_str = f"{self.first_name} {self.last_name} ({self.email})"
        if self.archived:
            return_str += " - Archived"
        return return_str

    def full_name(self):
        return f"{self.first_name} {self.last_name}"


class OccurrenceTenurePurpose(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()

    label = models.CharField(
        max_length=100, blank=True, null=True, validators=[no_commas_validator]
    )
    code = models.CharField(
        max_length=20, blank=True, null=True, validators=[no_commas_validator]
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Occurrence Tenure Purpose"
        verbose_name_plural = "Occurrence Tenure Purposes"

    def __str__(self):
        return f"{self.code} - {self.label}"


class OccurrenceTenureVesting(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()

    label = models.CharField(
        max_length=100, blank=True, null=True, validators=[no_commas_validator]
    )
    code = models.CharField(
        max_length=20, blank=True, null=True, validators=[no_commas_validator]
    )

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Occurrence Tenure Vesting"
        verbose_name_plural = "Occurrence Tenure Vestings"

    def __str__(self):
        return f"{self.code} - {self.label}"


def SET_NULL_AND_HISTORICAL(collector, field, sub_objs, using):
    sub_objs.update(status="historical")
    occurrence_geometry_set = collector.data.get(OccurrenceGeometry, {})
    if len(occurrence_geometry_set) > 0:
        # Create a shallow copy first to not modify the original set
        occurrence_geometry = occurrence_geometry_set.copy().pop()
        occurrence_geometry.occurrence.id
        occurrence_geometry.geometry.ewkt
        # Populate historical_occurrence_geometry_ewkb and historical_occurrence id
        sub_objs.update(historical_occurrence=occurrence_geometry.occurrence.id)
        sub_objs.update(
            historical_occurrence_geometry_ewkb=occurrence_geometry.geometry.ewkb
        )
    collector.add_field_update(field, None, sub_objs)


class OccurrenceTenure(RevisionedMixin):
    STATUS_CURRENT = "current"
    STATUS_HISTORICAL = "historical"
    STATUS_CHOICES = ((STATUS_CURRENT, "Current"), (STATUS_HISTORICAL, "Historical"))

    status = models.CharField(
        max_length=100, choices=STATUS_CHOICES, default=STATUS_CHOICES[0][0]
    )
    occurrence_geometry = models.ForeignKey(
        OccurrenceGeometry,
        related_name="occurrence_tenures",
        blank=True,
        null=True,
        on_delete=SET_NULL_AND_HISTORICAL,
    )
    historical_occurrence_geometry_ewkb = models.BinaryField(
        blank=True, null=True, editable=True
    )  # the geometry after setting the occurrence_geometry to None
    historical_occurrence = models.IntegerField(blank=True, null=True)

    tenure_area_id = models.CharField(
        max_length=100, blank=True, null=True
    )  # E.g. CPT_CADASTRE_SCDB.314159265
    tenure_area_ewkb = models.BinaryField(blank=True, null=True, editable=True)
    owner_name = models.CharField(max_length=255, blank=True, null=True)
    owner_count = models.IntegerField(blank=True, null=True)

    datetime_created = models.DateTimeField(auto_now_add=True)
    datetime_updated = models.DateTimeField(default=datetime.now)

    purpose = models.ForeignKey(
        OccurrenceTenurePurpose,
        related_name="occurrence_tenures",
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
    )
    vesting = models.ForeignKey(
        OccurrenceTenureVesting,
        related_name="occurrence_vestings",
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
    )

    comments = models.TextField(blank=True, null=True)
    significant_to_occurrence = models.BooleanField(
        null=True, blank=True, default=False
    )

    def save(self, *args, **kwargs):

        force_insert = kwargs.pop("force_insert", False)
        if force_insert:
            super().save(no_revision=True, force_insert=force_insert)
            self.save(*args, **kwargs)
        else:
            override_datetime_updated = kwargs.pop("override_datetime_updated", False)
            if not override_datetime_updated:
                self.datetime_updated = datetime.now()
            super().save(*args, **kwargs)

    class Meta:
        app_label = "boranga"
        unique_together = ("occurrence_geometry", "tenure_area_id", "status")
        verbose_name = "Occurrence Tenure"
        verbose_name_plural = "Occurrence Tenures"

    def __str__(self):
        owner_name = self.owner_name.strip() if self.owner_name else None
        owner_name_display = f": {self.owner_name}" if owner_name else ""
        return f"Tenure Area {self.tenure_area_id}{owner_name_display} [{self.get_status_display()}]"

    @property
    def typename(self):
        # The typeName (layer name) part of the tenure_area_id
        return self.tenure_area_id.split(".")[0]

    @property
    def featureid(self):
        # The featureId part of the tenure_area_id
        return self.tenure_area_id.split(".")[-1]

    @property
    def geometry(self):
        from boranga.components.spatial.utils import wkb_to_geojson

        # Return from historical geometry if historical, else from occurrence_geometry's geometry
        if self.status == self.STATUS_HISTORICAL:
            if self.historical_occurrence_geometry_ewkb:
                return wkb_to_geojson(self.historical_occurrence_geometry_ewkb)
            return None
        return wkb_to_geojson(self.occurrence_geometry.geometry.ewkb)

    @property
    def occurrence(self):
        # Return from historical occurrence if historical, else from occurrence_geometry's occurrence
        if self.status == self.STATUS_HISTORICAL:
            try:
                return Occurrence.objects.get(id=self.historical_occurrence)
            except Occurrence.DoesNotExist:
                logger.warning(
                    f"OccurrenceTenure {self.id} has historical_occurrence "
                    f"{self.historical_occurrence} which does not exist"
                )
                return None
        return self.occurrence_geometry.occurrence

    @property
    def tenure_area_centroid(self):
        from boranga.components.spatial.utils import (
            feature_json_to_geosgeometry,
            wkb_to_geojson,
        )

        if self.tenure_area_ewkb:
            geo_json = wkb_to_geojson(self.tenure_area_ewkb)
            centroid = feature_json_to_geosgeometry(geo_json).centroid
            return wkb_to_geojson(centroid.ewkb)
        return None

    @property
    def tenure_area_point_on_surface(self):
        from boranga.components.spatial.utils import (
            feature_json_to_geosgeometry,
            wkb_to_geojson,
        )

        if self.tenure_area_ewkb:
            geo_json = wkb_to_geojson(self.tenure_area_ewkb)
            centroid = feature_json_to_geosgeometry(geo_json).point_on_surface
            return wkb_to_geojson(centroid.ewkb)
        return None


class BufferGeometry(GeometryBase):
    buffered_from_geometry = models.OneToOneField(
        OccurrenceGeometry,
        on_delete=models.CASCADE,
        null=False,
        blank=False,
        related_name="buffer_geometry",
    )
    color = ColorField(default="#FFFF00")  # Yellow
    stroke = ColorField(default="#FF9900")  # Orange
    opacity = models.FloatField(
        default=0.5, validators=[MinValueValidator(0.0), MaxValueValidator(1.0)]
    )  # Used for map layer opacity

    class Meta:
        app_label = "boranga"
        verbose_name = "Buffer Geometry"
        verbose_name_plural = "Buffer Geometries"

    def related_model_field(self):
        return self.buffered_from_geometry


class SiteType(OrderedModel, ArchivableModel):
    objects = OrderedArchivableManager()
    """
    # Admin List

    Used by:
    - OccurrenceSite

    """

    name = models.CharField(max_length=250, blank=False, null=False, unique=True)

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Site Type"
        verbose_name_plural = "Site Type"

    def __str__(self):
        return str(self.name)


class OccurrenceSite(GeometryBase, DrawnByGeometry, RevisionedMixin):
    site_number = models.CharField(max_length=9, blank=True, default="")
    occurrence = models.ForeignKey(
        "Occurrence", related_name="sites", on_delete=models.CASCADE
    )

    site_name = models.CharField(max_length=255, null=True, blank=True)

    site_type = models.ForeignKey(
        SiteType, on_delete=models.SET_NULL, null=True, blank=True
    )

    related_occurrence_reports = models.ManyToManyField(OccurrenceReport, blank=True)

    comments = models.TextField(blank=True, null=True)

    visible = models.BooleanField(default=True)

    color = ColorField(default="#FF3300")  # Light red
    stroke = ColorField(default="#CC0000")  # Dark red

    def related_model_field(self):
        return self.occurrence

    class Meta:
        app_label = "boranga"
        verbose_name = "Occurrence Site"

    def save(self, *args, **kwargs):
        # Prefix "ST" char to site_number.
        if self.site_number == "":
            force_insert = kwargs.pop("force_insert", False)
            super().save(force_insert=force_insert)
            new_site_id = f"ST{str(self.pk)}"
            self.site_number = new_site_id
            self.save(*args, **kwargs)
        else:
            super().save(*args, **kwargs)


def validate_bulk_import_file_extension(value):
    ext = os.path.splitext(value.name)[1]
    valid_extensions = [".xlsx"]
    if ext not in valid_extensions:
        raise ValidationError("The bulk import file must be a .xlsx file")


def validate_bulk_import_associated_files_extension(value):
    ext = os.path.splitext(value.name)[1]
    valid_extensions = [".zip"]
    if ext not in valid_extensions:
        raise ValidationError("The associated documents file must be a .zip file")


# TODO: Would be nice to have the object id in the file path
# bit tricky before the object has been saved..
def get_occurrence_report_bulk_import_path(instance, filename):
    return f"occurrence_report/bulk-imports/{timezone.now()}/{filename}"


def get_occurrence_report_bulk_import_associated_files_path(instance, filename):
    return f"occurrence_report/bulk-imports/{timezone.now()}/{filename}"


class OccurrenceReportBulkImportTask(ArchivableModel):
    schema = models.ForeignKey(
        "OccurrenceReportBulkImportSchema",
        on_delete=models.PROTECT,
        null=True,
        blank=True,
    )
    _file = models.FileField(
        upload_to=get_occurrence_report_bulk_import_path,
        max_length=512,
        storage=private_storage,
        validators=[validate_bulk_import_file_extension],
    )
    _associated_files_zip = models.FileField(
        upload_to=get_occurrence_report_bulk_import_associated_files_path,
        max_length=512,
        storage=private_storage,
        validators=[validate_bulk_import_associated_files_extension],
        null=True,
        blank=True,
    )
    # A hash of the file to allow for duplicate detection
    file_hash = models.CharField(max_length=64, null=True, blank=True)

    rows = models.IntegerField(null=True, editable=False)
    rows_processed = models.IntegerField(default=0)

    datetime_queued = models.DateTimeField(auto_now_add=True)
    datetime_started = models.DateTimeField(null=True, blank=True)
    datetime_completed = models.DateTimeField(null=True, blank=True)

    datetime_error = models.DateTimeField(null=True, blank=True)
    error_row = models.IntegerField(null=True, blank=True)
    error_message = models.TextField(null=True, blank=True)

    email_user = models.IntegerField(null=False)

    PROCESSING_STATUS_QUEUED = "queued"
    PROCESSING_STATUS_STARTED = "started"
    PROCESSING_STATUS_FAILED = "failed"
    PROCESSING_STATUS_COMPLETED = "completed"
    PROCESSING_STATUS_ARCHIVED = "archived"

    PROCESSING_STATUS_CHOICES = (
        (PROCESSING_STATUS_QUEUED, "Queued"),
        (PROCESSING_STATUS_STARTED, "Started"),
        (PROCESSING_STATUS_FAILED, "Failed"),
        (PROCESSING_STATUS_COMPLETED, "Completed"),
        (PROCESSING_STATUS_ARCHIVED, "Archived"),
    )

    processing_status = models.CharField(
        max_length=20,
        choices=PROCESSING_STATUS_CHOICES,
        default=PROCESSING_STATUS_QUEUED,
    )

    class Meta:
        app_label = "boranga"
        verbose_name = "Occurrence Report Bulk Import Task"
        verbose_name_plural = "Occurrence Report Bulk Import Tasks"

    def save(self, *args, **kwargs):
        if not self.file_hash and self._file:
            self._file.seek(0)
            self.file_hash = hashlib.sha256(self._file.read()).hexdigest()
        super().save(*args, **kwargs)

    @property
    def file_name(self):
        return os.path.basename(self._file.name)

    @property
    def percentage_complete(self):
        if self.rows:
            return round((self.rows_processed / self.rows) * 100, 2)
        return 0

    @property
    def total_time_taken(self):
        if self.datetime_started and self.datetime_completed:
            delta = self.datetime_completed - self.datetime_started
            return delta.total_seconds()
        return None

    @property
    def total_time_taken_seconds(self):
        if self.datetime_started and self.datetime_completed:
            delta = self.datetime_completed - self.datetime_started
            return delta.seconds
        return None

    @property
    def total_time_taken_minues(self):
        if self.total_time_taken:
            return round(self.total_time_taken / 60, 2)
        return None

    @property
    def total_time_taken_human_readable(self):
        if self.total_time_taken is None:
            return None

        if self.total_time_taken < 1:
            return "Less than a second"

        if self.total_time_taken < 60:
            return f"{round(self.total_time_taken)} seconds"

        if self.total_time_taken:
            whole_minutes = int(self.total_time_taken // 60)
            remaining_seconds = round(self.total_time_taken - (whole_minutes * 60))
            if not remaining_seconds:
                return f"{whole_minutes} minutes"
            return f"{whole_minutes} minutes and {remaining_seconds} seconds"
        return None

    @property
    def time_taken_per_row(self):
        if self.datetime_started and self.datetime_completed:
            value = self.total_time_taken / self.rows_processed
            return round(value, 6)
        return None

    @property
    def file_size_bytes(self):
        if self._file:
            return self._file.size
        return None

    @property
    def file_size_megabytes(self):
        if self.file_size_bytes:
            return round(self.file_size_bytes / 1024 / 1024, 2)
        return None

    @classmethod
    def average_time_taken_per_row(cls):
        task_count = cls.objects.filter(
            datetime_completed__isnull=False, rows_processed__gt=0
        ).count()
        if task_count == 0:
            return None

        total_time_taken = 0
        for task in cls.objects.filter(
            datetime_completed__isnull=False, rows_processed__gt=0
        ):
            total_time_taken += task.time_taken_per_row

        return total_time_taken / task_count

    @property
    def estimated_processing_time_seconds(self):
        average_time_taken_per_row = (
            OccurrenceReportBulkImportTask.average_time_taken_per_row()
        )

        if self.rows and self.datetime_queued and average_time_taken_per_row:
            precisely = (self.rows - self.rows_processed) * average_time_taken_per_row
            return round(precisely)

        return None

    @property
    def estimated_processing_time_minutes(self):
        seconds = self.estimated_processing_time_seconds
        if seconds is None:
            return None

        return round(seconds / 60)

    @property
    def estimated_processing_time_human_readable(self):
        seconds = self.estimated_processing_time_seconds

        if seconds is None:
            return "No processing data available to estimate time"

        if seconds == 0:
            return "Less than a second"

        if seconds < 60:
            return f"~{seconds} seconds"

        minutes = self.estimated_processing_time_minutes

        return f"~{minutes} minutes"

    def count_rows(self):
        logger.info(f"Beginning row count for OCR Bulk Import Task {self.id}")

        try:
            workbook = openpyxl.load_workbook(self._file)
        except Exception as e:
            logger.error(f"Error opening bulk import file {self._file.name}: {e}")
            logger.error("Unable to count rows. Returning from method.")
            return

        sheet = workbook.active

        # Count the rows that have data in them
        all_rows = len(
            [row for row in sheet if not all([cell.value is None for cell in row])]
        )

        # Remove the header row
        self.rows = all_rows - 1

        logger.info(f"Found {self.rows} rows in OCR Bulk Import Task {self._file.name}")
        self.save()

    @classmethod
    def validate_headers(self, _file, schema):
        logger.info(f"Validating headers for bulk import task {self.id}")

        try:
            workbook = openpyxl.load_workbook(_file, read_only=True)
        except Exception as e:
            raise ValidationError(f"Error opening bulk import file {_file}: {e}")

        sheet = workbook.active

        headers = [cell.value for cell in sheet[1] if cell.value is not None]

        if not headers:
            raise ValidationError("No headers found in the file")

        # Check that the headers match the schema (group type and version headings)
        schema_headers = list(
            schema.columns.all().values_list("xlsx_column_header_name", flat=True)
        )
        if headers == schema_headers:
            return

        extra_headers = ",".join(map(repr, set(headers) - set(schema_headers)))
        missing_headers = ",".join(map(repr, set(schema_headers) - set(headers)))
        error_string = (
            f"The headers of the uploaded file do not match schema: {schema}."
        )
        if missing_headers:
            error_string += (
                " The file is missing the following headers that are part of the schema: "
                f"{missing_headers}."
            )
        if extra_headers:
            error_string += f" The file has the following headers that are not part of the schema: {extra_headers}"
        raise ValidationError(error_string)

    @transaction.atomic
    def process(self):
        if self.processing_status == self.PROCESSING_STATUS_COMPLETED:
            logger.info(f"Bulk import task {self.id} has already been processed")
            return

        if self.processing_status == self.PROCESSING_STATUS_FAILED:
            logger.info(
                f"Bulk import task {self.id} failed. Please correct the issues and try again"
            )
            return

        if self.processing_status == self.PROCESSING_STATUS_STARTED:
            logger.info(f"Bulk import task {self.id} is already in progress")
            return

        self.processing_status = self.PROCESSING_STATUS_STARTED
        self.datetime_started = timezone.now()
        self.save()

        if not self.rows:
            self.count_rows()

        # Open the file
        logger.info(f"Opening bulk import file {self._file.name}")
        try:
            workbook = openpyxl.load_workbook(self._file, read_only=True)
        except Exception as e:
            logger.error(f"Error opening bulk import file {self._file.name}: {e}")
            self.processing_status = (
                OccurrenceReportBulkImportTask.PROCESSING_STATUS_FAILED
            )
            self.datetime_error = timezone.now()
            self.error_message = f"Error opening bulk import file: {e}"
            self.save()
            return

        # Get the first sheet
        sheet = workbook.active

        # headers
        headers = [cell.value for cell in sheet[1] if cell.value is not None]

        # Get the rows
        rows = list(
            sheet.iter_rows(
                min_row=2,
                max_row=self.rows + 1,
                max_col=self.schema.columns.count(),
                values_only=True,
            )
        )

        errors = []
        ocr_migrated_from_ids = []
        # Process the rows
        for index, row in enumerate(rows):
            self.rows_processed = index + 1
            if self.rows_processed > self.rows:
                logger.warning(
                    f"Bulk import task {self.id} tried to process row {index + 1} "
                    "which is greater than the total number of rows"
                )
                break

            self.save()

            self.process_row(ocr_migrated_from_ids, index, headers, row, errors)

        if errors and len(errors) > 0:
            # If a single thing went wrong roll back everything
            # Imports either work completely or fail completely
            transaction.set_rollback(True)

        return errors

    def process_row(self, ocr_migrated_from_ids, index, headers, row, errors):
        row_hash = hashlib.sha256(str(row).encode()).hexdigest()
        if OccurrenceReport.objects.filter(import_hash=row_hash).exists():
            duplicate_ocr = OccurrenceReport.objects.get(import_hash=row_hash)
            error_message = (
                f"Row {index} has the exact same data as "
                f"Occurrence Report {duplicate_ocr.occurrence_report_number} did when it was imported."
            )
            errors.append(
                {
                    "row_index": index,
                    "error_type": "row",
                    "data": row,
                    "error_message": error_message,
                }
            )
            return

        ocr_migrated_from_id = row[0]

        # If the migrated from id is None then complain and return
        if not ocr_migrated_from_id:
            error_message = "Row does not have an Occurrence Report migrated from id"
            errors.append(
                {
                    "row_index": index,
                    "error_type": "missing_migrated_from_id",
                    "data": row,
                    "error_message": error_message,
                }
            )
            return

        mode = "create"
        if (
            ocr_migrated_from_id in ocr_migrated_from_ids
            or OccurrenceReport.objects.filter(
                migrated_from_id=ocr_migrated_from_id
            ).exists()
        ):
            mode = "update"

        if ocr_migrated_from_id not in ocr_migrated_from_ids:
            # Because this is happening in a transaction we need to keep track of
            # the ocr_migrated_from_ids that have been processed in this transaction
            ocr_migrated_from_ids.append(ocr_migrated_from_id)

        row_error_count = 0
        total_column_error_count = 0

        models = {}
        geometries = {}
        many_to_many_fields = {}

        # Find any sets of columns (models) where the fields are all empty
        # so that we can exclude them from the row data (and therefor not bother validating them)
        required_content_types = []
        for column_index, column in enumerate(self.schema.columns.all()):
            cell_value = row[column_index]
            if (
                cell_value is not None
                and column.django_import_content_type not in required_content_types
            ):
                required_content_types.append(column.django_import_content_type)

        indexes_to_remove = []
        for column_index, column in enumerate(self.schema.columns.all()):
            if column.django_import_content_type not in required_content_types:
                indexes_to_remove.append(column_index)

        row = [i for j, i in enumerate(row) if j not in indexes_to_remove]
        headers = [i for j, i in enumerate(headers) if j not in indexes_to_remove]

        # Validate each cell
        for column_index, column in enumerate(
            self.schema.columns.filter(
                django_import_content_type__in=required_content_types
            )
        ):
            column_error_count = 0

            cell_value = row[column_index]

            cell_value, errors_added = column.validate(
                self, cell_value, mode, index, headers, row, errors
            )

            model_class = apps.get_model(
                "boranga", column.django_import_content_type.model
            )
            field = model_class._meta.get_field(column.django_import_field_name)

            # Special case: geojson feature collection
            # TODO: Consider modifying this so that it can support multiple geometry fields
            # in the same model like the m2m one does below
            if (
                type(field) is gis_models.GeometryField
                and type(cell_value) is list
                and len(cell_value) > 0
            ):
                # Store every feature / geometry in the geometries dict
                geometries[model_class._meta.model_name] = cell_value

                # Set the first geometry as the cell value so that it is
                # created with the main model instance
                cell_value = geometries[model_class._meta.model_name][0]

            # Special case: Many to many fields
            if type(field) is ManyToManyField:
                # Store the many to many field in the many_to_many_fields dict
                # As you can't assign a list of models directly to a many to many field
                # via model_data
                if model_class._meta.model_name not in many_to_many_fields:
                    many_to_many_fields[model_class._meta.model_name] = []
                many_to_many_fields[model_class._meta.model_name].append(
                    {"field": column.django_import_field_name, "value": cell_value}
                )

            column_error_count += errors_added

            row_error_count += column_error_count
            total_column_error_count += column_error_count

            if column_error_count > 0:
                continue

            model_name = column.django_import_content_type.model

            if model_name not in models:
                models[model_name] = {"field_names": [], "values": []}

            models[model_name]["field_names"].append(column.django_import_field_name)

            # Attempting to assign the cell value directly to the m2m field
            # will raise an error, so we skip it here and handle it later
            if type(field) is ManyToManyField:
                continue

            models[model_name]["values"].append(cell_value)

        if row_error_count > 0:
            return

        model_instances = {}
        for current_model_name in models:
            model_data = dict(
                zip(
                    models[current_model_name]["field_names"],
                    models[current_model_name]["values"],
                )
            )

            # Create an instance of the model that is going to be created
            model_class = apps.get_model(
                "boranga",
                current_model_name,
            )

            if mode == "update":
                # Remove empty values from the model data
                model_data = {k: v for k, v in model_data.items() if v is not None}

                if current_model_name == OccurrenceReport._meta.model_name:
                    # Remove the migrated_from_id field as we don't want to update it
                    model_data.pop("migrated_from_id", None)

            # For OccurrenceReport check if we are creating or updating
            # and set appropriate fields if so
            if current_model_name == OccurrenceReport._meta.model_name:
                if mode == "create":
                    current_model_instance = OccurrenceReport(**model_data)
                    current_model_instance.bulk_import_task_id = self.pk
                    current_model_instance.import_hash = row_hash
                    current_model_instance.group_type_id = self.schema.group_type_id
                    current_model_instance.submitter = self.email_user
                    current_model_instance.lodgement_date = timezone.now()
                else:
                    current_model_instance = OccurrenceReport.objects.get(
                        migrated_from_id=row[0]
                    )
                    for field, value in model_data.items():
                        setattr(current_model_instance, field, value)
            elif current_model_name == Occurrence._meta.model_name:
                occ_migrated_from_id = model_data.pop("migrated_from_id", None)
                occurrence_number = model_data.pop("occurrence_number", None)

                if occ_migrated_from_id and occurrence_number:
                    error_message = (
                        "Both migrated_from_id and occurrence_number were provided. "
                        "Please provide only one of these fields."
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "ambiguous_occurrence_idenfitier",
                            "data": model_data,
                            "error_message": error_message,
                        }
                    )
                    return

                if occurrence_number:
                    try:
                        current_model_instance = Occurrence.objects.get(
                            occurrence_number=occurrence_number
                        )
                    except Occurrence.DoesNotExist:
                        error_message = "The occurrence number provided does not exist in the database"
                        errors.append(
                            {
                                "row_index": index,
                                "error_type": "invalid_occurrence_number",
                                "data": model_data,
                                "error_message": error_message,
                            }
                        )
                        return
                else:
                    if (
                        model_data.get("occurrence_name", None)
                        and Occurrence.objects.filter(
                            occurrence_name=model_data["occurrence_name"]
                        ).exists()
                    ):
                        error_message = (
                            f"An occurrence with the name '{model_data['occurrence_name']}' "
                            "already exists in the database. Please populate the Occurrence "
                            "Number or OCC Occurrence Migrated From ID field instead."
                        )
                        errors.append(
                            {
                                "row_index": index,
                                "error_type": "duplicate_occurrence_name",
                                "data": model_data,
                                "error_message": error_message,
                            }
                        )
                        return

                    if not occ_migrated_from_id:
                        if not model_data.get("group_type"):
                            model_data["group_type"] = self.schema.group_type

                        current_model_instance = Occurrence(**model_data)
                        current_model_instance.occurrence_source = (
                            Occurrence.OCCURRENCE_CHOICE_OCR
                        )
                        ocr_instance = model_instances[
                            OccurrenceReport._meta.model_name
                        ]
                        if ocr_instance.species:
                            current_model_instance.species = ocr_instance.species
                        else:
                            current_model_instance.community = ocr_instance.community
                    else:
                        if Occurrence.objects.filter(
                            migrated_from_id=occ_migrated_from_id
                        ).exists():
                            current_model_instance = Occurrence.objects.get(
                                migrated_from_id=occ_migrated_from_id
                            )
                        else:
                            current_model_instance = Occurrence.objects.create(
                                migrated_from_id=occ_migrated_from_id,
                                group_type=self.schema.group_type,
                                occurrence_source=Occurrence.OCCURRENCE_CHOICE_OCR,
                            )
                            ocr_instance = model_instances[
                                OccurrenceReport._meta.model_name
                            ]
                            if ocr_instance.species:
                                current_model_instance.species = ocr_instance.species
                            else:
                                current_model_instance.community = (
                                    ocr_instance.community
                                )
                        if (
                            not current_model_instance.group_type
                            == self.schema.group_type
                        ):
                            error_message = (
                                "The group type of the occurrence does not "
                                "match the group type of the schema"
                            )
                            errors.append(
                                {
                                    "row_index": index,
                                    "error_type": "invalid_occurrence_group_type",
                                    "data": model_data,
                                    "error_message": error_message,
                                }
                            )
                            return

                        occurrence_report = model_instances[
                            OccurrenceReport._meta.model_name
                        ]
                        if (
                            current_model_instance.species
                            and not current_model_instance.species
                            == occurrence_report.species
                        ):
                            error_message = (
                                "The species of the occurrence does not match "
                                "the species of the occurrence report"
                            )
                            errors.append(
                                {
                                    "row_index": index,
                                    "error_type": "invalid_occurrence_species",
                                    "data": model_data,
                                    "error_message": error_message,
                                }
                            )
                            return

                        if (
                            current_model_instance.community
                            and not current_model_instance.community
                            == occurrence_report.community
                        ):
                            error_message = (
                                "The community of the occurrence does not match "
                                "the community of the occurrence report"
                            )
                            errors.append(
                                {
                                    "row_index": index,
                                    "error_type": "invalid_occurrence_community",
                                    "data": model_data,
                                    "error_message": error_message,
                                }
                            )
                            return

                        for field, value in model_data.items():
                            setattr(current_model_instance, field, value)

            elif current_model_name == SubmitterInformation._meta.model_name:
                # Submitter information is created automatically when an OccurrenceReport is created
                occurrence_report = model_instances[OccurrenceReport._meta.model_name]
                current_model_instance = occurrence_report.submitter_information
                for field, value in model_data.items():
                    setattr(current_model_instance, field, value)
            else:
                current_model_instance = model_class(**model_data)

            # If we are at the top level model (OccurrenceReport) we don't need to relate it to anything
            # We also don't need to relate the Occurrence model to anything here as it is dealt with
            # as a special case further down
            # TODO: Optimize this relationships finder to minimise looping and move to a separate function
            if current_model_name not in [
                OccurrenceReport._meta.model_name,
                Occurrence._meta.model_name,
                SubmitterInformation._meta.model_name,
            ]:
                # Relate this model to it's parent instance

                related_to_parent = False
                # Look through all the model instances that have already been saved
                for potential_parent_model_key in [m for m in model_instances]:

                    # Check if this model has a relationship with the current model
                    potential_parent_instance = model_instances[
                        potential_parent_model_key
                    ]

                    # First search the current model instance for the relationship
                    # This is often faster as the child model often has the foreign key
                    # to the parent model
                    for field in current_model_instance._meta.get_fields():
                        if field.related_model == potential_parent_instance.__class__:
                            # If it does, set the relationship
                            setattr(
                                current_model_instance,
                                field.name,
                                potential_parent_instance,
                            )
                            related_to_parent = True
                            break

                    # If we didn't find a relationship in the current model, search the parent model
                    for field in potential_parent_instance.__class__._meta.get_fields():
                        if field.related_model == current_model_instance:
                            # If it does, set the relationship
                            setattr(
                                current_model_instance,
                                field.name,
                                potential_parent_instance,
                            )
                            related_to_parent = True
                            break

                if not related_to_parent:
                    error_message = (
                        "Could not find a parent model to relate this model to "
                        "(Probably due to an error saving the parent model instance)"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "relationship",
                            "data": model_data,
                            "error_message": error_message,
                        }
                    )
                    return

            try:
                if mode == "create" or (mode == "update" and len(model_data.keys())):
                    current_model_instance.save()

                self.ocr_bulk_import_generate_action_logs(mode, current_model_instance)

                model_instances[current_model_instance._meta.model_name] = (
                    current_model_instance
                )
                logger.info(f"Model instance saved: {current_model_instance}")

                # Deal with special case of relating Occurrence to OccurrenceReport
                if (
                    current_model_instance._meta.model_name
                    == Occurrence._meta.model_name
                ):
                    current_model_instance.occurrence_reports.add(
                        model_instances[OccurrenceReport._meta.model_name]
                    )

                # Deal with special case of creating mutliple geometries based on the
                # geojson text from the column
                if current_model_instance._meta.model_name in geometries:
                    logger.info(
                        f"Creating multiple geometries for {current_model_instance}"
                    )
                    for geometry in geometries[current_model_instance._meta.model_name][
                        1:
                    ]:
                        current_model_instance.pk = None
                        current_model_instance.geometry = geometry
                        current_model_instance.save()

                # Deal with special case of many to many fields
                if current_model_instance._meta.model_name in many_to_many_fields:
                    logger.info(
                        f"Adding many to many fields for {current_model_instance}"
                    )
                    for m2m_field in many_to_many_fields[
                        current_model_instance._meta.model_name
                    ]:
                        field = getattr(current_model_instance, m2m_field["field"])
                        field.set(m2m_field["value"])

            except IntegrityError as e:
                logger.error(f"Error creating model instance: {e}")
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "integrity",
                        "data": model_data,
                        "error_message": f"Error creating model instance: {e}",
                    }
                )

        return

    def ocr_bulk_import_generate_action_logs(self, mode, model_instance):
        if not model_instance._meta.model_name == OccurrenceReport._meta.model_name:
            return

        log_suffix = " (via bulk importer)"
        bulk_importer = EmailUser.objects.get(id=self.email_user)
        request = get_mock_request(bulk_importer)
        if mode == "create":
            model_instance.log_user_action(
                OccurrenceReportUserAction.ACTION_LODGE_PROPOSAL.format(
                    model_instance.occurrence_report_number
                )
                + log_suffix,
                request,
            )
            if model_instance.processing_status in [
                OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
                OccurrenceReport.PROCESSING_STATUS_APPROVED,
            ]:
                assessor = EmailUser.objects.get(id=model_instance.assigned_officer)
                request = get_mock_request(assessor)
                model_instance.log_user_action(
                    OccurrenceReportUserAction.ACTION_PROPOSED_APPROVAL.format(
                        model_instance.occurrence_report_number
                    )
                    + log_suffix,
                    request,
                )
            if (
                model_instance.processing_status
                == OccurrenceReport.PROCESSING_STATUS_APPROVED
            ):
                approver = EmailUser.objects.get(id=model_instance.assigned_approver)
                request = get_mock_request(approver)
                model_instance.log_user_action(
                    OccurrenceReportUserAction.ACTION_APPROVE.format(
                        model_instance.occurrence_report_number,
                        approver.get_full_name(),
                    )
                    + log_suffix,
                    request,
                )
        elif mode == "update":
            model_instance.log_user_action(
                OccurrenceReportUserAction.ACTION_EDIT_APPLICATION.format(
                    model_instance.occurrence_report_number
                )
                + log_suffix,
                request,
            )

    def retry(self):
        self.processing_status = self.PROCESSING_STATUS_QUEUED
        self.datetime_started = None
        self.datetime_completed = None
        self.datetime_error = None
        self.error_row = None
        self.error_message = None
        self.save()

    def revert(self):
        # Note: Using delete here due to the sheer number of records that could be created
        OccurrenceReport.objects.filter(bulk_import_task=self).delete()

        self.processing_status = self.PROCESSING_STATUS_ARCHIVED
        self.archived = True
        self.save()


class OccurrenceReportBulkImportSchema(models.Model):
    group_type = models.ForeignKey(
        GroupType, on_delete=models.PROTECT, null=False, blank=False
    )
    version = models.IntegerField(default=1)
    name = models.CharField(max_length=255, blank=True, null=True)
    tags = TaggableManager(blank=True)
    datetime_created = models.DateTimeField(auto_now_add=True)
    datetime_updated = models.DateTimeField(auto_now=True)
    is_master = models.BooleanField(default=False)

    class Meta:
        app_label = "boranga"
        verbose_name = "Occurrence Report Bulk Import Schema"
        verbose_name_plural = "Occurrence Report Bulk Import Schemas"
        ordering = ["group_type", "-version"]
        constraints = [
            models.UniqueConstraint(
                fields=[
                    "group_type",
                    "version",
                ],
                name="unique_schema_version",
            )
        ]

    def __str__(self):
        return f"Group type: {self.group_type.name} (Version: {self.version})"

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        # Every schema should have a migrated_from_id column regardless if it is used
        # for create new OCR records or updating existing ones
        content_type = ct_models.ContentType.objects.get_for_model(OccurrenceReport)
        if not self.columns.filter(
            django_import_content_type=content_type,
            django_import_field_name="migrated_from_id",
        ).exists():
            OccurrenceReportBulkImportSchemaColumn.objects.create(
                schema=self,
                xlsx_column_header_name="OCR Migrated From ID",
                django_import_content_type=content_type,
                django_import_field_name="migrated_from_id",
            )

    @property
    def mandatory_fields(self):
        if self.group_type.name == "community":
            return [
                "migrated_from_id",
                "community",
                "processing_status",
                "assigned_officer",
            ]

        return [
            "migrated_from_id",
            "species",
            "processing_status",
            "assigned_officer",
        ]

    @property
    def preview_import_file(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        columns = self.columns.all()
        if not columns.exists() or columns.count() == 0:
            logger.warning(
                f"No columns found for bulk import schema {self}. Returning empty preview file"
            )
            return workbook

        headers = [column.xlsx_column_header_name for column in columns]
        worksheet.append(headers)

        dv_types = dict(zip(DataValidation.type.values, DataValidation.type.values))
        dv_operators = dict(
            zip(DataValidation.operator.values, DataValidation.operator.values)
        )

        # Add the data validation for each column
        for index, column in enumerate(columns):
            column_letter = get_column_letter(index + 1)
            cell_range = f"{column_letter}2:{column_letter}1048576"  # 1048576 is the maximum number of rows in Excel

            model_class = column.django_import_content_type.model_class()
            if not hasattr(model_class, column.django_import_field_name):
                raise ValidationError(
                    f"Model {model_class} does not have field {column.django_import_field_name}"
                )
            model_field = model_class._meta.get_field(column.django_import_field_name)

            allow_blank = model_field.null
            if allow_blank and column.xlsx_data_validation_allow_blank is False:
                allow_blank = False

            dv = None
            if column.default_value is not None:
                dv = DataValidation(
                    type=dv_types["list"],
                    allow_blank=allow_blank,
                    formula1=column.default_value,
                    error=f"This field may only contain the value '{column.default_value}'",
                    errorTitle="Invalid value for column with default value",
                    prompt="Either leave the field blank or enter the default value",
                    promptTitle="Value",
                )
            elif isinstance(
                model_field, MultiSelectField
            ):  # MultiSelectField is a custom field, not a standard Django field
                # Unfortunately there is no easy way to embed validation in .xlsx
                # for a comma separated list of values so this will be validated
                # during the import process
                continue
            elif (
                isinstance(model_field, models.fields.CharField) and model_field.choices
            ):
                choices = [c[0] for c in model_field.choices]
                if (
                    model_class is OccurrenceReport
                    and model_field.name == "processing_status"
                ):
                    choices = [
                        c[0]
                        for c in OccurrenceReport.VALID_BULK_IMPORT_PROCESSING_STATUSES
                    ]
                dv = DataValidation(
                    type=dv_types["list"],
                    allow_blank=allow_blank,
                    formula1=",".join(choices),
                    error="Please select a valid option from the list",
                    errorTitle="Invalid selection",
                    prompt="Select a value from the list",
                    promptTitle="List selection",
                )
            elif isinstance(model_field, models.fields.CharField):
                dv = DataValidation(
                    type=dv_types["textLength"],
                    allow_blank=allow_blank,
                    operator=dv_operators["lessThanOrEqual"],
                    formula1=f"{model_field.max_length}",
                    error="Text must be less than or equal to {model_field.max_length} characters",
                    errorTitle="Text too long",
                    prompt=f"Maximum {model_field.max_length} characters",
                    promptTitle="Text length",
                )
            elif isinstance(
                model_field, (models.fields.DateTimeField, models.fields.DateField)
            ):
                dv = DataValidation(
                    type=dv_types["date"],
                    operator=dv_operators["greaterThanOrEqual"],
                    formula1="1900-01-01",
                    allow_blank=allow_blank,
                    error="Please enter a valid date",
                    errorTitle="Invalid date",
                    prompt="Enter a date",
                    promptTitle="Date",
                )
                if isinstance(model_field, models.fields.DateTimeField):
                    date_style = NamedStyle(
                        name="datetime", number_format="DD/MM/YYYY HH:MM:MM"
                    )
                    for cell in worksheet[column_letter]:
                        cell.style = date_style
            elif (
                isinstance(model_field, models.fields.IntegerField)
                and column.is_emailuser_column is False
            ):
                dv = DataValidation(
                    type=dv_types["whole"],
                    operator=dv_operators["greaterThanOrEqual"],
                    formula1="0",
                    allow_blank=allow_blank,
                    error="Please enter a whole number",
                    errorTitle="Invalid number",
                    prompt="Enter a whole number",
                    promptTitle="Whole number",
                )
            elif isinstance(model_field, models.fields.DecimalField):
                dv = DataValidation(
                    type=dv_types["decimal"],
                    allow_blank=allow_blank,
                    error="Please enter a decimal number",
                    errorTitle="Invalid number",
                    prompt="Enter a decimal number",
                    promptTitle="Decimal number",
                )
            elif isinstance(model_field, models.fields.BooleanField):
                dv = DataValidation(
                    type=dv_types["list"],
                    allow_blank=allow_blank,
                    formula1='"True,False"',
                    error="Please select True or False",
                    errorTitle="Invalid selection",
                    prompt="Select True or False",
                    promptTitle="Boolean selection",
                )
            elif (
                isinstance(model_field, models.fields.related.ForeignKey)
                and model_field.related_model
            ):
                related_model = model_field.related_model
                related_model_qs = related_model.objects.all()

                # Check if the related model is Archivable
                if issubclass(related_model, ArchivableModel):
                    related_model_qs = related_model_qs.exclude(archived=True)

                if (
                    not related_model_qs.exists()
                    or related_model_qs.count() == 0
                    or related_model._meta.model_name in ["species", "community"]
                    or related_model_qs.count()
                    > settings.OCR_BULK_IMPORT_LOOKUP_TABLE_RECORD_LIMIT
                ):
                    # If there are no records or too many records, we don't embed a data validation
                    # Instead, the field will be validated during the import process
                    continue

                display_field = get_display_field_for_model(related_model)

                dv = DataValidation(
                    type=dv_types["list"],
                    allow_blank=allow_blank,
                    formula1=f'"{",".join([str(getattr(obj, display_field)) for obj in related_model_qs])}"',
                    error="Please select a valid option from the list",
                    errorTitle="Invalid selection",
                    prompt="Select a value from the list",
                    promptTitle="List selection",
                )
            else:
                # Mostly covers TextField
                # Postgresql Text field can handle up to 65,535 characters, .xlsx can handle 32,767 characters
                # We'll gleefully assume this won't be an issue and not add a data validation for text fields =D
                continue

            dv.showErrorMessage = True
            worksheet.add_data_validation(dv)
            dv.add(cell_range)

        # Make the headers bold
        for cell in worksheet["A0:ZZ0"][0]:
            cell.font = Font(bold=True)

        # Make the column widths appropriate
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = (
                        max((dims.get(cell.column, 0), len(str(cell.value)))) + 2
                    ) + 2
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value

        return workbook

    @transaction.atomic
    def validate(self, request_user_id: int):
        # Tries adding valid sample data to each column and then saving to the database
        # Returns any errors that occur
        # Rolls back the transaction regardless of success or failure

        errors = []

        columns = self.columns.all()
        if not columns.exists() or columns.count() == 0:
            errors.append(
                {
                    "error_type": "no_columns",
                    "error_message": "No columns found in the schema",
                }
            )
            return errors

        for field in self.mandatory_fields:
            if not columns.filter(
                django_import_content_type=ct_models.ContentType.objects.get_for_model(
                    OccurrenceReport
                ),
                django_import_field_name=field,
            ).exists():
                errors.append(
                    {
                        "error_type": "missing_mandatory_column",
                        "error_message": f"Missing mandatory occurrence report column for django field: {field}",
                    }
                )
                return errors

        # Create a sample row of data
        sample_row = []

        # Special case where only a migrated_from_id or occurrence_number
        # Should be provided when both columns are present otherwise validation will fail
        row_contains_occ_migrated_from_id = (
            columns.filter(
                django_import_content_type=ct_models.ContentType.objects.get_for_model(
                    Occurrence
                ),
                django_import_field_name="migrated_from_id",
            ).exists()
            and columns.filter(
                django_import_content_type=ct_models.ContentType.objects.get_for_model(
                    Occurrence
                ),
                django_import_field_name="occurrence_number",
            ).exists()
        )
        species_or_community_identifier = None
        for column in columns:
            sample_value = column.get_sample_value(
                errors, species_or_community_identifier
            )
            if (
                column.django_import_content_type.model == Occurrence._meta.model_name
                and column.django_import_field_name == "species"
            ):
                species_or_community_identifier = Species.objects.get(
                    taxonomy__scientific_name=sample_value
                )

            if (
                column.django_import_content_type.model == Occurrence._meta.model_name
                and column.django_import_field_name == "community"
            ):
                species_or_community_identifier = Community.objects.get(
                    taxonomy__community_migrated_id=sample_value
                )

            if (
                column.django_import_content_type.model == Occurrence._meta.model_name
                and column.django_import_field_name == "migrated_from_id"
                and row_contains_occ_migrated_from_id
            ):
                sample_value = ""
            sample_row.append(sample_value)

        logger.info(f"Sample row: {sample_row}")

        preview_import_file = self.preview_import_file

        # Write the sample row to the file
        sheet = preview_import_file.active
        sheet.append(sample_row)

        # Convert the import file to a Django File object
        import_file_name = f"sample_{self.group_type.name}_{self.version}.xlsx"
        import_file = InMemoryUploadedFile(
            preview_import_file,
            "_file",
            import_file_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            None,
            None,
        )
        sample_associated_file = InMemoryUploadedFile(
            BytesIO(b"I am a test file"),
            "file",
            "sample_file.txt",
            "text/plain",
            0,
            None,
        )
        # Create a .zip file to house the sample associated file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            zip_file.writestr(
                sample_associated_file.name, sample_associated_file.read()
            )
        zip_file_name = f"sample_{self.group_type.name}_{self.version}.zip"
        zip_file = InMemoryUploadedFile(
            zip_buffer,
            "file",
            zip_file_name,
            "application/zip",
            0,
            None,
        )

        import_task = OccurrenceReportBulkImportTask(
            schema=self,
            _file=import_file,
            _associated_files_zip=zip_file,
            rows=1,
            email_user=request_user_id,
        )

        # Convert the file to bytes
        buffer = BytesIO()
        preview_import_file.save(buffer)
        buffer.seek(0)

        try:
            import_task.validate_headers(buffer, self)
        except ValidationError as e:
            logger.error(f"Error validating headers: {e}")
            logger.error(traceback.format_exc())
            errors.append(
                {
                    "error_type": "header_validation",
                    "error_message": e.message,
                }
            )
            transaction.set_rollback(True)
            return errors

        headers = [cell.value for cell in sheet[1] if cell.value is not None]

        # Get the rows
        row = list(
            sheet.iter_rows(
                min_row=2,
                max_row=2,
                max_col=self.columns.count(),
                values_only=True,
            )
        )[0]

        ocr_migrated_from_ids = []

        try:
            import_task.process_row(ocr_migrated_from_ids, 0, headers, row, errors)
        except Exception as e:
            logger.error(f"Error processing sample row: {e}")
            logger.error(traceback.format_exc())
            errors.append(
                {
                    "error_type": "row_validation",
                    "error_message": f"Error processing sample row: {e}",
                }
            )
            transaction.set_rollback(True)
            return errors

        transaction.set_rollback(True)
        return errors

    @transaction.atomic
    def copy(self, request):
        if not self.pk:
            raise ValueError("Schema must be saved before it can be copied")

        if OccurrenceReportBulkImportSchema.objects.filter(
            group_type=self.group_type
        ).exists():
            highest_version = OccurrenceReportBulkImportSchema.objects.filter(
                group_type=self.group_type
            ).aggregate(Max("version"))["version__max"]
        else:
            highest_version = 0

        new_schema = OccurrenceReportBulkImportSchema(
            group_type=self.group_type,
            version=highest_version + 1,
        )
        # If the user copying the schema has elevated permissions, the new schema
        # will also be a master schema otherwise it will be a regular schema
        if self.is_master and is_django_admin(request):
            new_schema.is_master = True

        if self.name:
            new_schema.name = f"{self.name} (Copy)"
        else:
            new_schema.name = f"Copy of Version {self.version}"
        new_schema.save()
        django_import_content_type = ct_models.ContentType.objects.get_for_model(
            OccurrenceReport
        )
        # Copy all columns except those that were automatically created
        for column in self.columns.exclude(
            django_import_content_type=django_import_content_type,
            django_import_field_name="migrated_from_id",
        ):
            # Note: Due to an issue in django-ordered-model you can't use the
            # method of cloning objects here where to make a copy of the object
            # you set the pk to None and save it (that will mess up the ordering of
            # the original schema's columns) Strange bug but can't be bothered looking into it further
            new_column = OccurrenceReportBulkImportSchemaColumn()
            for field in column._meta.fields:
                if field.name == "id":
                    continue
                setattr(new_column, field.name, getattr(column, field.name))
            new_column.schema = new_schema
            new_column.save()

            for lookup_filter in column.lookup_filters.all():
                new_lookup_filter = SchemaColumnLookupFilter.objects.create(
                    schema_column=new_column,
                    filter_field_name=lookup_filter.filter_field_name,
                    filter_type=lookup_filter.filter_type,
                )

                for value in lookup_filter.values.all():
                    SchemaColumnLookupFilterValue.objects.create(
                        lookup_filter=new_lookup_filter,
                        filter_value=value.filter_value,
                    )

        new_schema.tags.add(*self.tags.all())

        if self.is_master and not is_django_admin(request):
            # Columns copied from a 'master' schema will not be editable
            # for occurrence approvers (only by superusers and django admins)
            new_schema.columns.all().update(is_editable=False)

        return new_schema


class OccurrenceReportBulkImportSchemaColumnManager(OrderedModelManager):
    def get_queryset(self):
        return super().get_queryset().prefetch_related("lookup_filters")


class OccurrenceReportBulkImportSchemaColumn(OrderedModel):
    objects = OccurrenceReportBulkImportSchemaColumnManager()
    schema = models.ForeignKey(
        OccurrenceReportBulkImportSchema,
        related_name="columns",
        on_delete=models.CASCADE,
    )

    # These two fields define where the data from the column will be imported to
    django_import_content_type = models.ForeignKey(
        ct_models.ContentType,
        on_delete=models.PROTECT,
        null=True,
        blank=True,
        related_name="import_columns",
    )
    django_import_field_name = models.CharField(max_length=50, blank=False, null=False)
    django_lookup_field_name = models.CharField(max_length=50, blank=True, null=True)

    # The name of the column header in the .xlsx file
    xlsx_column_header_name = models.CharField(max_length=50, blank=False, null=False)

    # This field allows the user that is generating the schema the ability to
    # make non mandatory fields mandatory for a specific schema
    xlsx_data_validation_allow_blank = models.BooleanField(default=False)

    # Columns copied from a 'master' schema will not be editable
    # by occurrence approvers (only by superusers and django admins)
    is_editable = models.BooleanField(default=True)

    order_with_respect_to = "schema"

    DEFAULT_VALUE_BULK_IMPORT_SUBMITTER = "bulk_import_submitter"
    DEFAULT_VALUE_CHOICES = (
        (DEFAULT_VALUE_BULK_IMPORT_SUBMITTER, "Bulk Import Submitter"),
    )

    default_value = models.CharField(
        max_length=255, choices=DEFAULT_VALUE_CHOICES, blank=True, null=True
    )
    is_emailuser_column = models.BooleanField(default=False)

    class Meta(OrderedModel.Meta):
        app_label = "boranga"
        verbose_name = "Occurrence Report Bulk Import Schema Column"
        verbose_name_plural = "Occurrence Report Bulk Import Schema Columns"
        ordering = ["schema", "order"]
        constraints = [
            models.UniqueConstraint(
                fields=[
                    "schema",
                    "django_import_content_type",
                    "django_import_field_name",
                ],
                name="unique_schema_column_import",
                violation_error_message="This field already exists in the schema",
            ),
            models.UniqueConstraint(
                fields=["schema", "xlsx_column_header_name"],
                name="unique_schema_column_header",
                violation_error_message="This column name already exists in the schema",
            ),
            # models.UniqueConstraint(
            #     fields=["schema", "order"],
            #     name="unique_schema_column_order",
            #     violation_error_message="A column with this order value already exists in the same schema",
            # ),
        ]

    def __str__(self):
        return f"{self.xlsx_column_header_name} - {self.schema}"

    @property
    def model_name(self):
        if not self.django_import_content_type:
            return None

        model = self.django_import_content_type.model_class()
        return model._meta.verbose_name

    @property
    def field(self):
        if not self.django_import_content_type or not self.django_import_field_name:
            return None

        return self.django_import_content_type.model_class()._meta.get_field(
            self.django_import_field_name
        )

    @property
    def field_type(self):
        if not self.field:
            return None

        return self.field.get_internal_type()

    @property
    def xlsx_validation_type(self):
        if not self.field:
            return None

        return get_openpyxl_data_validation_type_for_django_field(
            self.field, column=self
        )

    @property
    def text_length(self):
        if not self.field:
            return None

        if not isinstance(self.field, models.CharField):
            return 32767

        return self.field.max_length

    @property
    def choices(self):
        if not self.field:
            return None

        model_class = self.django_import_content_type.model_class()

        return get_choices_for_field(model_class, self.field)

    @cached_property
    def related_model(self):
        if not self.django_import_content_type or not self.django_import_field_name:
            return None

        field = self.django_import_content_type.model_class()._meta.get_field(
            self.django_import_field_name
        )
        if not isinstance(field, (models.ForeignKey, models.ManyToManyField)):
            return None

        return field.related_model

    @cached_property
    def display_field(self):
        related_model = self.related_model

        if not related_model:
            return None

        if self.django_lookup_field_name:
            display_field = self.django_lookup_field_name
        else:
            display_field = get_display_field_for_model(related_model)

        return display_field

    @cached_property
    def related_model_qs(self):
        display_field = self.display_field

        filter_dict = {f"{display_field}__isnull": False}
        related_model_qs = self.related_model.objects.filter(**filter_dict)

        if issubclass(self.related_model, ArchivableModel):
            related_model_qs = self.related_model.objects.exclude(archived=True)

        return related_model_qs.order_by(display_field)

    @cached_property
    def filtered_related_model_qs(self):
        if not self.related_model_qs:
            return None

        if not self.lookup_filters.exists():
            return self.related_model_qs

        related_model_qs = self.related_model_qs

        related_model = self.related_model

        # If the related model has a group_type field, filter by the schema's group type
        # (i.e. flora, fauna or community)
        if hasattr(related_model, "group_type"):
            related_model_qs = related_model_qs.filter(
                group_type=self.schema.group_type
            )

        # Apply any lookup filters if they exist
        for lookup_filter in self.lookup_filters.all():
            lookup_filter.filter_field_name + "__" + lookup_filter.filter_type
            lookup_filter_value = lookup_filter.values.first().filter_value
            if lookup_filter.values.count() > 1:
                lookup_filter_value = lookup_filter.values.values_list(
                    "filter_value", flat=True
                )
            if lookup_filter.filter_type == "in" and not isinstance(
                lookup_filter_value, list
            ):
                lookup_filter_value = [lookup_filter_value]

            related_model_qs = related_model_qs.filter(
                **{
                    lookup_filter.filter_field_name
                    + "__"
                    + lookup_filter.filter_type: lookup_filter_value
                }
            )

        return related_model_qs

    @cached_property
    def foreign_key_count(self):
        if not self.related_model_qs:
            return 0

        # Don't return the filtered count here as if we add filters on the front end
        # the count will change which will then result in the advanced lookup interface
        # No longer being shown
        return self.related_model_qs.count()

    @cached_property
    def requires_lookup_field(self):
        if not self.django_import_content_type or not self.django_import_field_name:
            return False

        if (
            self.django_import_content_type
            == ct_models.ContentType.objects.get_for_model(OccurrenceReport)
            and self.django_import_field_name in ["species", "community"]
        ):
            return True

        return (
            self.foreign_key_count > settings.OCR_BULK_IMPORT_LOOKUP_TABLE_RECORD_LIMIT
        )

    @cached_property
    def filtered_foreign_key_count(self):
        if not self.filtered_related_model_qs:
            return 0

        return self.filtered_related_model_qs.count()

    def get_sample_value(self, errors, species_or_community_identifier=None):
        if not self.django_import_content_type:
            errors.append(
                {
                    "error_type": "no_import_content_type",
                    "error_message": f"No import content type set for column {self}",
                }
            )
            return None

        if not self.django_import_field_name:
            errors.append(
                {
                    "error_type": "no_import_field_name",
                    "error_message": f"No import field name set for column {self}",
                }
            )
            return None

        try:
            field = self.django_import_content_type.model_class()._meta.get_field(
                self.django_import_field_name
            )
        except FieldDoesNotExist:
            error_message = (
                f"Field {self.django_import_field_name} not found in model "
                f"{self.django_import_content_type.model_class()}"
            )
            errors.append(
                {
                    "error_type": "field_not_found",
                    "error_message": error_message,
                }
            )
            return None

        if (
            self.django_import_content_type
            == ct_models.ContentType.objects.get_for_model(OccurrenceReport)
            and self.django_import_field_name == "processing_status"
        ):
            # Special case as there are only 3 processing statuses allow for OCR
            return random.choice(
                [
                    OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
                    OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
                    OccurrenceReport.PROCESSING_STATUS_APPROVED,
                ]
            )

        if isinstance(field, models.ForeignKey):
            related_model_qs = self.filtered_related_model_qs

            # Special case for species or community
            # Ensure only species or communities that have occurrences are selected
            # in case the schema includes the occurrence model (quite likely)
            if field.name == "species":
                related_model_qs = related_model_qs.annotate(
                    occurrence_count=Subquery(
                        Occurrence.objects.filter(species__pk=OuterRef("pk"))
                        .values("species")
                        .annotate(count=Count("id"))
                        .values("count")
                    )
                ).filter(occurrence_count__gt=0)
            if field.name == "community":
                related_model_qs = related_model_qs.annotate(
                    occurrence_count=Subquery(
                        Occurrence.objects.filter(community__pk=OuterRef("pk"))
                        .values("community")
                        .annotate(count=Count("id"))
                        .values("count")
                    )
                ).filter(occurrence_count__gt=0)

            if not related_model_qs.exists():
                error_message = f"No records found for foreign key field {field.related_model._meta.model_name}"
                errors.append(
                    {
                        "error_type": "no_records",
                        "error_message": error_message,
                    }
                )

            display_field = self.display_field

            random_value = (
                related_model_qs.order_by("?")
                .values_list(display_field, flat=True)
                .first()
            )

            return random_value

        if isinstance(field, models.ManyToManyField):
            related_model_qs = self.filtered_related_model_qs

            if not related_model_qs.exists():
                error_message = f"No records found for many to many field {field.related_model._meta.model_name}"
                errors.append(
                    {
                        "error_type": "no_records",
                        "error_message": error_message,
                    }
                )

            display_field = self.display_field

            random_values = list(
                related_model_qs.order_by("?")
                .values_list(display_field, flat=True)
                .distinct()[: random.randint(1, 3)]
            )

            return ",".join(random_values)

        if isinstance(field, MultiSelectField):
            model_class = self.django_import_content_type.model_class()
            # Unfortunatly have to have an actual model instance to get the choices
            # as they are defined in __init__
            model_instance = model_class()
            choices = model_instance._meta.get_field(
                self.django_import_field_name
            ).choices
            if not choices or len(choices) == 0:
                errors.append(
                    {
                        "error_type": "no_choices",
                        "error_message": f"No choices found for column {self.xlsx_column_header_name}",
                    }
                )
                return None
            return random.choice([choice[1] for choice in choices])

        if isinstance(field, models.BooleanField):
            return random.choice([True, False])

        if isinstance(field, models.DecimalField):
            decimal_places = field.decimal_places if field.decimal_places else 2
            max_digits = field.max_digits if field.max_digits else 5
            max_value = 10 ** (max_digits - decimal_places - 1)
            return round(random.uniform(0, max_value), decimal_places)

        if isinstance(field, models.IntegerField):
            if self.is_emailuser_column:
                user_qs = EmailUser.objects.filter(is_active=True)
                if field.name == "assigned_officer":
                    ocr_officer_ids = member_ids(
                        settings.GROUP_NAME_OCCURRENCE_ASSESSOR
                    )
                    user_qs = user_qs.filter(id__in=ocr_officer_ids)
                if field.name == "assigned_approver":
                    ocr_approver_ids = member_ids(
                        settings.GROUP_NAME_OCCURRENCE_APPROVER
                    )
                    user_qs = user_qs.filter(id__in=ocr_approver_ids)
                user = user_qs.order_by("?").first()
                return user.email
            else:
                min_value = field.min_value if hasattr(field, "min_value") else 0
                max_value = field.max_value if hasattr(field, "max_value") else 10000
                return random.randint(min_value, max_value)

        if isinstance(field, models.DateField):
            start = datetime(1900, 1, 1, 0, 0, 0, tzinfo=dt_timezone.utc).date()
            end = timezone.now().date()
            random_date = start + (end - start) * random.random()
            return random_date.strftime("%d/%m/%Y")

        if isinstance(field, models.DateTimeField):
            start = datetime(1900, 1, 1, 0, 0, 0, tzinfo=dt_timezone.utc)
            end = timezone.now()
            random_datetime = start + (end - start) * random.random()
            return random_datetime.strftime("%d/%m/%Y %H:%M:%S")

        if isinstance(field, (models.CharField, models.TextField)):
            if (
                self.django_import_content_type.model_class() == Occurrence
                and self.django_import_field_name == "occurrence_number"
            ):
                # Special case for occurrence number (Make sure that the occurrence is of the same
                # group type as the schema and same species or community name as the sample data)
                group_type = self.schema.group_type
                random_occurrence = Occurrence.objects.filter(group_type=group_type)
                filter_field = {
                    "species__taxonomy__scientific_name": species_or_community_identifier
                }
                if group_type.name == "community":
                    filter_field = {
                        "community__taxonomy__community_migrated_id": species_or_community_identifier
                    }
                if not random_occurrence.filter(**filter_field).exists():
                    error_message = (
                        f"No occurrences found where species or community identifier = "
                        f"{species_or_community_identifier}"
                    )
                    errors.append(
                        {
                            "error_type": "no_occurrences",
                            "error_message": error_message,
                        }
                    )
                    return None
                return (
                    random_occurrence.filter(**filter_field)
                    .order_by("?")
                    .first()
                    .occurrence_number
                )

            if hasattr(field, "max_length") and field.max_length:
                random_length = random.randint(1, field.max_length)
            else:
                random_length = random.randint(1, 1000)
            return "".join(
                random.choices(
                    string.ascii_letters + string.digits + " ", k=random_length
                )
            )

        if isinstance(field, gis_models.GeometryField):
            # Generate a random point and polygon that falls within Western Australia
            # In this case the dbca building in Kensington
            return json.dumps(
                {
                    "type": "FeatureCollection",
                    "features": [
                        {
                            "type": "Feature",
                            "geometry": {
                                "type": "Point",
                                "coordinates": [115.8840195356077, -31.99563118840819],
                            },
                        },
                        {
                            "type": "Feature",
                            "geometry": {
                                "type": "Polygon",
                                "coordinates": [
                                    [
                                        [115.88337912139104, -31.995016820738698],
                                        [115.88337912139104, -31.99586499034117],
                                        [115.88434603648648, -31.99586499034117],
                                        [115.88434603648648, -31.995016820738698],
                                        [115.88337912139104, -31.995016820738698],
                                    ]
                                ],
                            },
                        },
                    ],
                }
            )

        if isinstance(field, models.FileField):
            return "sample_file.txt"

        raise ValueError(
            f"Not able to generate sample data for field {field} of type {type(field)}"
        )

    @property
    def preview_foreign_key_values_xlsx(self):
        related_model_qs = self.filtered_related_model_qs

        if not related_model_qs:
            raise ValueError("No related model queryset found")

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        display_field = self.display_field

        # Query the max character length of the display field
        max_length = related_model_qs.aggregate(
            max_length=Max(Length(Cast(display_field, output_field=CharField())))
        )["max_length"]

        if len(self.xlsx_column_header_name) > max_length:
            max_length = len(self.xlsx_column_header_name)

        headers = [self.xlsx_column_header_name]
        worksheet.append(headers)
        for cell_value in related_model_qs.order_by(display_field).values_list(
            display_field, flat=True
        ):
            worksheet.append([cell_value])

        # Make the headers bold
        worksheet["A1"].font = Font(bold=True)

        # Make the column widths appropriate
        worksheet.column_dimensions["A"].width = max_length + 2

        return workbook

    def validate(self, task, cell_value, mode, index, headers, row, errors):
        from boranga.components.spatial.utils import get_geometry_array_from_geojson

        errors_added = 0

        if mode == "update" and (cell_value is None or cell_value == ""):
            # When updating an OCR many of the columns may be blank
            # So we don't need to validate them if they don't have a value
            return cell_value, errors_added

        try:
            model_class = apps.get_model(
                "boranga", self.django_import_content_type.model
            )
        except LookupError:
            errors.append(
                {
                    "row_index": index,
                    "error_type": "column",
                    "data": cell_value,
                    "error_message": f"Model class {self.django_import_content_type.model} not found",
                }
            )
            errors_added += 1
            return cell_value, errors_added

        if not hasattr(model_class, self.django_import_field_name):
            errors.append(
                {
                    "row_index": index,
                    "error_type": "column",
                    "data": cell_value,
                    "error_message": f"Field {self.django_import_field_name} not found in model {model_class}",
                }
            )
            errors_added += 1
            return cell_value, errors_added

        field = model_class._meta.get_field(self.django_import_field_name)

        if self.default_value:
            if self.default_value == self.DEFAULT_VALUE_BULK_IMPORT_SUBMITTER:
                cell_value = task.email_user
            return cell_value, errors_added

        if (
            self.django_import_content_type
            == ct_models.ContentType.objects.get_for_model(OccurrenceReport)
            and self.django_import_field_name == "processing_status"
        ):
            if cell_value not in [
                c[0] for c in OccurrenceReport.VALID_BULK_IMPORT_PROCESSING_STATUSES
            ]:
                error_message = (
                    f"Value {cell_value} in column {self.xlsx_column_header_name} "
                    f"is not a valid processing status (must be one of "
                    f"{str([c[0] for c in OccurrenceReport.VALID_BULK_IMPORT_PROCESSING_STATUSES])})"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            if cell_value in [
                OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
                OccurrenceReport.PROCESSING_STATUS_APPROVED,
            ]:
                assigned_officer_column = self.schema.columns.filter(
                    django_import_content_type=ct_models.ContentType.objects.get_for_model(
                        OccurrenceReport
                    ),
                    django_import_field_name="assigned_officer",
                ).first()
                assigned_officer_email = row[
                    headers.index(assigned_officer_column.xlsx_column_header_name)
                ]
                assigned_officer_id = None
                try:
                    assigned_officer_id = EmailUser.objects.get(
                        email=assigned_officer_email
                    ).id
                except EmailUser.DoesNotExist:
                    error_message = (
                        "No ledger user found for assigned_officer with "
                        f"email address: {assigned_officer_email}"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": assigned_officer_email,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1
                    return cell_value, errors_added

                if not belongs_to_by_user_id(
                    assigned_officer_id, settings.GROUP_NAME_OCCURRENCE_ASSESSOR
                ):
                    error_message = (
                        f"User with email {assigned_officer_email} "
                        f"(Ledger Emailuser ID: {assigned_officer_id}) is not a member"
                        " of the occurrence assessor group"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": assigned_officer_email,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1
            if cell_value == OccurrenceReport.PROCESSING_STATUS_APPROVED:
                assigned_approver_column = self.schema.columns.filter(
                    django_import_content_type=ct_models.ContentType.objects.get_for_model(
                        OccurrenceReport
                    ),
                    django_import_field_name="assigned_approver",
                ).first()
                assigned_approver_email = row[
                    headers.index(assigned_approver_column.xlsx_column_header_name)
                ]
                assigned_approver_id = None
                try:
                    assigned_approver_id = EmailUser.objects.get(
                        email=assigned_approver_email
                    ).id
                except EmailUser.DoesNotExist:
                    error_message = (
                        "No ledger user found for assigned_approver with "
                        f"email address: {assigned_approver_email}"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": assigned_approver_email,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1
                    return cell_value, errors_added

                if not belongs_to_by_user_id(
                    assigned_approver_id, settings.GROUP_NAME_OCCURRENCE_APPROVER
                ):
                    error_message = (
                        f"User with email {assigned_approver_email} "
                        f"(Ledger Emailuser ID: {assigned_approver_id}) is not a member"
                        " of the occurrence approver group"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": assigned_approver_email,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1

            return cell_value, errors_added

        if not self.xlsx_data_validation_allow_blank and (
            cell_value is None or cell_value == ""
        ):
            errors.append(
                {
                    "row_index": index,
                    "error_type": "column",
                    "data": cell_value,
                    "error_message": f"Value in column {self.xlsx_column_header_name} is blank",
                }
            )
            errors_added += 1

        xlsx_data_validation_type = self.xlsx_validation_type

        if self.is_emailuser_column:
            if not cell_value:
                # Depending on the status of the OCR the assigned_officer and / or assigned_approver
                # fields may be blank
                return cell_value, errors_added
            try:
                cell_value = EmailUser.objects.get(email=cell_value).id
            except EmailUser.DoesNotExist:
                error_message = f"No ledger user found with email address: {cell_value}"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
            return cell_value, errors_added

        if isinstance(field, MultiSelectField):
            if not cell_value:
                return cell_value, errors_added

            if not isinstance(cell_value, str):
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": f"Value {cell_value} in column {self.xlsx_column_header_name} is not a string",
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            # Unfortunatly have to have an actual model instance to get the choices
            # as they are defined in __init__
            model_instance = model_class()
            choices = model_instance._meta.get_field(
                self.django_import_field_name
            ).choices

            display_values = cell_value.split(",")
            cell_value = []
            for display_value in [
                display_value.strip() for display_value in display_values
            ]:
                if display_value not in [choice[1] for choice in choices]:
                    error_message = (
                        f"Value '{display_value}' in column {self.xlsx_column_header_name} "
                        "is not in the list"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": cell_value,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1
                else:
                    cell_value.append(
                        [
                            choice[0]
                            for choice in field.choices
                            if choice[1] == display_value
                        ][0]
                    )
            return cell_value, errors_added

        if isinstance(field, models.FileField):
            if not task._associated_files_zip:
                error_message = "No associated files zip found"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            associated_files_zip = zipfile.ZipFile(task._associated_files_zip, "r")

            # Check if the file exists in the zip
            if cell_value not in associated_files_zip.namelist():
                error_message = f"File {cell_value} not found in associated files zip"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            # Get a reference to the file in memory
            file_in_memory = associated_files_zip.open(cell_value)

            extension = cell_value.split(".")[-1].lower()
            # Get the content type of the file
            try:
                content_type = mimetypes.types_map["." + str(extension)]
            except KeyError:
                error_message = f"File extension {extension} not found in mimetypes"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            cell_value = InMemoryUploadedFile(
                file_in_memory, field.name, cell_value, content_type, None, None
            )

            return cell_value, errors_added

        if isinstance(field, gis_models.GeometryField):
            try:
                geom_json = json.loads(cell_value)
            except (json.JSONDecodeError, TypeError):
                error_message = f"Value {cell_value} in column {self.xlsx_column_header_name} is not a valid JSON"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            cell_value = []

            geojson_type = geom_json.get("type", None)
            if not geojson_type or geojson_type != "FeatureCollection":
                error_message = (
                    f"Value {cell_value} in column {self.xlsx_column_header_name} "
                    "does not contain a valid FeatureCollection"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1

                return cell_value, errors_added

            cell_value = get_geometry_array_from_geojson(
                geom_json,
                cell_value,
                index,
                self.xlsx_column_header_name,
                errors,
                errors_added,
            )

            return cell_value, errors_added

        if xlsx_data_validation_type == "textLength" and field.max_length:
            if len(str(cell_value)) > field.max_length:
                error_message = f"Value {cell_value} in column {self.xlsx_column_header_name} has too many characters"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1

        if xlsx_data_validation_type == "whole":
            if not isinstance(cell_value, int):
                errors_message = f"Value {cell_value} in column {self.xlsx_column_header_name} is not an integer"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": errors_message,
                    }
                )
                errors_added += 1

        if xlsx_data_validation_type == "decimal":
            try:
                cell_value = Decimal(cell_value)
            except Exception:
                error_message = f"Value {cell_value} in column {self.xlsx_column_header_name} is not a decimal"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1

        if xlsx_data_validation_type in ["date", "time"]:
            # The cell_value should already be a datetime object since openpyxl
            # converts cells formatted as dates to datetime objects automatically
            # but when validating the schema
            if not isinstance(cell_value, datetime):
                try:
                    if xlsx_data_validation_type == "date":
                        cell_value = datetime.strptime(cell_value, "%d/%m/%Y")
                    elif xlsx_data_validation_type == "time":
                        cell_value = datetime.strptime(cell_value, "%d/%m/%Y %H:%M:%S")
                except ValueError:
                    error_message = (
                        f"Value {cell_value} in column {self.xlsx_column_header_name} "
                        "was not able to be converted to a datetime object"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": cell_value,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1
                    return cell_value, errors_added

            # Make the datetime object timezone aware
            cell_value = cell_value.replace(
                tzinfo=zoneinfo.ZoneInfo(settings.TIME_ZONE)
            )

            return cell_value, errors_added

        if isinstance(field, models.ForeignKey):
            related_model = field.related_model
            related_model_qs = self.related_model_qs

            # Check if the related model is Archivable
            if issubclass(related_model, ArchivableModel):
                related_model_qs = related_model_qs.exclude(archived=True)

            if not related_model_qs.exists() or related_model_qs.count() == 0:
                error_message = f"No records found for foreign key {field.related_model._meta.model_name}"
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                return cell_value, errors_added

            # Use the django lookup field to find the value
            if self.django_lookup_field_name:
                lookup_field = self.django_lookup_field_name
            else:
                lookup_field = get_display_field_for_model(related_model)

            try:
                related_model_instance = related_model_qs.get(
                    **{lookup_field: cell_value}
                )

            except FieldError:
                error_message = (
                    f"Can't find {self.django_import_field_name} record by looking up "
                    f"{lookup_field} with value {cell_value} "
                    f"for column {self.xlsx_column_header_name} "
                    f" no field {lookup_field} found in {related_model}"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added
            except related_model.DoesNotExist:
                error_message = (
                    f"Can't find {self.django_import_field_name} record by looking up "
                    f"{self.django_lookup_field_name} with value {cell_value} "
                    f"for column {self.xlsx_column_header_name}"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            # Replace the lookup cell_value with the actual instance to be assigned
            cell_value = related_model_instance
            return cell_value, errors_added

        if isinstance(field, models.ManyToManyField):
            related_model = field.related_model
            related_model_qs = related_model.objects.all()

            # Check if the related model is Archivable
            if issubclass(related_model, ArchivableModel):
                related_model_qs = related_model_qs.exclude(archived=True)

            if not related_model_qs.exists() or related_model_qs.count() == 0:
                return cell_value, errors_added

            # Use the django lookup field to find the value
            if self.django_lookup_field_name:
                lookup_field = self.django_lookup_field_name
            else:
                lookup_field = get_display_field_for_model(related_model)

            lookup_field += "__in"

            cell_value = [
                c.strip()
                for c in cell_value.split(settings.OCR_BULK_IMPORT_M2M_DELIMITER)
            ]

            try:
                related_model_instances = related_model_qs.filter(
                    **{lookup_field: cell_value}
                )
            except FieldError:
                error_message = (
                    f"Can't find {self.django_import_field_name} record by looking up "
                    f"{lookup_field} with value {cell_value} "
                    f"for column {self.xlsx_column_header_name} "
                    f" no field {lookup_field} found in {related_model}"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added
            except related_model.DoesNotExist:
                error_message = (
                    f"Can't find {self.django_import_field_name} record by looking up "
                    f"{self.django_lookup_field_name} with value {cell_value} "
                    f"for column {self.xlsx_column_header_name}"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            if related_model_instances.count() == 0:
                error_message = (
                    f"Can't find any {self.django_import_field_name} records by looking up "
                    f"{lookup_field} with value {cell_value} "
                    f"for column {self.xlsx_column_header_name}"
                )
                if "," in cell_value[0]:
                    error_message += (
                        " (Hint: The delimiter for many to many fields is '"
                        f"{settings.OCR_BULK_IMPORT_M2M_DELIMITER}' not ',')"
                    )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
                return cell_value, errors_added

            # Replace the lookup cell_value with a list of model instances to be assigned
            cell_value = list(related_model_instances)
            return cell_value, errors_added

        if isinstance(field, models.BooleanField):
            if cell_value is None or cell_value == "":
                if not field.null:
                    error_message = (
                        f"Value {cell_value} in column {self.xlsx_column_header_name} "
                        "is required to be a boolean"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": cell_value,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1
                return cell_value, errors_added

            # Convert the cell value to a boolean as the application that was used
            # to create the xlsx may have stored it as a string. For example, it often
            # automatically formats cells with a text value of 'TRUE' or 'FALSE' as a boolean
            # then when the file is saved the value in the cell will be converted to '=TRUE()' or '=FALSE()'
            if isinstance(cell_value, str):
                cell_value = cell_value.lower()
                if cell_value in ["=true()", "true", "1", "yes", "y"]:
                    cell_value = True
                elif cell_value in ["=false()", "false", "0", "no", "n"]:
                    cell_value = False
                else:
                    error_message = (
                        f"Value {cell_value} in column {self.xlsx_column_header_name} "
                        "is not a valid boolean. The bulk importer is able to convert "
                        "the following values to booleans: 'True', '1', 'yes', 'y', "
                        "'=TRUE()', 'False', '0', 'no', 'n', '=False()'"
                    )
                    errors.append(
                        {
                            "row_index": index,
                            "error_type": "column",
                            "data": cell_value,
                            "error_message": error_message,
                        }
                    )
                    errors_added += 1

            if cell_value not in [True, False]:
                error_message = (
                    f"Value {cell_value} in column {self.xlsx_column_header_name} "
                    "is not a valid boolean"
                )
                errors.append(
                    {
                        "row_index": index,
                        "error_type": "column",
                        "data": cell_value,
                        "error_message": error_message,
                    }
                )
                errors_added += 1
            return cell_value, errors_added

        return cell_value, errors_added


class SchemaColumnLookupFilter(models.Model):
    schema_column = models.ForeignKey(
        OccurrenceReportBulkImportSchemaColumn,
        related_name="lookup_filters",
        on_delete=models.CASCADE,
    )

    LOOKUP_FILTER_TYPE_EXACT = "exact"
    LOOKUP_FILTER_TYPE_IEXACT = "iexact"
    LOOKUP_FILTER_TYPE_CONTAINS = "contains"
    LOOKUP_FILTER_TYPE_ICONTAINS = "icontains"
    LOOKUP_FILTER_TYPE_STARTSWITH = "startswith"
    LOOKUP_FILTER_TYPE_ISTARTSWITH = "istartswith"
    LOOKUP_FILTER_TYPE_ENDSWITH = "endswith"
    LOOKUP_FILTER_TYPE_IENDSWITH = "iendswith"
    LOOKUP_FILTER_TYPE_GT = "gt"
    LOOKUP_FILTER_TYPE_GTE = "gte"
    LOOKUP_FILTER_TYPE_LT = "lt"
    LOOKUP_FILTER_TYPE_LTE = "lte"
    # Only supporting a single value per lookup filter at this stage
    # LOOKUP_FILTER_TYPE_IN = "in"

    LOOKUP_FILTER_TYPES = (
        (LOOKUP_FILTER_TYPE_EXACT, "Exact"),
        (LOOKUP_FILTER_TYPE_IEXACT, "Case-insensitive Exact"),
        (LOOKUP_FILTER_TYPE_CONTAINS, "Contains"),
        (LOOKUP_FILTER_TYPE_ICONTAINS, "Case-insensitive Contains"),
        (LOOKUP_FILTER_TYPE_STARTSWITH, "Starts with"),
        (LOOKUP_FILTER_TYPE_ISTARTSWITH, "Case-insensitive Starts with"),
        (LOOKUP_FILTER_TYPE_ENDSWITH, "Ends with"),
        (LOOKUP_FILTER_TYPE_IENDSWITH, "Case-insensitive Ends with"),
        (LOOKUP_FILTER_TYPE_GT, "Greater than"),
        (LOOKUP_FILTER_TYPE_GTE, "Greater than or equal to"),
        (LOOKUP_FILTER_TYPE_LT, "Less than"),
        (LOOKUP_FILTER_TYPE_LTE, "Less than or equal to"),
        # (LOOKUP_FILTER_TYPE_IN, "In"),
    )

    filter_field_name = models.CharField(max_length=50, blank=False, null=False)
    filter_type = models.CharField(
        max_length=50,
        choices=LOOKUP_FILTER_TYPES,
        default=LOOKUP_FILTER_TYPE_EXACT,
        blank=False,
        null=False,
    )

    class Meta:
        app_label = "boranga"
        verbose_name = "Schema Column Lookup Filter"
        verbose_name_plural = "Schema Column Lookup Filters"
        constraints = [
            models.UniqueConstraint(
                fields=["schema_column", "filter_field_name", "filter_type"],
                name="unique_schema_column_lookup_field",
                violation_error_message=(
                    "A lookup filter with the same name and type "
                    "already exists for this schema column"
                ),
            )
        ]

    def __str__(self):
        return f"{self.schema_column} - {self.filter_field_name}"


class SchemaColumnLookupFilterValue(models.Model):
    lookup_filter = models.ForeignKey(
        SchemaColumnLookupFilter,
        related_name="values",
        on_delete=models.CASCADE,
    )

    filter_value = models.CharField(max_length=255, blank=True, null=True)

    class Meta:
        app_label = "boranga"
        verbose_name = "Schema Column Lookup Filter Value"
        verbose_name_plural = "Schema Column Lookup Filter Values"

    def __str__(self):
        return f"{self.lookup_filter} - {self.filter_value}"


# Occurrence Report Document
reversion.register(OccurrenceReportDocument)

# Occurrence Report Threat
reversion.register(OCRConservationThreat)

# Occurrence Report Observer Detail
reversion.register(OCRObserverDetail)

reversion.register(OCRHabitatComposition)
reversion.register(OCRHabitatCondition)
reversion.register(OCRVegetationStructure)
reversion.register(OCRFireHistory)
reversion.register(OCRAssociatedSpecies)
reversion.register(OCRObservationDetail)
reversion.register(OCRPlantCount)
reversion.register(OCRAnimalObservation)
reversion.register(OCRIdentification)

# Occurrence Report
reversion.register(
    OccurrenceReport,
    follow=[
        "species",
        "community",
        "habitat_composition",
        "habitat_condition",
        "vegetation_structure",
        "fire_history",
        "associated_species",
        "observation_detail",
        "plant_count",
        "animal_observation",
        "identification",
    ],
)

# Occurrence Document
reversion.register(OccurrenceDocument)

# Occurrence Threat
reversion.register(OCCConservationThreat)

# Occurrence Contact Detail
reversion.register(OCCContactDetail)

# Occurrence Site
reversion.register(OccurrenceSite)

# Occurrence Tenure
reversion.register(OccurrenceTenure)

reversion.register(OCCHabitatComposition)
reversion.register(OCCHabitatCondition)
reversion.register(OCCVegetationStructure)
reversion.register(OCCFireHistory)
reversion.register(OCCAssociatedSpecies)
reversion.register(OCCObservationDetail)
reversion.register(OCCPlantCount)
reversion.register(OCCAnimalObservation)
reversion.register(OCCIdentification)

# Occurrence
reversion.register(
    Occurrence,
    follow=[
        "species",
        "community",
        "habitat_composition",
        "habitat_condition",
        "vegetation_structure",
        "fire_history",
        "associated_species",
        "observation_detail",
        "plant_count",
        "animal_observation",
        "identification",
    ],
)

reversion.register(OccurrenceReportGeometry)
reversion.register(OccurrenceGeometry)
