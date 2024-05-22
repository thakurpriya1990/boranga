import json
import logging
from datetime import datetime, time
from io import BytesIO

import pandas as pd
from django.db import transaction
from django.db.models import CharField, Q, Value
from django.db.models.functions import Concat
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from ledger_api_client.ledger_models import EmailUserRO as EmailUser
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import serializers, views, viewsets, mixins
from rest_framework.decorators import action as detail_route
from rest_framework.decorators import action as list_route
from rest_framework.decorators import renderer_classes
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework_datatables.filters import DatatablesFilterBackend
from rest_framework_datatables.pagination import DatatablesPageNumberPagination

from boranga.components.conservation_status.serializers import SendReferralSerializer
from boranga.components.main.api import (
    DatumSearchMixing,
    UserActionLoggingViewset,
    search_datums,
)
from boranga.components.main.related_item import RelatedItemsSerializer
from boranga.components.main.spatial_utils import (
    spatially_process_geometry,
    transform_json_geometry,
)
from boranga.components.main.utils import validate_threat_request
from boranga.components.occurrence.models import (
    AnimalHealth,
    CoordinationSource,
    CountedSubject,
    DeathReason,
    Drainage,
    IdentificationCertainty,
    Intensity,
    LandForm,
    Location,
    LocationAccuracy,
    ObservationMethod,
    OCCAnimalObservation,
    OCCAssociatedSpecies,
    OCCConservationThreat,
    OCCFireHistory,
    OCCHabitatComposition,
    OCCHabitatCondition,
    OCCIdentification,
    OCCObservationDetail,
    OCCPlantCount,
    Occurrence,
    OccurrenceDocument,
    OccurrenceReport,
    OccurrenceReportAmendmentRequest,
    OccurrenceReportAmendmentRequestDocument,
    OccurrenceReportDocument,
    OccurrenceReportGeometry,
    OccurrenceReportReferral,
    OccurrenceReportUserAction,
    OccurrenceSource,
    OccurrenceUserAction,
    OCRAnimalObservation,
    OCRAssociatedSpecies,
    OCRConservationThreat,
    OCRFireHistory,
    OCRHabitatComposition,
    OCRHabitatCondition,
    OCRIdentification,
    OCRObservationDetail,
    OCRObserverDetail,
    OCRPlantCount,
    PermitType,
    PlantCondition,
    PlantCountAccuracy,
    PlantCountMethod,
    PrimaryDetectionMethod,
    ReproductiveMaturity,
    RockType,
    SampleDestination,
    SampleType,
    SecondarySign,
    SoilColour,
    SoilCondition,
    SoilType,
    WildStatus,
)
from boranga.components.occurrence.serializers import (
    BackToAssessorSerializer,
    CreateOccurrenceReportSerializer,
    CreateOccurrenceSerializer,
    InternalOccurrenceReportReferralSerializer,
    InternalOccurrenceReportSerializer,
    ListInternalOccurrenceReportSerializer,
    ListOccurrenceReportSerializer,
    ListOccurrenceSerializer,
    ListOCRReportMinimalSerializer,
    OCCConservationThreatSerializer,
    OccurrenceDocumentSerializer,
    OccurrenceLogEntrySerializer,
    OccurrenceReportAmendmentRequestSerializer,
    OccurrenceReportDocumentSerializer,
    OccurrenceReportLogEntrySerializer,
    OccurrenceReportReferralSerializer,
    OccurrenceReportSerializer,
    OccurrenceReportUserActionSerializer,
    OccurrenceSerializer,
    OccurrenceUserActionSerializer,
    OCRConservationThreatSerializer,
    OCRObserverDetailSerializer,
    ProposeApproveSerializer,
    ProposeDeclineSerializer,
    SaveLocationSerializer,
    SaveOCCAnimalObservationSerializer,
    SaveOCCAssociatedSpeciesSerializer,
    SaveOCCConservationThreatSerializer,
    SaveOCCFireHistorySerializer,
    SaveOCCHabitatCompositionSerializer,
    SaveOCCHabitatConditionSerializer,
    SaveOCCIdentificationSerializer,
    SaveOCCObservationDetailSerializer,
    SaveOCCPlantCountSerializer,
    SaveOccurrenceDocumentSerializer,
    SaveOccurrenceReportDocumentSerializer,
    SaveOccurrenceReportSerializer,
    SaveOccurrenceSerializer,
    SaveOCRAnimalObservationSerializer,
    SaveOCRAssociatedSpeciesSerializer,
    SaveOCRConservationThreatSerializer,
    SaveOCRFireHistorySerializer,
    SaveOCRHabitatCompositionSerializer,
    SaveOCRHabitatConditionSerializer,
    SaveOCRIdentificationSerializer,
    SaveOCRObservationDetailSerializer,
    SaveOCRPlantCountSerializer,
)
from boranga.components.occurrence.utils import (
    ocr_proposal_submit,
    process_shapefile_document,
    save_geometry,
    validate_map_files,
)
from boranga.components.species_and_communities.models import (
    CommunityTaxonomy,
    GroupType,
    Species,
)
from boranga.helpers import is_customer, is_internal

logger = logging.getLogger(__name__)


class OccurrenceReportFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        if "internal" in view.name:
            total_count = queryset.count()

            filter_group_type = request.GET.get("filter_group_type")
            if filter_group_type and not filter_group_type.lower() == "all":
                queryset = queryset.filter(group_type__name=filter_group_type)

            filter_occurrence = request.GET.get("filter_occurrence")
            if filter_occurrence and not filter_occurrence.lower() == "all":
                queryset = queryset.filter(occurrence_id=filter_occurrence)

            filter_scientific_name = request.GET.get("filter_scientific_name")
            if filter_scientific_name and not filter_scientific_name.lower() == "all":
                queryset = queryset.filter(species__taxonomy__id=filter_scientific_name)

            filter_community_name = request.GET.get("filter_community_name")
            if filter_community_name and not filter_community_name.lower() == "all":
                queryset = queryset.filter(
                    community__taxonomy__id=filter_community_name
                )

            filter_status = request.GET.get("filter_status")
            if filter_status and not filter_status.lower() == "all":
                queryset = queryset.filter(processing_status=filter_status)

            def get_date(filter_date):
                date = request.GET.get(filter_date)
                if date:
                    date = datetime.strptime(date, "%Y-%m-%d")
                return date

            filter_submitted_from_date = get_date("filter_submitted_from_date")
            filter_submitted_to_date = get_date("filter_submitted_to_date")
            if filter_submitted_to_date:
                filter_submitted_to_date = datetime.combine(
                    filter_submitted_to_date, time.max
                )

            if filter_submitted_from_date and not filter_submitted_to_date:
                queryset = queryset.filter(
                    reported_date__gte=filter_submitted_from_date
                )

            if filter_submitted_from_date and filter_submitted_to_date:
                queryset = queryset.filter(
                    reported_date__range=[
                        filter_submitted_from_date,
                        filter_submitted_to_date,
                    ]
                )

            if filter_submitted_to_date and not filter_submitted_from_date:
                queryset = queryset.filter(reported_date__lte=filter_submitted_to_date)


            filter_from_effective_from_date = request.GET.get("filter_from_effective_from_date")
            filter_to_effective_from_date = request.GET.get("filter_to_effective_from_date")

            filter_from_effective_to_date = request.GET.get("filter_from_effective_to_date")
            filter_to_effective_to_date = request.GET.get("filter_to_effective_to_date")

            if filter_from_effective_from_date:
                queryset = queryset.filter(
                    effective_from__gte=filter_from_effective_from_date
                )
            if filter_to_effective_from_date:
                queryset = queryset.filter(
                    effective_from__lte=filter_to_effective_from_date
                )

            if filter_from_effective_to_date:
                queryset = queryset.filter(
                    effective_to__gte=filter_from_effective_to_date
                )
            if filter_to_effective_to_date:
                queryset = queryset.filter(
                    effective_to__lte=filter_to_effective_to_date
                )

            filter_from_review_due_date = request.GET.get("filter_from_review_due_date")
            filter_to_review_due_date = request.GET.get("filter_to_review_due_date")

            if filter_from_review_due_date:
                queryset = queryset.filter(
                    review_due_date__gte=filter_from_review_due_date
                )
            if filter_to_review_due_date:
                queryset = queryset.filter(
                    review_due_date__lte=filter_to_review_due_date
                )

        if "external" in view.name:
            total_count = queryset.count()

            filter_group_type = request.GET.get("filter_group_type")
            if filter_group_type and not filter_group_type.lower() == "all":
                queryset = queryset.filter(group_type__name=filter_group_type)

            # filter_scientific_name is the species_id
            filter_scientific_name = request.GET.get("filter_scientific_name")
            if filter_scientific_name and not filter_scientific_name.lower() == "all":
                queryset = queryset.filter(species=filter_scientific_name)

            # filter_community_name is the community_id
            filter_community_name = request.GET.get("filter_community_name")
            if filter_community_name and not filter_community_name.lower() == "all":
                queryset = queryset.filter(community=filter_community_name)

            filter_application_status = request.GET.get("filter_application_status")
            if (
                filter_application_status
                and not filter_application_status.lower() == "all"
            ):
                queryset = queryset.filter(customer_status=filter_application_status)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset


# class OccurrenceReportRenderer(DatatablesRenderer):
#     def render(self, data, accepted_media_type=None, renderer_context=None):
#         if 'view' in renderer_context and hasattr(renderer_context['view'], '_datatables_total_count'):
#             data['recordsTotal'] = renderer_context['view']._datatables_total_count
#         return super(OccurrenceReportRenderer, self).render(data, accepted_media_type, renderer_context)


class OccurrenceReportPaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    filter_backends = (OccurrenceReportFilterBackend,)
    pagination_class = DatatablesPageNumberPagination
    queryset = OccurrenceReport.objects.none()
    serializer_class = ListOccurrenceReportSerializer
    page_size = 10

    def get_serializer_class(self):
        if self.action == "occurrence_report_internal":
            return ListInternalOccurrenceReportSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = OccurrenceReport.objects.all()
        if is_customer(self.request):
            qs = qs.filter(submitter=self.request.user.id)

        return qs

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_external(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = qs.filter(internal_application=False)
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListOccurrenceReportSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListInternalOccurrenceReportSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_external_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "group_type",
            "scientific_name",
            "community_name",
            "customer_status",
            "occurrence_report_number",
        ]

        serializer = ListOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Type",
            "Scientific Name",
            "Community Name",
            "Status",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Type",
            "Scientific Name",
            "Community Name",
            "Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ExternalOccurrenceReports.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ExternalOccurrenceReports.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_internal_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "scientific_name",
            "reported_date",
            "submitter",
            "processing_status",
            "occurrence_report_number",
            "occurrence_name",
        ]

        serializer = ListInternalOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
            "Occurrence",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Occurrence",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def community_occurrence_report_internal_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "community_name",
            "reported_date",
            "submitter",
            "processing_status",
            "occurrence_report_number",
            "occurrence_name",
        ]

        serializer = ListInternalOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Community Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
            "Occurrence",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Occurrence",
            "Community Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Community.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Community.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")


class OccurrenceReportViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OccurrenceReport.objects.none()
    serializer_class = OccurrenceReportSerializer
    lookup_field = "id"

    def get_queryset(self):
        user = self.request.user
        if is_internal(self.request):  # user.is_authenticated():
            qs = OccurrenceReport.objects.all()
            return qs
        elif is_customer(self.request):
            # user_orgs = [org.id for org in user.boranga_organisations.all()]
            qs = OccurrenceReport.objects.filter(Q(submitter=user.id))
            return qs
        logger.warn(
            "User is neither customer nor internal user: {} <{}>".format(
                user.get_full_name(), user.email
            )
        )
        return OccurrenceReport.objects.none()

    def get_serializer_class(self):
        if is_internal(self.request):
            return InternalOccurrenceReportSerializer
        return OccurrenceReportSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        group_type_id = GroupType.objects.get(id=request.data.get("group_type_id"))

        new_instance = OccurrenceReport(
            submitter=request.user.id,
            group_type=group_type_id,
        )
        if is_internal(request):
            new_instance.internal_application = True

        new_instance.save(version_user=request.user)
        data = {"occurrence_report_id": new_instance.id}

        # create Location for new instance
        serializer = SaveLocationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatComposition for new instance
        serializer = SaveOCRHabitatCompositionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatCondition for new instance
        serializer = SaveOCRHabitatConditionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create FireHistory for new instance
        serializer = SaveOCRFireHistorySerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create FireHistory for new instance
        serializer = SaveOCRAssociatedSpeciesSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create ObservationDetail for new instance
        serializer = SaveOCRObservationDetailSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create PlantCount for new instance
        serializer = SaveOCRPlantCountSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create AnimalObservation for new instance
        serializer = SaveOCRAnimalObservationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create Identification for new instance
        serializer = SaveOCRIdentificationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        serialized_obj = CreateOccurrenceReportSerializer(new_instance)
        return Response(serialized_obj.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="transform-geometry",
    )
    def transform_geometry(self, request, *args, **kwargs):
        geometry = request.GET.get("geometry", None)
        from_srid = int(request.GET.get("from", 4326))
        to_srid = int(request.GET.get("to", 4326))

        if not geometry:
            return HttpResponse({}, content_type="application/json")

        json_geom = json.loads(geometry)

        transformed = transform_json_geometry(json_geom, from_srid, to_srid)

        return HttpResponse(transformed, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="spatially-process-geometries",
    )
    def spatially_process_geometries(self, request, *args, **kwargs):
        geometry = request.GET.get("geometry", None)
        operation = request.GET.get("operation", None)
        parameters = request.GET.get("parameters", None)
        parameters = [float(p) for p in parameters.split(",")] if parameters else []
        unit = request.GET.get("unit", None)

        if not geometry:
            raise serializers.ValidationError("Geometry is required")
        if not operation:
            raise serializers.ValidationError("Operation is required")
        if not unit:
            raise serializers.ValidationError("Unit is required")

        try:
            res_json = spatially_process_geometry(
                json.loads(geometry), operation, parameters, unit
            )
        except Exception as e:
            raise e
        else:
            return HttpResponse(res_json, content_type="application/json")

    # used for Location Tab of Occurrence Report external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="location-list-of-values",
    )
    def location_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        qs = self.get_queryset()
        datum_list = []

        id = request.GET.get("id", None)
        try:
            qs = qs.get(id=id)
        except OccurrenceReport.DoesNotExist:
            logger.error(f"Occurrence Report with id {id} not found")
        else:
            ocr_geometries = qs.ocr_geometry.all().exclude(**{"geometry": None})
            epsg_codes = [
                str(g.srid)
                for g in ocr_geometries.values_list("geometry", flat=True).distinct()
            ]
            # Add the srids of the original geometries to epsg_codes
            original_geometry_srids = [
                str(g.original_geometry_srid) for g in ocr_geometries
            ]
            epsg_codes += [g for g in original_geometry_srids if g.isnumeric()]
            epsg_codes = list(set(epsg_codes))
            datum_list = search_datums("", codes=epsg_codes)

        coordination_source_list = []
        values = CoordinationSource.objects.all()
        if values:
            for val in values:
                coordination_source_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        location_accuracy_list = []
        values = LocationAccuracy.objects.all()
        if values:
            for val in values:
                location_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "datum_list": datum_list,
            "coordination_source_list": coordination_source_list,
            "location_accuracy_list": location_accuracy_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    # used for Occurrence Report external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def list_of_values(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        land_form_list = []
        types = LandForm.objects.all()
        if types:
            for val in types:
                land_form_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        rock_type_list = []
        types = RockType.objects.all()
        if types:
            for val in types:
                rock_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_type_list = []
        types = SoilType.objects.all()
        if types:
            for val in types:
                soil_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_colour_list = []
        colours = SoilColour.objects.all()
        if colours:
            for val in colours:
                soil_colour_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_condition_list = []
        conditions = SoilCondition.objects.all()
        if conditions:
            for val in conditions:
                soil_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        drainage_list = []
        drainages = Drainage.objects.all()
        if drainages:
            for val in drainages:
                drainage_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        intensity_list = []
        intensities = Intensity.objects.all()
        if intensities:
            for val in intensities:
                intensity_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "land_form_list": land_form_list,
            "rock_type_list": rock_type_list,
            "soil_type_list": soil_type_list,
            "soil_colour_list": soil_colour_list,
            "soil_condition_list": soil_condition_list,
            "drainage_list": drainage_list,
            "intensity_list": intensity_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    # used for Occurrence Report Observation external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def observation_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        observation_method_list = []
        values = ObservationMethod.objects.all()
        if values:
            for val in values:
                observation_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_method_list = []
        values = PlantCountMethod.objects.all()
        if values:
            for val in values:
                plant_count_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_accuracy_list = []
        values = PlantCountAccuracy.objects.all()
        if values:
            for val in values:
                plant_count_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_condition_list = []
        values = PlantCondition.objects.all()
        if values:
            for val in values:
                plant_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        counted_subject_list = []
        values = CountedSubject.objects.all()
        if values:
            for val in values:
                counted_subject_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        primary_detection_method_list = []
        values = PrimaryDetectionMethod.objects.all()
        if values:
            for val in values:
                primary_detection_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        secondary_sign_list = []
        values = SecondarySign.objects.all()
        if values:
            for val in values:
                secondary_sign_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        reprod_maturity_list = []
        values = ReproductiveMaturity.objects.all()
        if values:
            for val in values:
                reprod_maturity_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        death_reason_list = []
        values = DeathReason.objects.all()
        if values:
            for val in values:
                death_reason_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        animal_health_list = []
        values = AnimalHealth.objects.all()
        if values:
            for val in values:
                animal_health_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        identification_certainty_list = []
        values = IdentificationCertainty.objects.all()
        if values:
            for val in values:
                identification_certainty_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_type_list = []
        values = SampleType.objects.all()
        if values:
            for val in values:
                sample_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_dest_list = []
        values = SampleDestination.objects.all()
        if values:
            for val in values:
                sample_dest_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        permit_type_list = []
        values = PermitType.objects.all()
        if values:
            for val in values:
                permit_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "observation_method_list": observation_method_list,
            "plant_count_method_list": plant_count_method_list,
            "plant_count_accuracy_list": plant_count_accuracy_list,
            "plant_condition_list": plant_condition_list,
            "counted_subject_list": counted_subject_list,
            "primary_detection_method_list": primary_detection_method_list,
            "secondary_sign_list": secondary_sign_list,
            "reprod_maturity_list": reprod_maturity_list,
            "death_reason_list": death_reason_list,
            "animal_health_list": animal_health_list,
            "identification_certainty_list": identification_certainty_list,
            "sample_type_list": sample_type_list,
            "sample_dest_list": sample_dest_list,
            "permit_type_list": permit_type_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    def is_authorised_to_update(self):
        #To update an occurrence report, the user must be:
        # - the original submitter and the OCR in draft or
        # - an internal assessor and the OCR under assessment or
        instance = self.get_object()
        user = self.request.user
        if not ((instance.can_user_edit and (
            user.id == instance.submitter #or 
            #(instance.internal_application and is_internal(self.request))
        )) or (instance.has_assessor_mode(user)) or (instance.has_unlocked_mode(user))):
            raise serializers.ValidationError("User not authorised to update Occurrence Report")

    def is_authorised_to_assign(self, assigner, assignee=None):
        #To assign a report:
        # - the report must be under assessment, the assigner must be in the assessment group, and the assignee must be in the assessment group or
        # - the report must be under approval, the assigner must be in the approver group, and the assignee must be in the approval group
        # AND the Assignee must be the proposed assignee, or already assigned
        instance = self.get_object()

        in_assessor_group = assignee and (assignee.id in instance.get_assessor_group().get_system_group_member_ids())
        in_approver_group = assignee and (assignee.id in instance.get_approver_group().get_system_group_member_ids())
        
        self_assigning = assigner == assignee
        
        assigner_assigned = instance.assigned_officer == assigner.id
        assigner_approver = instance.assigned_approver == assigner.id

        if (instance.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
            OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
            OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
        ]) and (\
            (self_assigning and (in_assessor_group or in_approver_group)) or \
            (not(assignee) and assigner_assigned and instance.has_assessor_mode(assigner)) or \
            ((in_assessor_group or in_approver_group) and assigner_assigned and instance.has_assessor_mode(assigner)) \
        ):
            return
        elif (instance.processing_status in [
            OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
        ]) and (\
            (self_assigning and in_approver_group) or \
            (not(assignee) and assigner_approver and instance.has_approver_mode(assigner)) or \
            ((in_approver_group) and assigner_assigned and instance.has_assessor_mode(assigner)) \
        ):
            return

        raise serializers.ValidationError("User not authorised to manage assignments for Occurrence Report")

    def is_authorised_to_assess(self):
        instance = self.get_object()
        user = self.request.user
        if not instance.has_assessor_mode(user):
            raise serializers.ValidationError("User not authorised to make Assessment Actions for Occurrence Report")

    def is_authorised_to_approve(self):
        instance = self.get_object()
        user = self.request.user
        if not instance.has_approver_mode(user):
            raise serializers.ValidationError("User not authorised to make Approval Actions for Occurrence Report")

    def is_authorised_to_change_lock(self):
        instance = self.get_object()
        user = self.request.user

        if not instance.can_change_lock(user):
            raise serializers.ValidationError("User not authorised to change lock status for Occurrence Report")

    def unlocked_back_to_assessor(self):
        instance = self.get_object()
        request = self.request
        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            serializer = BackToAssessorSerializer(data={"reason":"Change made after unlock"})
            serializer.is_valid(raise_exception=True)
            instance.back_to_assessor(request, serializer.validated_data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def lock_occurrence_report(self, request, *args, **kwargs):
        self.is_authorised_to_change_lock()
        instance = self.get_object()
        instance.lock(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def unlock_occurrence_report(self, request, *args, **kwargs):
        self.is_authorised_to_change_lock()
        instance = self.get_object()
        instance.unlock(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_location_details(self, request, *args, **kwargs):
        
        self.is_authorised_to_update()
        ocr_instance = self.get_object()

        location_instance, created = Location.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # species_id saved seperately as its not field of Location but OCR
        species = request.data.get("species_id")
        ocr_instance.species_id = species
        # ocr_instance.save()
        # community_id saved seperately as its not field of Location but OCR
        community = request.data.get("community_id")
        ocr_instance.community_id = community

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()
        else:
            ocr_instance.save(version_user=request.user)

        # ocr geometry data to save seperately
        geometry_data = request.data.get("ocr_geometry")
        if geometry_data:
            save_geometry(request, ocr_instance, geometry_data)

        # print(request.data.get('geojson_polygon'))
        # polygon = request.data.get('geojson_polygon')
        # if polygon:
        #     coords_list = [list(map(float, coord.split(' '))) for coord in polygon.split(',')]
        #     coords_list.append(coords_list[0])
        #     request.data['geojson_polygon'] = GEOSGeometry(f'POLYGON(({", ".join(map(lambda
        # x: " ".join(map(str, x)), coords_list))}))')

        # the request.data is only the habitat composition data thats been sent from front end
        location_data = request.data.get("location")
        serializer = SaveLocationSerializer(
            location_instance, data=location_data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_composition_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        
        habitat_instance, created = OCRHabitatComposition.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCRHabitatCompositionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_condition_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()        
        ocr_instance = self.get_object()
        habitat_instance, created = OCRHabitatCondition.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat condition data thats been sent from front end
        serializer = SaveOCRHabitatConditionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_fire_history_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        fire_instance, created = OCRFireHistory.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCRFireHistorySerializer(
            fire_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_associated_species_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        assoc_species_instance, created = OCRAssociatedSpecies.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCRAssociatedSpeciesSerializer(
            assoc_species_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_observation_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()        
        ocr_instance = self.get_object()
        obs_det_instance, created = OCRObservationDetail.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the observation detail data thats been sent from front end
        serializer = SaveOCRObservationDetailSerializer(
            obs_det_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_plant_count_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()        
        ocr_instance = self.get_object()
        plant_count_instance, created = OCRPlantCount.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the plant count data thats been sent from front end
        serializer = SaveOCRPlantCountSerializer(
            plant_count_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_animal_observation_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        animal_obs_instance, created = OCRAnimalObservation.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the animal obs data thats been sent from front end
        serializer = SaveOCRAnimalObservationSerializer(
            animal_obs_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_identification_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()        
        ocr_instance = self.get_object()
        identification_instance, created = OCRIdentification.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the identification data thats been sent from front end
        serializer = SaveOCRIdentificationSerializer(
            identification_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    # used for observer detail datatable on location tab
    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def observer_details(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.observer_detail.all()
        serializer = OCRObserverDetailSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @list_route(methods=["GET"], detail=False)
    def list_for_map(self, request, *args, **kwargs):
        """Returns the proposals for the map"""
        proposal_ids = [
            int(id)
            for id in request.query_params.get("proposal_ids", "").split(",")
            if id.lstrip("-").isnumeric()
        ]
        # application_type = request.query_params.get("application_type", None)
        # processing_status = request.query_params.get("processing_status", None)

        # cache_key = settings.CACHE_KEY_MAP_PROPOSALS
        # qs = cache.get(cache_key)
        # priya added qs=None as we don't have cache data yet
        qs = None
        if qs is None:
            qs = (
                self.get_queryset()
                .exclude(ocr_geometry__isnull=True)
                .prefetch_related("ocr_geometry")
            )
            # cache.set(cache_key, qs, settings.CACHE_TIMEOUT_2_HOURS)

        if len(proposal_ids) > 0:
            qs = qs.filter(id__in=proposal_ids)

        # if (
        #     application_type
        #     and application_type.isnumeric()
        #     and int(application_type) > 0
        # ):
        #     qs = qs.filter(application_type_id=application_type)

        # if processing_status:
        #     qs = qs.filter(processing_status=processing_status)

        # qs = self.filter_queryset(qs)
        serializer = ListOCRReportMinimalSerializer(
            qs, context={"request": request}, many=True
        )
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def draft(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        instance = self.get_object()
        # request_data = request.data
        proposal_data = (
            request.data.get("proposal") if request.data.get("proposal") else {}
        )
        # request.data['submitter'] = u'{}'.format(request.user.id)
        if "submitter" in proposal_data and proposal_data["submitter"]:
            request.data.get("proposal")["submitter"] = "{}".format(
                proposal_data["submitter"].get("id")
            )
        if proposal_data.get("habitat_composition"):
            habitat_instance, created = OCRHabitatComposition.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRHabitatCompositionSerializer(
                habitat_instance, data=proposal_data.get("habitat_composition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("habitat_condition"):
            hab_cond_instance, created = OCRHabitatCondition.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRHabitatConditionSerializer(
                hab_cond_instance, data=proposal_data.get("habitat_condition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("fire_history"):
            fire_instance, created = OCRFireHistory.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRFireHistorySerializer(
                fire_instance, data=proposal_data.get("fire_history")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("associated_species"):
            assoc_species_instance, created = (
                OCRAssociatedSpecies.objects.get_or_create(occurrence_report=instance)
            )
            serializer = SaveOCRAssociatedSpeciesSerializer(
                assoc_species_instance,
                data=proposal_data.get("associated_species"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("observation_detail"):
            obs_det_instance, created = OCRObservationDetail.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRObservationDetailSerializer(
                obs_det_instance, data=proposal_data.get("observation_detail")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("plant_count"):
            plant_count_instance, created = OCRPlantCount.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRPlantCountSerializer(
                plant_count_instance, data=proposal_data.get("plant_count")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("animal_observation"):
            animal_obs_instance, created = OCRAnimalObservation.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRAnimalObservationSerializer(
                animal_obs_instance,
                data=proposal_data.get("animal_observation"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("identification"):
            identification_instance, created = OCRIdentification.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRIdentificationSerializer(
                identification_instance,
                data=proposal_data.get("identification"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("location"):
            location_instance, created = Location.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveLocationSerializer(
                location_instance, data=proposal_data.get("location")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        # ocr geometry data to save seperately
        geometry_data = proposal_data.get("ocr_geometry", None)
        if geometry_data:
            save_geometry(request, instance, geometry_data)

        serializer = SaveOccurrenceReportSerializer(
            instance, data=proposal_data, partial=True
        )

        serializer.is_valid(raise_exception=True)
        if serializer.is_valid():
            if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
                saved_instance = serializer.save(no_revision=True)
                self.unlocked_back_to_assessor()
            else:
                saved_instance = serializer.save(version_user=request.user)

        # return redirect(reverse('external'))

        final_instance = self.get_object()
        serializer = self.get_serializer(final_instance)
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    def submit(self, request, *args, **kwargs):

        self.is_authorised_to_update()        
        instance = self.get_object()
        # instance.submit(request,self)
        ocr_proposal_submit(instance, request)
        instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
        # return redirect(reverse('external'))

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def action_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.action_logs.all()
        serializer = OccurrenceReportUserActionSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.comms_logs.all()
        serializer = OccurrenceReportLogEntrySerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def add_comms_log(self, request, *args, **kwargs):

        if not is_internal(self):
            raise serializers.ValidationError("User not authorised to add Communication Logs to the Occurrence Report")

        instance = self.get_object()
        mutable = request.data._mutable
        request.data._mutable = True
        request.data["occurrence_report"] = f"{instance.id}"
        request.data["staff"] = f"{request.user.id}"
        request.data._mutable = mutable
        serializer = OccurrenceReportLogEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comms = serializer.save()
        # Save the files
        for f in request.FILES:
            document = comms.documents.create()
            document.name = str(request.FILES[f])
            document._file = request.FILES[f]
            document.save()
        # End Save Documents

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.documents.all()
        if is_internal(self.request):
            qs = instance.documents.all()
        elif is_customer(self.request):
            qs = instance.documents.filter(Q(uploaded_by=request.user.id))
        # qs = qs.exclude(input_name='occurrence_report_approval_doc')
        # TODO do we need/not to show approval doc in cs documents tab
        qs = qs.order_by("-uploaded_date")
        serializer = OccurrenceReportDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def threats(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.ocr_threats.all()
        if is_internal(self.request):
            qs = instance.ocr_threats.all()
        elif is_customer(self.request):
            # TODO Do we need to sort the threats for external user (similar like documents)
            # qs = qs.filter(Q(uploaded_by=request.user.id))
            qs = instance.ocr_threats.all()
        filter_backend = OCCConservationThreatFilterBackend()
        qs = filter_backend.filter_queryset(self.request,qs,self)
        serializer = OCRConservationThreatSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["POST"], detail=True)
    @renderer_classes((JSONRenderer,))
    def process_shapefile_document(self, request, *args, **kwargs):
        instance = self.get_object()
        returned_data = None
        returned_data = process_shapefile_document(request, instance)
        if returned_data:
            return Response(returned_data)
        else:
            return Response({})

    @detail_route(methods=["POST"], detail=True)
    @renderer_classes((JSONRenderer,))
    def validate_map_files(self, request, *args, **kwargs):
        self.is_authorised_to_update()    
        instance = self.get_object()
        validate_map_files(request, instance, "occurrence_report")
        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()
            instance.save(no_revision=True)
        else:
            instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)
        logger.debug(f"validate_map_files response: {serializer.data}")
        
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def assign_request_user(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_assign(request.user,request.user)
        instance.assign_officer(request, request.user)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def assign_to(self, request, *args, **kwargs):

        instance = self.get_object()
        user_id = request.data.get("assessor_id", None)
        user = None
        if not user_id:
            raise serializers.ValidationError("An assessor id is required")
        try:
            user = EmailUser.objects.get(id=user_id)
        except EmailUser.DoesNotExist:
            raise serializers.ValidationError(
                "A user with the id passed in does not exist"
            )
        assigner = self.request.user
        self.is_authorised_to_assign(assigner,user)
        instance.assign_officer(request, user)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def unassign(self, request, *args, **kwargs):

        user = self.request.user
        self.is_authorised_to_assign(user)

        instance = self.get_object()
        instance.unassign(request)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def amendment_request(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.amendment_requests
        qs = qs.filter(status="requested")
        serializer = OccurrenceReportAmendmentRequestSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def propose_decline(self, request, *args, **kwargs):

        self.is_authorised_to_assess()

        instance = self.get_object()
        serializer = ProposeDeclineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.propose_decline(request, serializer.validated_data)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def decline(self, request, *args, **kwargs):

        self.is_authorised_to_approve()

        instance = self.get_object()

        original_occ = instance.occurrence

        serializer = ProposeDeclineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.decline(request, serializer.validated_data) #ensure occ set to None

        #run occ check
        if original_occ:
            original_occ.check_ocr_count_for_discard(request)

        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def back_to_assessor(self, request, *args, **kwargs):
        instance = self.get_object()
        
        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER:
            self.is_authorised_to_approve()
        elif instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.is_authorised_to_update()
        
        serializer = BackToAssessorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.back_to_assessor(request, serializer.validated_data)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def propose_approve(self, request, *args, **kwargs):

        self.is_authorised_to_assess()

        instance = self.get_object()
        serializer = ProposeApproveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.propose_approve(request, serializer.validated_data)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def approve(self, request, *args, **kwargs):

        self.is_authorised_to_approve()

        instance = self.get_object()

        original_occ = instance.occurrence

        instance.approve(request)

        if original_occ and original_occ.id != instance.occurrence.id:
            original_occ.check_ocr_count_for_discard(request)

        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    # used on referral form
    @detail_route(methods=["post"], detail=True)
    def send_referral(self, request, *args, **kwargs):
        self.is_authorised_to_assess()
        instance = self.get_object()
        serializer = SendReferralSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        instance.send_referral(
            request,
            serializer.validated_data["email"],
            serializer.validated_data["text"],
        )
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def referrals(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.referrals.all()
        serializer = InternalOccurrenceReportReferralSerializer(qs, many=True)
        return Response(serializer.data)


class ObserverDetailViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OCRObserverDetail.objects.none()
    serializer_class = OCRObserverDetailSerializer

    def is_authorised_to_update(self,occurrence_report):
        user = self.request.user
        if not ((occurrence_report.can_user_edit and (
            user.id == occurrence_report.submitter #or 
            #(occurrence_report.internal_application and is_internal(self.request))
        )) or (occurrence_report.has_assessor_mode(user)) or (occurrence_report.has_unlocked_mode(user))):
            raise serializers.ValidationError("User not authorised to update Occurrence Report")

    def unlocked_back_to_assessor(self,occurrence_report):
        request = self.request
        if occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            serializer = BackToAssessorSerializer(data={"reason":"Change made after unlock"})
            serializer.is_valid(raise_exception=True)
            occurrence_report.back_to_assessor(request, serializer.validated_data)

    def get_queryset(self):
        qs = OCRObserverDetail.objects.none()

        if is_internal(self.request):
            qs = OCRObserverDetail.objects.all().order_by("id")
        elif is_customer(self.request):
            # not sure what qs it should be for api security check
            qs = OCRObserverDetail.objects.all().order_by("id")
            return qs
        return qs

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        serializer = OCRObserverDetailSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report)    
        # instance.community.log_user_action(CommunityUserAction.ACTION_ADD_THREAT.format(instance.threat_number,instance.community.community_number),request)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = OCRObserverDetailSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        self.is_authorised_to_update(occurrence_report)
        serializer.save()

        if occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(occurrence_report)

        return Response(serializer.data)
    
    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = False
        instance.save()

        serializer = self.get_serializer(instance)
        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = True
        instance.save()

        serializer = self.get_serializer(instance)
        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
        return Response(serializer.data)


class OccurrenceReportAmendmentRequestViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OccurrenceReportAmendmentRequest.objects.none()
    serializer_class = OccurrenceReportAmendmentRequestSerializer

    def get_queryset(self):
        if is_internal(self.request):  # user.is_authenticated():
            qs = OccurrenceReportAmendmentRequest.objects.all().order_by("id")
            return qs
        return OccurrenceReportAmendmentRequest.objects.none()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=json.loads(request.data.get("data")))
        serializer.is_valid(raise_exception=True)
        user = self.request.user
        occurrence_report = serializer.validated_data["occurrence_report"]
        if not (occurrence_report.has_assessor_mode(user)):
            raise serializers.ValidationError("User not authorised to update Occurrence Report")
        instance = serializer.save()
        instance.add_documents(request)
        instance.generate_amendment(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    def delete_document(self, request, *args, **kwargs):
        instance = self.get_object()

        user = self.request.user
        occurrence_report = instance.occurrence_report
        if not (occurrence_report.has_assessor_mode(user)):
            raise serializers.ValidationError("User not authorised to update Occurrence Report")

        OccurrenceReportAmendmentRequestDocument.objects.get(
            id=request.data.get("id")
        ).delete()
        return Response(
            [
                dict(id=i.id, name=i.name, _file=i._file.url)
                for i in instance.cs_amendment_request_documents.all()
            ]
        )


class OccurrenceReportDocumentViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OccurrenceReportDocument.objects.none()
    serializer_class = OccurrenceReportDocumentSerializer

    def is_authorised_to_update(self,occurrence_report):
        user = self.request.user
        if not ((occurrence_report.can_user_edit and (
            user.id == occurrence_report.submitter #or 
            #(occurrence_report.internal_application and is_internal(self.request))
        )) or (occurrence_report.has_assessor_mode(user)) or (occurrence_report.has_unlocked_mode(user))):
            raise serializers.ValidationError("User not authorised to update Occurrence Report")

    def unlocked_back_to_assessor(self,occurrence_report):
        request = self.request
        if occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            serializer = BackToAssessorSerializer(data={"reason":"Change made after unlock"})
            serializer.is_valid(raise_exception=True)
            occurrence_report.back_to_assessor(request, serializer.validated_data)

    def get_queryset(self):
        request_user = self.request.user
        qs = OccurrenceReportDocument.objects.none()

        if is_internal(self.request):
            qs = OccurrenceReportDocument.objects.all().order_by("id")
        elif is_customer(self.request):
            qs = OccurrenceReportDocument.objects.filter(Q(uploaded_by=request_user.id))
            return qs
        return qs

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_DISCARD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_REINSTATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        serializer = SaveOccurrenceReportDocumentSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(no_revision=True)
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_UPDATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOccurrenceReportDocumentSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        self.is_authorised_to_update(occurrence_report)
        instance = serializer.save(no_revision=True)
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_ADD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
        return Response(serializer.data)


class OCRConservationThreatFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):

        total_count = queryset.count()

        filter_threat_category = request.GET.get("filter_threat_category")
        if filter_threat_category and not filter_threat_category.lower() == "all":
            queryset = queryset.filter(threat_category_id=filter_threat_category)

        filter_threat_current_impact = request.GET.get("filter_threat_current_impact")
        if filter_threat_current_impact and not filter_threat_current_impact.lower() == "all":
            queryset = queryset.filter(current_impact=filter_threat_current_impact)

        filter_threat_potential_impact = request.GET.get("filter_threat_potential_impact")
        if filter_threat_potential_impact and not filter_threat_potential_impact.lower() == "all":
            queryset = queryset.filter(potential_impact=filter_threat_potential_impact)

        filter_threat_status = request.GET.get(
            "filter_threat_status"
        )
        if (
            filter_threat_status
            and not filter_threat_status.lower() == "all"
        ):
            if filter_threat_status == "active":
                queryset = queryset.filter(visible=True)
            elif filter_threat_status == "removed":
                queryset = queryset.filter(visible=False)

        def get_date(filter_date):
            date = request.GET.get(filter_date)
            if date:
                date = datetime.strptime(date, "%Y-%m-%d")
            return date
        
        filter_observed_from_date = get_date("filter_observed_from_date")
        if filter_observed_from_date:
            queryset = queryset.filter(date_observed__gte=filter_observed_from_date)

        filter_observed_to_date = get_date("filter_observed_to_date")
        if filter_observed_to_date:
            queryset = queryset.filter(date_observed__lte=filter_observed_to_date)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OCRConservationThreatViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OCRConservationThreat.objects.none()
    serializer_class = OCRConservationThreatSerializer

    def is_authorised_to_update(self,occurrence_report):
        user = self.request.user
        if not ((occurrence_report.can_user_edit and (
            user.id == occurrence_report.submitter #or 
            #(occurrence_report.internal_application and is_internal(self.request))
        )) or (occurrence_report.has_assessor_mode(user)) or (occurrence_report.has_unlocked_mode(user))):
            raise serializers.ValidationError("User not authorised to update Occurrence Report")

    def unlocked_back_to_assessor(self,occurrence_report):
        request = self.request
        if occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            serializer = BackToAssessorSerializer(data={"reason":"Change made after unlock"})
            serializer.is_valid(raise_exception=True)
            occurrence_report.back_to_assessor(request, serializer.validated_data)

    def get_queryset(self):
        request_user = self.request.user
        qs = OCRConservationThreat.objects.none()

        if is_internal(self.request):
            qs = OCRConservationThreat.objects.all().order_by("id")
        elif is_customer(self.request):
            # TODO filter qs as per added_by - using the OCR submitter for now
            qs = OCRConservationThreat.objects.filter(occurrence_report__submitter=request_user.id).order_by("id")
            return qs
        return qs

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_DISCARD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_REINSTATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 
            
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        serializer = SaveOCRConservationThreatSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        serializer.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_UPDATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 

        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOCRConservationThreatSerializer(
            data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        self.is_authorised_to_update(occurrence_report)
        instance = serializer.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_ADD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if instance.occurrence_report.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor(instance.occurrence_report) 

        return Response(serializer.data)


class GetOCCProfileDict(views.APIView):
    def get(self, request, format=None):
        group_type = request.GET.get("group_type", "")

        species_list = []
        if group_type:
            exclude_status = ["draft"]
            species = Species.objects.filter(
                ~Q(processing_status__in=exclude_status)
                & ~Q(taxonomy=None)
                & Q(group_type__name=group_type)
            )
            if species:
                for specimen in species:
                    species_list.append(
                        {
                            "id": specimen.id,
                            "name": specimen.taxonomy.scientific_name,
                            "taxon_previous_name": specimen.taxonomy.taxon_previous_name,
                            "common_name": specimen.taxonomy.taxon_vernacular_name,
                        }
                    )
        community_list = []
        exculde_status = ["draft"]
        communities = CommunityTaxonomy.objects.filter(
            ~Q(community__processing_status__in=exculde_status)
        )  # TODO remove later as every community will have community name
        if communities:
            for specimen in communities:
                community_list.append(
                    {
                        "id": specimen.community.id,
                        "name": specimen.community_name,
                    }
                )

        occurrence_source_list = list(
            OccurrenceSource.objects.all().values("id", "name")
        )
        wild_status_list = list(WildStatus.objects.all().values("id", "name"))

        res_json = {
            "species_list": species_list,
            "community_list": community_list,
            "source_list": occurrence_source_list,
            "wild_status_list": wild_status_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class OccurrenceFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        logger.debug(f"OccurrenceFilterBackend:filter_queryset: {view.name}")

        total_count = queryset.count()

        filter_group_type = request.GET.get("filter_group_type")
        if filter_group_type and not filter_group_type.lower() == "all":
            queryset = queryset.filter(group_type__name=filter_group_type)

        filter_occurrence_name = request.GET.get("filter_occurrence_name")
        if filter_occurrence_name and not filter_occurrence_name.lower() == "all":
            queryset = queryset.filter(occurrence_name=filter_occurrence_name)

        filter_scientific_name = request.GET.get("filter_scientific_name")
        if filter_scientific_name and not filter_scientific_name.lower() == "all":
            queryset = queryset.filter(species__taxonomy__id=filter_scientific_name)

        filter_community_name = request.GET.get("filter_community_name")
        if filter_community_name and not filter_community_name.lower() == "all":
            queryset = queryset.filter(community__taxonomy__id=filter_community_name)

        filter_status = request.GET.get("filter_status")
        if filter_status and not filter_status.lower() == "all":
            queryset = queryset.filter(processing_status=filter_status)

        filter_from_effective_from_date = request.GET.get("filter_from_effective_from_date")
        filter_to_effective_from_date = request.GET.get("filter_to_effective_from_date")

        filter_from_effective_to_date = request.GET.get("filter_from_effective_to_date")
        filter_to_effective_to_date = request.GET.get("filter_to_effective_to_date")

        if filter_from_effective_from_date:
            queryset = queryset.filter(
                effective_from__gte=filter_from_effective_from_date
            )
        if filter_to_effective_from_date:
            queryset = queryset.filter(
                effective_from__lte=filter_to_effective_from_date
            )

        if filter_from_effective_to_date:
            queryset = queryset.filter(
                effective_to__gte=filter_from_effective_to_date
            )
        if filter_to_effective_to_date:
            queryset = queryset.filter(
                effective_to__lte=filter_to_effective_to_date
            )

        filter_from_review_due_date = request.GET.get("filter_from_review_due_date")
        filter_to_review_due_date = request.GET.get("filter_to_review_due_date")

        if filter_from_review_due_date:
            queryset = queryset.filter(
                review_due_date__gte=filter_from_review_due_date
            )
        if filter_to_review_due_date:
            queryset = queryset.filter(
                review_due_date__lte=filter_to_review_due_date
            )

        fields = self.get_fields(request)

        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OccurrencePaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    pagination_class = DatatablesPageNumberPagination
    queryset = Occurrence.objects.none()
    serializer_class = OccurrenceSerializer
    page_size = 10
    filter_backends = (OccurrenceFilterBackend,)

    def get_serializer_class(self):
        if self.action in ["list", "occurrence_internal", "occurrence_external"]:
            return ListOccurrenceSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = Occurrence.objects.all()
        if is_customer(self.request):
            qs = qs.filter(submitter=self.request.user.id)
        return qs

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListOccurrenceSerializer(
            result_page, context={"request": request}, many=True
        )

        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_internal_export(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "species",
            "scientific_name",
            "reported_date",
            "submitter",
            "processing_status",
            "occurrence_report_number",
        ]

        serializer = ListInternalOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Occurrence",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Occurrence",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        group_type_id = request.GET.get("group_type_id", None)
        if group_type_id:
            queryset = queryset.filter(group_type_id=group_type_id)
        search_term = request.GET.get("term", "")
        if search_term:
            queryset = queryset.values_list("occurrence_number", flat=True)
            queryset = (
                queryset.filter(occurrence_number__icontains=search_term)
                .distinct()
                .values("id", "occurrence_number")[:10]
            )
            queryset = [
                {"id": occurrence["id"], "text": occurrence["occurrence_number"]}
                for occurrence in queryset
            ]
        return Response({"results": queryset})

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_name_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset().filter(processing_status=Occurrence.PROCESSING_STATUS_ACTIVE)
        group_type_id = request.GET.get("group_type_id", None)
        if group_type_id:
            try:
                group_type = GroupType.objects.get(id=group_type_id)
            except GroupType.DoesNotExist:
                logger.warning(f"GroupType with id {group_type_id} does not exist")
                return Response({"results": []})

            queryset = queryset.filter(group_type=group_type)
            occurrence_report_id = request.GET.get("occurrence_report_id", None)
            if occurrence_report_id:
                try:
                    occurrence_report = OccurrenceReport.objects.get(
                        id=occurrence_report_id
                    )
                except OccurrenceReport.DoesNotExist:
                    logger.warning(
                        "OccurrenceReport with id {} does not exist".format(
                            occurrence_report_id
                        )
                    )
                    return Response({"results": []})

                if group_type.name in [
                    GroupType.GROUP_TYPE_FLORA,
                    GroupType.GROUP_TYPE_FAUNA,
                ]:
                    queryset = queryset.filter(species=occurrence_report.species)
                elif group_type.name == GroupType.GROUP_TYPE_COMMUNITY:
                    queryset = queryset.filter(community=occurrence_report.community)

        search_term = request.GET.get("term", None)
        if search_term:
            if occurrence_report_id:
                queryset = (
                    queryset.annotate(
                        display_name=Concat(
                            "occurrence_number",
                            Value(" - "),
                            "occurrence_name",
                            Value(" ("),
                            "group_type__name",
                            Value(")"),
                            output_field=CharField(),
                        ),
                    )
                    .filter(display_name__icontains=search_term)
                    .distinct()
                    .values("id", "display_name")[:10]
                )
                queryset = [
                    {"id": occurrence["id"], "text": occurrence["display_name"]}
                    for occurrence in queryset
                ]
            else:
                queryset = (
                    queryset.filter(occurrence_name__icontains=search_term)
                    .distinct()
                    .values("id", "occurrence_name")[:10]
                )

                queryset = [
                    {"id": occurrence["id"], "text": occurrence["occurrence_name"]}
                    for occurrence in queryset
                ]
        return Response({"results": queryset})

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.documents.all()
        if is_internal(self.request):
            qs = instance.documents.all()
        else:
            qs = instance.documents.none()
        qs = qs.order_by("-uploaded_date")
        serializer = OccurrenceDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def threats(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_internal(self.request):
            qs = instance.occ_threats.all()
        else:
            qs = instance.occ_threats.none()
        qs = qs.order_by("-date_observed")
        serializer = OCCConservationThreatSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_occurrence_reports(self, request, *args, **kwargs):
        instance = self.get_object()
        related_reports = instance.get_related_occurrence_reports()
        if is_internal(self.request):
            related_reports = related_reports.all()
        else:
            related_reports = related_reports.none()
        print(related_reports)
        serializer = ListInternalOccurrenceReportSerializer(
            related_reports, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_items(self, request, *args, **kwargs):
        instance = self.get_object()
        related_filter_type = request.GET.get("related_filter_type")
        related_items = instance.get_related_items(related_filter_type)
        serializer = RelatedItemsSerializer(related_items, many=True)
        return Response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def filter_list(self, request, *args, **kwargs):
        """Used by the Related Items dashboard filters"""
        related_type = Occurrence.RELATED_ITEM_CHOICES
        res_json = json.dumps(related_type)
        return HttpResponse(res_json, content_type="application/json")


class OccurrenceDocumentViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OccurrenceDocument.objects.none()
    serializer_class = OccurrenceDocumentSerializer

    def get_queryset(self):
        qs = OccurrenceDocument.objects.none()

        if is_internal(self.request):
            qs = OccurrenceDocument.objects.all().order_by("id")

        return qs

    def is_authorised_to_update(self, occurrence):
        user = self.request.user
        if not (user.id in occurrence.get_occurrence_approver_group().get_system_group_member_ids()):
            raise serializers.ValidationError("User not authorised to update Occurrence")

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_DISCARD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_REINSTATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        serializer = SaveOccurrenceDocumentSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(no_revision=True)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_UPDATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOccurrenceDocumentSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        self.is_authorised_to_update(serializer.validated_data["occurrence"])
        instance = serializer.save(no_revision=True)
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_ADD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        return Response(serializer.data)

class OCCConservationThreatFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):

        total_count = queryset.count()

        filter_threat_source = request.GET.get("filter_threat_source")
        if filter_threat_source and not filter_threat_source.lower() == "all":
            queryset = queryset.filter((Q(occurrence__occurrence_number=filter_threat_source) & 
            Q(occurrence_report_threat__occurrence_report=None))|
            Q(occurrence_report_threat__occurrence_report__occurrence_report_number=filter_threat_source))

        filter_threat_category = request.GET.get("filter_threat_category")
        if filter_threat_category and not filter_threat_category.lower() == "all":
            queryset = queryset.filter(threat_category_id=filter_threat_category)

        filter_threat_current_impact = request.GET.get("filter_threat_current_impact")
        if filter_threat_current_impact and not filter_threat_current_impact.lower() == "all":
            queryset = queryset.filter(current_impact=filter_threat_current_impact)

        filter_threat_potential_impact = request.GET.get("filter_threat_potential_impact")
        if filter_threat_potential_impact and not filter_threat_potential_impact.lower() == "all":
            queryset = queryset.filter(potential_impact=filter_threat_potential_impact)

        filter_threat_status = request.GET.get(
            "filter_threat_status"
        )
        if (
            filter_threat_status
            and not filter_threat_status.lower() == "all"
        ):
            if filter_threat_status == "active":
                queryset = queryset.filter(visible=True)
            elif filter_threat_status == "removed":
                queryset = queryset.filter(visible=False)

        def get_date(filter_date):
            date = request.GET.get(filter_date)
            if date:
                date = datetime.strptime(date, "%Y-%m-%d")
            return date
        
        filter_observed_from_date = get_date("filter_observed_from_date")
        if filter_observed_from_date:
            queryset = queryset.filter(date_observed__gte=filter_observed_from_date)

        filter_observed_to_date = get_date("filter_observed_to_date")
        if filter_observed_to_date:
            queryset = queryset.filter(date_observed__lte=filter_observed_to_date)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        try:
            queryset = super().filter_queryset(request, queryset, view)
        except Exception as e:
            print(e)
        setattr(view, "_datatables_total_count", total_count)
        return queryset

class OCCConservationThreatViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OCCConservationThreat.objects.none()
    serializer_class = OCCConservationThreatSerializer
    filter_backends = (OCCConservationThreatFilterBackend,)

    def get_queryset(self):
        qs = OCCConservationThreat.objects.none()

        if is_internal(self.request):
            qs = OCCConservationThreat.objects.all().order_by("id")

        return qs

    def is_authorised_to_update(self, occurrence):
        user = self.request.user
        if not (user.id in occurrence.get_occurrence_approver_group().get_system_group_member_ids()):
            raise serializers.ValidationError("User not authorised to update Occurrence")

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_DISCARD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_REINSTATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        serializer = SaveOCCConservationThreatSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        serializer.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_UPDATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOCCConservationThreatSerializer(
            data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        self.is_authorised_to_update(serializer.validated_data["occurrence"])
        instance = serializer.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_ADD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class GetWildStatus(views.APIView):
    def get(self, request, format=None):
        search_term = request.GET.get("term", "")
        if search_term:
            data = WildStatus.objects.filter(name__icontains=search_term).values(
                "id", "name"
            )[:10]
            data_transform = [
                {"id": wild_status["id"], "text": wild_status["name"]}
                for wild_status in data
            ]
            return Response({"results": data_transform})
        return Response()


class GetOccurrenceSource(views.APIView):
    def get(self, request, format=None):
        search_term = request.GET.get("term", "")
        if search_term:
            data = OccurrenceSource.objects.filter(name__icontains=search_term).values(
                "id", "name"
            )[:10]
            data_transform = [
                {"id": occurrence_source["id"], "text": occurrence_source["name"]}
                for occurrence_source in data
            ]
            return Response({"results": data_transform})
        return Response()


class OccurrenceViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = Occurrence.objects.none()
    serializer_class = OccurrenceSerializer
    lookup_field = "id"

    def get_queryset(self):
        qs = Occurrence.objects.all()
        if is_customer(self.request):
            qs = qs.filter(submitter=self.request.user.id)
        return qs
    
    def is_authorised_to_update(self):
        user = self.request.user
        instance = self.get_object()
        if not (user.id in instance.get_occurrence_approver_group().get_system_group_member_ids() and instance.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE):
            raise serializers.ValidationError("User not authorised to update Occurrence")

    @transaction.atomic
    def create(self, request, *args, **kwargs):

        group_type_id = GroupType.objects.get(id=request.data.get("group_type_id"))

        new_instance = Occurrence(
            submitter=request.user.id,
            group_type=group_type_id,
        )

        if not (request.user.id in new_instance.get_occurrence_approver_group().get_system_group_member_ids()):
            raise serializers.ValidationError("User not authorised to create Occurrence")

        new_instance.save(version_user=request.user)
        data = {"occurrence_id": new_instance.id}

        # create Location for new instance TODO
        # serializer = SaveLocationSerializer(data=data)
        # serializer.is_valid(raise_exception=True)
        # serializer.save()

        # create HabitatComposition for new instance
        serializer = SaveOCCHabitatCompositionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatCondition for new instance
        serializer = SaveOCCHabitatConditionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create FireHistory for new instance
        serializer = SaveOCCFireHistorySerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create AssociatedSpecies for new instance
        serializer = SaveOCCAssociatedSpeciesSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create ObservationDetail for new instance
        serializer = SaveOCCObservationDetailSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create PlantCount for new instance
        serializer = SaveOCCPlantCountSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create AnimalObservation for new instance
        serializer = SaveOCCAnimalObservationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create Identification for new instance
        serializer = SaveOCCIdentificationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        serialized_obj = CreateOccurrenceSerializer(new_instance)
        return Response(serialized_obj.data)

    
    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def lock_occurrence(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        instance.lock(request)
        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def unlock_occurrence(self, request, *args, **kwargs):
        user = request.user
        instance = self.get_object()
        if not (user.id in instance.get_occurrence_approver_group().get_system_group_member_ids() and instance.processing_status == Occurrence.PROCESSING_STATUS_LOCKED):
            raise serializers.ValidationError("User not authorised to update Occurrence")
        instance.unlock(request)
        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def close_occurrence(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        instance.close(request)
        return redirect(reverse("internal"))
    
    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def action_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.action_logs.all()
        serializer = OccurrenceUserActionSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.comms_logs.all()
        serializer = OccurrenceLogEntrySerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def add_comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        mutable = request.data._mutable
        request.data._mutable = True
        request.data["occurrence"] = f"{instance.id}"
        request.data["staff"] = f"{request.user.id}"
        request.data._mutable = mutable
        serializer = OccurrenceLogEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comms = serializer.save()
        # Save the files
        for f in request.FILES:
            document = comms.documents.create()
            document.name = str(request.FILES[f])
            document._file = request.FILES[f]
            document.save()
        # End Save Documents

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.documents.all()
        if is_internal(self.request):
            qs = instance.documents.all()
        else:
            qs = instance.documents.none()
        qs = qs.order_by("-uploaded_date")
        serializer = OccurrenceDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def threats(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_internal(self.request):
            qs = instance.occ_threats.all()
        else:
            qs = instance.occ_threats.none()
        filter_backend = OCCConservationThreatFilterBackend()
        qs = filter_backend.filter_queryset(self.request,qs,self)
        serializer = OCCConservationThreatSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    #gets all distinct threat sources for threats pertaining to a specific OCC
    def threat_source_list(self, request, *args, **kwargs):
        instance = self.get_object()
        data = []
        if is_internal(self.request):
            #distinct on OCR
            qs = instance.occ_threats.distinct("occurrence_report_threat__occurrence_report").exclude(occurrence_report_threat=None)
            #format
            data = [threat.occurrence_report_threat.occurrence_report.occurrence_report_number for threat in qs]

        #if any occ threats exist with an ocr threat, then the source must be the occ
        if instance.occ_threats.filter(occurrence_report_threat=None).exists(): 
            data.append(instance.occurrence_number)

        return Response(data)

    @detail_route(methods=["get"], detail=True)
    def get_related_occurrence_reports(self, request, *args, **kwargs):
        instance = self.get_object()
        related_reports = instance.get_related_occurrence_reports()
        if is_internal(self.request):
            related_reports = related_reports.all()
        else:
            related_reports = related_reports.none()
        serializer = ListInternalOccurrenceReportSerializer(
            related_reports, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_existing_ocr_threats(self, request, *args, **kwargs):
        instance = self.get_object()
        related_reports = instance.get_related_occurrence_reports().values_list(
            "id", flat=True
        )
        addedThreats = (
            OCCConservationThreat.objects.filter(occurrence=instance)
            .exclude(occurrence_report_threat=None)
            .values_list("occurrence_report_threat_id", flat=True)
        )
        threats = OCRConservationThreat.objects.filter(
            occurrence_report_id__in=related_reports
        ).exclude(id__in=addedThreats)
        if is_internal(self.request):
            threats = threats.all()
        else:
            threats = threats.none()
        serializer = OCRConservationThreatSerializer(
            threats, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_items(self, request, *args, **kwargs):
        instance = self.get_object()
        related_filter_type = request.GET.get("related_filter_type")
        related_items = instance.get_related_items(related_filter_type)
        serializer = RelatedItemsSerializer(related_items, many=True)
        return Response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def filter_list(self, request, *args, **kwargs):
        """Used by the Related Items dashboard filters"""
        related_type = Occurrence.RELATED_ITEM_CHOICES
        res_json = json.dumps(related_type)
        return HttpResponse(res_json, content_type="application/json")

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def occurrence_save(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        instance = self.get_object()
        request_data = request.data

        if request_data.get("habitat_composition"):
            habitat_instance, created = OCCHabitatComposition.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCHabitatCompositionSerializer(
                habitat_instance, data=request_data.get("habitat_composition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("habitat_condition"):
            hab_cond_instance, created = OCCHabitatCondition.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCHabitatConditionSerializer(
                hab_cond_instance, data=request_data.get("habitat_condition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("fire_history"):
            fire_instance, created = OCCFireHistory.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCFireHistorySerializer(
                fire_instance, data=request_data.get("fire_history")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("associated_species"):
            assoc_species_instance, created = (
                OCCAssociatedSpecies.objects.get_or_create(occurrence=instance)
            )
            serializer = SaveOCCAssociatedSpeciesSerializer(
                assoc_species_instance,
                data=request_data.get("associated_species"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("observation_detail"):
            obs_det_instance, created = OCCObservationDetail.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCObservationDetailSerializer(
                obs_det_instance, data=request_data.get("observation_detail")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("plant_count"):
            plant_count_instance, created = OCCPlantCount.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCPlantCountSerializer(
                plant_count_instance, data=request_data.get("plant_count")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("animal_observation"):
            animal_obs_instance, created = OCCAnimalObservation.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCAnimalObservationSerializer(
                animal_obs_instance,
                data=request_data.get("animal_observation"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("identification"):
            identification_instance, created = OCCIdentification.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCIdentificationSerializer(
                identification_instance,
                data=request_data.get("identification"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        #TODO: adjust and enable when OCC location ready
        #if request_data.get("location"):
        #    location_instance, created = Location.objects.get_or_create(
        #        occurrence=instance
        #    )
        #    serializer = SaveLocationSerializer(
        #        location_instance, data=request_data.get("location")
        #    )
        #    serializer.is_valid(raise_exception=True)
        #    if serializer.is_valid():
        #        serializer.save()

        # occ geometry data to save seperately TODO: determine what is need here
        #geometry_data = request_data.get("occ_geometry", None)
        #if geometry_data:
        #    save_geometry(request, instance, geometry_data)

        serializer = SaveOccurrenceReportSerializer(
            instance, data=request_data, partial=True
        )

        serializer = SaveOccurrenceSerializer(instance, data=request_data, partial=True)
        serializer.is_valid(raise_exception=True)

        if serializer.is_valid():
            serializer.save(version_user=request.user)

            instance.log_user_action(
                OccurrenceUserAction.ACTION_SAVE_OCCURRENCE.format(
                    instance.occurrence_number
                ),
                request,
            )

        return redirect(reverse("internal"))

    @detail_route(methods=["post"], detail=True)
    @transaction.atomic
    def copy_ocr_section(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        data = json.loads(request.data["data"])

        ocrId = data["occurrence_report_id"]
        section = data["section"]
        merge = data["merge"]

        ocr = OccurrenceReport.objects.get(id=ocrId)
        ocrSection = getattr(ocr, section)
        occSection = getattr(instance, section)

        section_fields = type(ocrSection)._meta.get_fields()
        for i in section_fields:
            if (
                i.name != "id"
                and i.name != "occurrence_report"
                and hasattr(occSection, i.name)
            ):
                ocrValue = getattr(ocrSection, i.name)
                if merge:
                    # if not ocrValue: #do not overwrite if None, 0, or empty string
                    #    #determine if field is one-to-many
                    #    many = False
                    # DEFERRED for now
                    pass
                else:
                    setattr(occSection, i.name, ocrValue)

        occSection.save()
        instance.save(version_user=request.user)

        serialized_obj = OccurrenceSerializer(instance, context={"request": request})
        return Response(serialized_obj.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_composition_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        habitat_instance, created = OCCHabitatComposition.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCCHabitatCompositionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_condition_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        habitat_instance, created = OCCHabitatCondition.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat condition data thats been sent from front end
        serializer = SaveOCCHabitatConditionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_fire_history_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        fire_instance, created = OCCFireHistory.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCCFireHistorySerializer(
            fire_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_associated_species_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        assoc_species_instance, created = OCCAssociatedSpecies.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCCAssociatedSpeciesSerializer(
            assoc_species_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_observation_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        obs_det_instance, created = OCCObservationDetail.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the observation detail data thats been sent from front end
        serializer = SaveOCCObservationDetailSerializer(
            obs_det_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_plant_count_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        plant_count_instance, created = OCCPlantCount.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the plant count data thats been sent from front end
        serializer = SaveOCCPlantCountSerializer(
            plant_count_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_animal_observation_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        animal_obs_instance, created = OCCAnimalObservation.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the animal obs data thats been sent from front end
        serializer = SaveOCCAnimalObservationSerializer(
            animal_obs_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_identification_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        identification_instance, created = OCCIdentification.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the identification data thats been sent from front end
        serializer = SaveOCCIdentificationSerializer(
            identification_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    # used for Occurrence external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def list_of_values(self, request, *args, **kwargs):
        """used for Occurrence external form"""
        land_form_list = []
        types = LandForm.objects.all()
        if types:
            for val in types:
                land_form_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        rock_type_list = []
        types = RockType.objects.all()
        if types:
            for val in types:
                rock_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_type_list = []
        types = SoilType.objects.all()
        if types:
            for val in types:
                soil_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_colour_list = []
        colours = SoilColour.objects.all()
        if colours:
            for val in colours:
                soil_colour_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_condition_list = []
        conditions = SoilCondition.objects.all()
        if conditions:
            for val in conditions:
                soil_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        drainage_list = []
        drainages = Drainage.objects.all()
        if drainages:
            for val in drainages:
                drainage_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        intensity_list = []
        intensities = Intensity.objects.all()
        if intensities:
            for val in intensities:
                intensity_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "land_form_list": land_form_list,
            "rock_type_list": rock_type_list,
            "soil_type_list": soil_type_list,
            "soil_colour_list": soil_colour_list,
            "soil_condition_list": soil_condition_list,
            "drainage_list": drainage_list,
            "intensity_list": intensity_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    # used for Occurrence Observation external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def observation_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence external form"""
        observation_method_list = []
        values = ObservationMethod.objects.all()
        if values:
            for val in values:
                observation_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_method_list = []
        values = PlantCountMethod.objects.all()
        if values:
            for val in values:
                plant_count_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_accuracy_list = []
        values = PlantCountAccuracy.objects.all()
        if values:
            for val in values:
                plant_count_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_condition_list = []
        values = PlantCondition.objects.all()
        if values:
            for val in values:
                plant_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        counted_subject_list = []
        values = CountedSubject.objects.all()
        if values:
            for val in values:
                counted_subject_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        primary_detection_method_list = []
        values = PrimaryDetectionMethod.objects.all()
        if values:
            for val in values:
                primary_detection_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        secondary_sign_list = []
        values = SecondarySign.objects.all()
        if values:
            for val in values:
                secondary_sign_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        reprod_maturity_list = []
        values = ReproductiveMaturity.objects.all()
        if values:
            for val in values:
                reprod_maturity_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        death_reason_list = []
        values = DeathReason.objects.all()
        if values:
            for val in values:
                death_reason_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        animal_health_list = []
        values = AnimalHealth.objects.all()
        if values:
            for val in values:
                animal_health_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        identification_certainty_list = []
        values = IdentificationCertainty.objects.all()
        if values:
            for val in values:
                identification_certainty_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_type_list = []
        values = SampleType.objects.all()
        if values:
            for val in values:
                sample_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_dest_list = []
        values = SampleDestination.objects.all()
        if values:
            for val in values:
                sample_dest_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        permit_type_list = []
        values = PermitType.objects.all()
        if values:
            for val in values:
                permit_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "observation_method_list": observation_method_list,
            "plant_count_method_list": plant_count_method_list,
            "plant_count_accuracy_list": plant_count_accuracy_list,
            "plant_condition_list": plant_condition_list,
            "counted_subject_list": counted_subject_list,
            "primary_detection_method_list": primary_detection_method_list,
            "secondary_sign_list": secondary_sign_list,
            "reprod_maturity_list": reprod_maturity_list,
            "death_reason_list": death_reason_list,
            "animal_health_list": animal_health_list,
            "identification_certainty_list": identification_certainty_list,
            "sample_type_list": sample_type_list,
            "sample_dest_list": sample_dest_list,
            "permit_type_list": permit_type_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="available-occurrence-reports-crs",
    )
    def available_occurrence_reports_crs(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        qs = self.get_queryset()
        crs = []

        id = request.GET.get("id", None)
        try:
            qs = qs.get(id=id)
        except Occurrence.DoesNotExist:
            logger.error(f"Occurrence with id {id} not found")
        else:
            ocr_geometries_ids = (
                qs.occurrence_reports.all()
                .values_list("ocr_geometry", flat=True)
                .distinct()
            )
            ocr_geometries = OccurrenceReportGeometry.objects.filter(
                id__in=ocr_geometries_ids
            ).exclude(**{"geometry": None})

            epsg_codes = [
                str(g.srid)
                for g in ocr_geometries.values_list("geometry", flat=True).distinct()
            ]
            # Add the srids of the original geometries to epsg_codes
            original_geometry_srids = [
                str(g.original_geometry_srid) for g in ocr_geometries
            ]
            epsg_codes += [g for g in original_geometry_srids if g.isnumeric()]
            epsg_codes = list(set(epsg_codes))
            crs = search_datums("", codes=epsg_codes)

        res_json = {
            "crs": crs,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class OccurrenceReportReferralViewSet(viewsets.GenericViewSet,mixins.RetrieveModelMixin):
    queryset = OccurrenceReportReferral.objects.all()
    serializer_class = OccurrenceReportReferralSerializer

    def get_serializer_class(self):
        if is_internal(self.request):
            return InternalOccurrenceReportReferralSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = super().get_queryset()
        if not is_internal(self.request):
            qs.filter(occurrence_report__submitter=self.request.user)
        return qs

    def is_authorised_to_refer(self):
        instance = self.get_object()
        user = self.request.user
        if not instance.occurrence_report.has_assessor_mode(user):
            raise serializers.ValidationError("User not authorised to manage Referrals for Occurrence Report")
        
    def is_authorised_to_referee(self):
        instance = self.get_object()
        user = self.request.user
        if not instance.referral == user:
            raise serializers.ValidationError("User is not the Referee for Occurrence Report Referral")

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def referral_list(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = self.get_queryset().filter(
            sent_by=instance.referral, occurrence_report=instance.occurrence_report
        )
        serializer = self.get_serializer(qs, many=True, context={"request": request})
        return Response(serializer.data)

    @detail_route(methods=["GET", "POST"], detail=True)
    def complete(self, request, *args, **kwargs):

        self.is_authorised_to_referee()

        instance = self.get_object()
        instance.complete(request)
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def remind(self, request, *args, **kwargs):

        self.is_authorised_to_refer()

        instance = self.get_object()
        instance.remind(request)
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def recall(self, request, *args, **kwargs):

        self.is_authorised_to_refer()

        instance = self.get_object()
        instance.recall(request)
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def resend(self, request, *args, **kwargs):

        self.is_authorised_to_refer()

        instance = self.get_object()
        instance.resend(request)
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def occurrence_report_referral_save(self, request, *args, **kwargs):

        self.is_authorised_to_referee()

        instance = self.get_object()
        request_data = request.data
        instance.referral_comment = request_data.get("referral_comment")
        instance.save()

        # Create a log entry for the occurrence report
        instance.occurrence_report.log_user_action(
            OccurrenceReportUserAction.COMMENT_REFERRAL.format(
                instance.id,
                instance.occurrence_report.occurrence_report_number,
                f"{instance.referral_as_email_user.get_full_name()}({instance.referral_as_email_user.email})",
            ),
            request,
        )
        return redirect(reverse("internal"))
